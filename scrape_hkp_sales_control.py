#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
import json
import time
import random
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import glob
from datetime import datetime, timezone, timedelta

try:
    import fitz
except ImportError:
    fitz = None

def parse_hkt_date(date_str):
    if not date_str or str(date_str).strip() in ['-', 'None', '']:
        return '-'
    s = str(date_str).strip()
    if 'T' in s and ('Z' in s or '+' in s or '-' in s):
        try:
            dt_utc = datetime.fromisoformat(s.replace('Z', '+00:00'))
            hkt = timezone(timedelta(hours=8))
            return dt_utc.astimezone(hkt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return s[:10]

# ==========================================
# 1. 简繁转换辅助函数与映射
# ==========================================
# 输出文件基础目录定位到当前脚本所在的价单文件夹
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 简繁转换对照文件仍引用“楼盘字典”下的原文件
zh2hans_path = "/Users/nb/google/Antigravity/工作/运营/楼盘字典/zh2hans.json"

if os.path.exists(zh2hans_path):
    with open(zh2hans_path, 'r', encoding='utf-8') as f:
        zh2hans_dict = json.load(f)
    # 只提取单字符映射，防止包含多字词汇导致的错位对齐映射
    char_map = {k: v for k, v in zh2hans_dict.items() if len(k) == 1 and len(v) == 1}
    TRADITIONAL_CHARS = "".join(char_map.keys())
    SIMPLIFIED_CHARS  = "".join(char_map.values())
else:
    # 备用简繁字符映射
    TRADITIONAL_CHARS = "港島九龍啟德灣仔堅尼地城筲箕灣紅磡鰂魚涌鴨脷洲黃竹坑東半山西半山中半山西九龍東九龍壽臣山山顶淺水灣深水灣跑马地大坑道司徒拔道渣甸山薄扶林香港仔上環中環土瓜灣小西湾西营盘铜锣湾天后鲗鱼涌康怡太古湾仔跑马地大坑渣甸山北角半山石塘咀坚尼地城摩星岭薄扶林香港仔黄竹坑深水湾浅水湾赤柱大潭石澳峯滙"
    SIMPLIFIED_CHARS  = "港岛九龙启德湾仔坚尼地城筲箕湾红磡鲗鱼涌鸭脷洲黄竹坑东半山西半山中半山西九龙东九龙寿臣山山顶浅水湾深水湾跑马地大坑道司徒拔道渣甸山薄扶林香港仔上环中环土瓜湾小西湾西营盘铜锣湾天后鲗鱼涌康怡太古湾仔跑马地大坑渣甸山北角半山石塘咀坚尼地城摩星岭薄扶林香港仔黄竹坑深水湾浅水湾赤柱大潭石澳峰汇"

TRANS_MAP = str.maketrans(TRADITIONAL_CHARS, SIMPLIFIED_CHARS)

def t2s(text):
    if not text:
        return ""
    if isinstance(text, dict):
        return {k: t2s(v) for k, v in text.items()}
    if isinstance(text, list):
        return [t2s(x) for x in text]
    return str(text).translate(TRANS_MAP)

def clean_name(name):
    """标准化名称比对，去除空格、特殊字符并转为小写简体"""
    if not name:
        return ""
    name = t2s(name).lower()
    name = re.sub(r'[\s\-\(\)\（\）\.\,\，\。]', '', name)
    return name

def get_safe_filename(project_dir, pname):
    """生成去除特殊字符的安全 Excel 文件名"""
    safe_pname = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '', pname)
    return os.path.join(project_dir, f"{safe_pname}_销控明细表.xlsx")

def normalize_bname(name):
    """标准化楼栋名称以进行比对映射"""
    if not name:
        return ""
    name = t2s(name).strip()
    name = re.sub(r'^第', '', name)
    return name

# ==========================================
# 2. 状态映射与颜色配置
# ==========================================
STATUS_MAP = {
    'pending': '待售',
    'priced': '已定价未售',
    'sale': '在售',
    'sold': '已售',
    'stopped': '暂停销售'
}

COLORS = {
    'sale': {'bg': 'A9F5A9', 'fg': '004D00'},
    'sold': {'bg': 'F5A9A9', 'fg': '660000'},
    'priced': {'bg': 'A9D0F5', 'fg': '002060'},
    'stopped': {'bg': 'FFC000', 'fg': '000000'},
    'pending': {'bg': 'FFFFFF', 'fg': '7F7F7F'}
}

# 销控单元格颜色填充配置 (Openpyxl Fills)
FILLS = {k: PatternFill(start_color=v['bg'], end_color=v['bg'], fill_type='solid') for k, v in COLORS.items()}

# ==========================================
# 3. 网络请求基础配置与 Token 获取
# ==========================================
USER_AGENTS = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:125.0) Gecko/20100101 Firefox/125.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 17_4_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Mobile/15E148 Safari/604.1'
]

HEADERS = {
    'User-Agent': USER_AGENTS[0]
}

def fetch_user_token():
    """从主列表页的 __NEXT_DATA__ 中解析最新的 Bearer Token"""
    url = "https://www.hkp.com.hk/zh-hk/list/new-property/"
    print("正在从主页获取最新 Token...")
    try:
        headers = {'User-Agent': random.choice(USER_AGENTS)}
        r = requests.get(url, headers=headers, timeout=15)
        r.raise_for_status()
        match = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>', r.text)
        if match:
            data = json.loads(match.group(1))
            token = data.get('props', {}).get('pageProps', {}).get('userToken')
            if token:
                print(f"Token 获取成功! 长度: {len(token)}")
                return token
        raise ValueError("__NEXT_DATA__ 中未找到 userToken")
    except Exception as e:
        print(f"错误: 无法获取 Token: {e}", file=sys.stderr)
        sys.exit(1)

def clean_pct(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('%', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0

def format_discount_rate_str(rate):
    if rate is None or rate <= 0:
        return '0%'
    val = round(rate * 100, 2)
    if val.is_integer():
        return f"{int(val)}%"
    val_str = f"{val:.2f}".rstrip('0').rstrip('.')
    return f"{val_str}%"

def calculate_unit_discount(u, project_payments, fallback_rate, fallback_title):
    unit_payment = u.get('payment')
    if not unit_payment or not project_payments:
        return fallback_rate, fallback_title
        
    pay_dict = {p['id']: p for p in project_payments if 'id' in p}
    max_rate = 0.0
    best_title = fallback_title
    
    for u_pay in unit_payment:
        pay_id = u_pay.get('id')
        if not pay_id or pay_id not in pay_dict:
            continue
            
        allowed_discount_ids = set(u_pay.get('discount_group_ids', []))
        proj_pay = pay_dict[pay_id]
        
        pct = clean_pct(proj_pay.get('percentage'))
        
        # 计算 bonuses
        bonuses_pct = 0.0
        for b in proj_pay.get('bonuses', []):
            if b.get('id') in allowed_discount_ids:
                bonuses_pct += clean_pct(b.get('percentage'))
                
        # 计算 sub_bonuses
        sub_bonuses_pct = 0.0
        for sb in proj_pay.get('sub_bonuses', []):
            if sb.get('id') in allowed_discount_ids:
                sub_bonuses_pct += clean_pct(sb.get('percentage'))
                
        total_pct = pct + bonuses_pct + sub_bonuses_pct
        if total_pct < 0:
            abs_rate = abs(total_pct) / 100.0
            if abs_rate > max_rate:
                max_rate = abs_rate
                best_title = proj_pay.get('title') or fallback_title
                
    if max_rate > 0.0:
        return max_rate, best_title
    return fallback_rate, fallback_title

def calculate_scale2_stamp_duty(price):
    if price <= 3000000:
        return 100
    elif price <= 3528000:
        return 100 + (price - 3000000) * 0.10
    elif price <= 4500000:
        return price * 0.015
    elif price <= 4935000:
        return 67500 + (price - 4500000) * 0.10
    elif price <= 6000000:
        return price * 0.0225
    elif price <= 6642900:
        return 135000 + (price - 6000000) * 0.10
    elif price <= 9000000:
        return price * 0.03
    elif price <= 10080000:
        return 270000 + (price - 9000000) * 0.10
    elif price <= 20000000:
        return price * 0.0375
    elif price <= 21739000:
        return 750000 + (price - 20000000) * 0.10
    else:
        return price * 0.0425

def is_floor_in_chunk(floor, chunk):
    """高级楼层匹配，支持繁简体 '楼/樓', 区间范围如 '2樓至39樓', '2/F至39/F' 及 '所有樓層'"""
    if not chunk:
        return False
    c_clean = chunk.replace(" ", "").replace("楼", "樓").upper()
    floor_str = str(floor).strip()
    
    range_pattern = r'(\d+)(?:樓|/F)(?:至|至|-|–|—)(\d+)(?:樓|/F)'
    ranges = re.findall(range_pattern, c_clean)
    if ranges:
        for r_min, r_max in ranges:
            try:
                if int(r_min) <= int(floor) <= int(r_max):
                    return True
            except ValueError:
                pass
                
    if "所有樓" in c_clean or "各樓" in c_clean or "所有單位" in c_clean:
        return True
        
    num_matches = re.findall(r'\d+', c_clean)
    if floor_str in num_matches:
        return True
        
    return False

def check_stamp_duty_eligibility(pdf_path, block, floor, flat):
    if not fitz:
        return False
    try:
        doc = fitz.open(pdf_path)
        for page_idx in range(len(doc)):
            page = doc.load_page(page_idx)
            text = page.get_text("text")
            if "代繳從價印花稅" not in text and "Ad Valorem Stamp Duty" not in text:
                continue
                
            blocks = page.get_text("blocks")
            # 依据 y 坐标将文本块分组（同水平行内的文本块 y 差值不超过 6 像素）
            sorted_by_y = sorted(blocks, key=lambda b: b[1])
            rows = []
            current_row = []
            current_y = None
            for b in sorted_by_y:
                if current_y is None:
                    current_y = b[1]
                    current_row.append(b)
                elif b[1] - current_y <= 6:
                    current_row.append(b)
                else:
                    rows.append(current_row)
                    current_row = [b]
                    current_y = b[1]
            if current_row:
                rows.append(current_row)
                
            block_pattern = f"第{block}座"
            
            # 在每行中精确匹配大厦与单元，并检查对应的层数是否在同一行中
            for row_blocks in rows:
                match_block_flat = False
                for b in row_blocks:
                    b_text = re.sub(r'\s+', '', b[4])
                    if block_pattern in b_text and flat in b_text:
                        match_block_flat = True
                        break
                        
                if match_block_flat:
                    row_text = "\n".join([b[4] for b in row_blocks])
                    if is_floor_in_chunk(floor, row_text):
                        return True
    except Exception:
        pass
    return False

def validate_henderson_pdf_structure(pdf_path):
    """验证 PDF 前几页文字中是否同时包含恒基核心的五列列头字样，避免错套格式"""
    if not fitz:
        return False
    try:
        doc = fitz.open(pdf_path)
        for i in range(min(3, len(doc))):
            page = doc.load_page(i)
            text = page.get_text("text")
            text_clean = re.sub(r'\s+', '', text)
            
            # 标准恒基大表中，五列头部包含大厦、楼层、单位、实用面积、售价
            has_block = "大廈" in text_clean or "大厦" in text_clean or "Block" in text_clean
            has_floor = "樓層" in text_clean or "楼层" in text_clean or "Floor" in text_clean
            has_flat = "單位" in text_clean or "单位" in text_clean or "Flat" in text_clean
            has_area = "實用面積" in text_clean or "实用面积" in text_clean or "Saleable" in text_clean
            has_price = "售價" in text_clean or "售价" in text_clean or "Price" in text_clean
            
            if has_block and has_floor and has_flat and has_area and has_price:
                return True
    except Exception:
        pass
    return False

def get_henderson_pdf_data(detail_data, api_headers, project_dir, fallback_rate, pname):
    pdf_db = {}
    if not fitz:
        return pdf_db
        
    attachments = detail_data.get('attachment', [])
    pdf_dir = os.path.join(project_dir, "pdfs")
    os.makedirs(pdf_dir, exist_ok=True)
    
    price_lists_pdf = {}
    for a in attachments:
        name = a.get('name', '')
        type_ = a.get('type', '')
        pl_num = a.get('pl_num', '')
        path = a.get('path', '')
        if pl_num and path and ('pl' in str(type_).lower() or '價單' in name):
            price_lists_pdf[pl_num] = path
            
    # 按照最新到最旧进行价单版本排序 (例如 6B > 5L > 4M > 4L > 3O > 2Q > 1N)
    def get_pl_sort_key(pl):
        m = re.match(r'(\d+)([A-Za-z]*)', str(pl))
        if m:
            num = int(m.group(1))
            chars = m.group(2)
            char_val = ord(chars[0]) if chars else 0
            return (num, char_val)
        return (0, 0)
        
    sorted_pls = sorted(price_lists_pdf.items(), key=lambda x: get_pl_sort_key(x[0]), reverse=True)
    


    # 下载价单并解析房源
    for pl_num, url in sorted_pls:
        cache_path = os.path.join(pdf_dir, f"cache_{pl_num}.json")
        # 1. 优先读取本地 JSON 缓存，避免重复扫描 PDF
        if os.path.exists(cache_path):
            try:
                with open(cache_path, 'r', encoding='utf-8') as cf:
                    cached_data = json.load(cf)
                for k_str, pdf_info in cached_data.items():
                    parts = k_str.split('_')
                    if len(parts) == 3:
                        key = (parts[0], parts[1], parts[2])
                        if key not in pdf_db:
                            pdf_db[key] = pdf_info
                print(f"    -> [缓存命中] 成功从本地 JSON 缓存加载 價單 {pl_num} 號 的解析数据。")
                continue
            except Exception as cache_err:
                print(f"    -> 提示: 读取價單 {pl_num} 號 的本地 JSON 缓存异常 ({cache_err})，将重新扫描 PDF。")
        
        pdf_path = os.path.join(pdf_dir, f"price_list_{pl_num}.pdf")
        if not os.path.exists(pdf_path):
            try:
                # 隐藏 API 头，用一般浏览头的模拟
                headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
                r = requests.get(url, headers=headers, timeout=30)
                if r.status_code == 200:
                    with open(pdf_path, 'wb') as f:
                        f.write(r.content)
            except Exception:
                continue
                
        # 2. 物理排版安全判断，防止错套非恒基模板的联合开发项目 PDF
        if not validate_henderson_pdf_structure(pdf_path):
            print(f"    -> [安全拦截] 價單 {pl_num} 號 物理排版不符合恒基 5 列标准列头，已跳过该价单的物理折算。")
            continue
            
        current_pl_db = {}
        try:
            doc = fitz.open(pdf_path)
            for page_idx in range(len(doc)):
                page = doc.load_page(page_idx)
                tables = page.find_tables()
                if not tables:
                    continue
                for table in tables:
                    df = table.to_pandas()
                    if len(df.columns) < 5:
                        continue
                    for row_idx, row in df.iterrows():
                        col0 = str(row.iloc[0]).replace("\n", "").replace(" ", "").strip()
                        col1 = str(row.iloc[1]).replace("\n", "").replace(" ", "").strip()
                        col2 = str(row.iloc[2]).replace("\n", "").replace(" ", "").strip()
                        col3 = str(row.iloc[3]).replace("\n", "").replace(" ", "").strip()
                        col4 = str(row.iloc[4]).replace("\n", "").replace(" ", "").strip()
                        
                        if "大廈" in col0 or "Block" in col0 or "樓層" in col1 or "Floor" in col1:
                            continue
                        if "座" not in col0:
                            continue
                        floor_clean = re.sub(r'[^0-9]', '', col1)
                        if not floor_clean:
                            continue
                        flat_clean = re.sub(r'[^A-Z0-9]', '', col2.upper())
                        if not flat_clean:
                            continue
                        price_str = col4.replace(",", "").replace("$", "").replace("元", "")
                        try:
                            price = int(price_str)
                        except ValueError:
                            continue
                        
                        area_match = re.search(r'\((\d+)\)', col3)
                        area = 0
                        if area_match:
                            area = int(area_match.group(1))
                        else:
                            area_nums = re.findall(r'\d+', col3.replace(",", ""))
                            if area_nums:
                                area = int(area_nums[-1])
                                
                        b_match = re.search(r'(\d+)([A-Za-z]?)', col0)
                        bname = b_match.group(1) + b_match.group(2) if b_match else col0
                        
                        key = (bname, floor_clean, flat_clean)
                        
                        # 无论如何，都为当前价单的独立缓存构造数据
                        direct_rate = fallback_rate
                        if "The Henley" in pname:
                            if pname == "The Henley III" and pl_num == "4L":
                                if area <= 420:
                                    direct_rate = 0.125
                                else:
                                    direct_rate = 0.05
                            else:
                                if area <= 420:
                                    direct_rate = 0.09
                                else:
                                    direct_rate = 0.035
                        elif "Henley Park" in pname:
                            direct_rate = 0.085
                                
                        contract_price = int(price * (1 - direct_rate))
                        contract_price = (contract_price // 100) * 100
                        
                        has_sd = check_stamp_duty_eligibility(pdf_path, bname, floor_clean, flat_clean)
                        sd_amount = 0
                        if has_sd:
                            sd_amount = calculate_scale2_stamp_duty(contract_price)
                            
                        final_price = contract_price - sd_amount
                        final_price = int((final_price // 100) * 100)
                        
                        total_benefit_pct = (price - final_price) / price
                        
                        payment_desc = f"現金付款計劃 (減{direct_rate * 100:.1f}%)"
                        if has_sd:
                            payment_desc += " + 代繳從價印花稅"
                        
                        pdf_info = {
                            "original_price": price,
                            "discount_price": final_price,
                            "discount_percent_str": f"{total_benefit_pct * 100:.2f}%",
                            "payment_method": payment_desc,
                            "price_list": pl_num
                        }
                        
                        current_pl_db[f"{bname}_{floor_clean}_{flat_clean}"] = pdf_info
                        if key not in pdf_db:
                            pdf_db[key] = pdf_info
                            
            # 保存到本地 JSON 缓存文件
            if current_pl_db:
                with open(cache_path, 'w', encoding='utf-8') as cf:
                    json.dump(current_pl_db, cf, ensure_ascii=False, indent=2)
                print(f"    -> [缓存更新] 成功为 價單 {pl_num} 號 生成并写出本地 JSON 缓存文件。")
        except Exception as scan_err:
            print(f"    -> 警告: 物理扫描 價單 {pl_num} 號 发生异常 ({scan_err})")
            continue
            
    # 针对 Henley Park 早期高座大厦（如 1B 座的 A, D, F 大户型）价单未被 HKP 挂载的现状，在内存中动态注入 8.5% 付款补偿模板
    if pname == "Henley Park":
        for floor_num in range(2, 35):
            floor_str = str(floor_num)
            for flat_let in ["A", "D", "F"]:
                # 兼容 1座(1A)/2座(1B) 和 1A/1B 两种不同物理命名表现
                for b_prefix in ["1", "2", "1A", "1B"]:
                    key = (b_prefix, floor_str, flat_let)
                    pdf_db[key] = {
                        "is_compensation": True,
                        "direct_rate": 0.085
                    }
                
    return pdf_db

def extract_max_discount_info(detail_data):
    """从新盘详情中提取最优惠的付款办法和总折扣率"""
    if 'result' in detail_data and isinstance(detail_data['result'], dict):
        detail_data = detail_data['result']
        
    payment = detail_data.get('payment', [])
    if not payment:
        return 0.0, "-", []
        
    max_rate = 0.0
    best_plan_title = "-"
    best_plan_bonuses = []
    
    for pay in payment:
        # 基础百分比 (通常为负数，如 -6，代表减 6%)
        pct = pay.get('percentage') or 0.0
        pct = clean_pct(pct)
            
        # 计算 bonuses 折扣
        bonuses = pay.get('bonuses', [])
        bonuses_pct = 0.0
        for b in bonuses:
            bp = b.get('percentage') or 0.0
            bonuses_pct += clean_pct(bp)
                 
        # 计算 sub_bonuses 折扣
        sub_bonuses = pay.get('sub_bonuses', [])
        sub_bonuses_pct = 0.0
        for sb in sub_bonuses:
            sbp = sb.get('percentage') or 0.0
            sub_bonuses_pct += clean_pct(sbp)
                 
        # 总折扣率之和 (注意是累加负值)
        total_pct = pct + bonuses_pct + sub_bonuses_pct
        if total_pct < 0:
            abs_rate = abs(total_pct) / 100.0
            if abs_rate > max_rate:
                max_rate = abs_rate
                best_plan_title = pay.get('title') or "-"
                best_plan_bonuses = [b.get('title') for b in bonuses if b.get('title')]
                
    return max_rate, best_plan_title, best_plan_bonuses

# ==========================================
# 4. 全局目录扫描与映射
# ==========================================
def scan_existing_folders():
    """扫描已有文件夹（包含价单和原楼盘字典目录），以防止命名拼写差异导致创建重复文件夹"""
    mapping = {}
    folders_to_scan = [BASE_DIR, "/Users/nb/google/Antigravity/工作/运营/楼盘字典"]
    for folder in folders_to_scan:
        if not os.path.exists(folder):
            continue
        for d in os.listdir(folder):
            path = os.path.join(folder, d)
            if os.path.isdir(path):
                parts = d.split('-')
                if len(parts) >= 3:
                    proj_name = parts[-1]
                    mapping[clean_name(proj_name)] = d
    return mapping

def load_existing_project_data(pname, pid, region, district, developer, existing_folders, global_units):
    """
    如果抓取失败，尝试从已有的本地 Excel 文件中加载数据并合并到 global_units 中。
    """
    clean_proj_name = clean_name(pname)
    if clean_proj_name not in existing_folders:
        return False
        
    folder_name = existing_folders[clean_proj_name]
    project_dir = os.path.join(BASE_DIR, folder_name)
    excel_filename = get_safe_filename(project_dir, pname)
    
    if not os.path.exists(excel_filename):
        return False
        
    print(f"  [历史沿用] 正在从本地历史数据加载: {excel_filename}")
    try:
        # 使用 read_only=True 加速读取
        wb = openpyxl.load_workbook(excel_filename, read_only=True, data_only=True)
        if "销控汇总明细" in wb.sheetnames:
            ws = wb["销控汇总明细"]
            count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                # 检查是否是空行 (第一列"楼栋"为空则视为结束)
                if not row or row[0] is None:
                    break
                
                global_units.append({
                    '项目ID': pid,
                    '项目名称': pname,
                    '区域': region,
                    '商圈': district,
                    '开发商': developer,
                    '楼栋名称': row[0],
                    '楼层': row[1],
                    '房号': row[2],
                    '户型': row[3],
                    '实用面积': row[4],
                    '销控状态': row[5],
                    '成交日期': row[6],
                    '总价': row[7],
                    '呎价': row[8],
                    '最高折扣': row[9] if len(row) > 9 else '-',
                    '折实总价': row[10] if len(row) > 10 else '暂无',
                    '折实呎价': row[11] if len(row) > 11 else '暂无',
                    '付款办法': row[12] if len(row) > 12 else '-',
                    '是否招标': row[13] if len(row) > 13 else '否'
                })
                count += 1
            print(f"  [成功] 从本地历史数据恢复了 {count} 条单位数据。")
            wb.close()
            return True
        wb.close()
    except Exception as e:
        print(f"  [警告] 读取本地历史 Excel 失败: {e}")
    return False

def is_project_sold_out(excel_filename):
    """
    检查已有的本地 Excel 文件是否显示该项目已售罄 (100%)，且所有已售单位均有成交日期。
    若存在缺失成交日期的已售单位，则返回 False 以重新抓取并反填日期。
    """
    if not os.path.exists(excel_filename):
        return False
    try:
        wb = openpyxl.load_workbook(excel_filename, read_only=True, data_only=True)
        if "销控汇总明细" in wb.sheetnames:
            ws = wb["销控汇总明细"]
            total = 0
            sold = 0
            missing_date = 0
            for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
                if not row or row[0] is None:
                    break
                total += 1
                status_str = str(row[5]).strip() if len(row) >= 6 and row[5] is not None else ""
                tx_date = str(row[6]).strip() if len(row) >= 7 and row[6] is not None else "-"
                if status_str == '已售':
                    sold += 1
                    if tx_date == '-' or not tx_date:
                        missing_date += 1
            wb.close()
            return total > 0 and sold == total and missing_date == 0
        wb.close()
    except Exception as e:
        print(f"  [警告] 读取本地 Excel 判断是否售罄时失败: {e}")
    return False


def extract_floor_number(floor_str):
    """从楼层字符串提取用于数字排序的数值"""
    if not floor_str:
        return -99
    floor_str = str(floor_str).upper().strip()
    match = re.search(r'(\d+)', floor_str)
    if match:
        val = int(match.group(1))
        if 'R' in floor_str:
            val += 0.5
        return val
    if 'G' in floor_str or 'UG' in floor_str:
        return 0
    if 'B' in floor_str:
        return -1
    return -99

# ==========================================
# 5. Excel 写入辅助逻辑 (使用 Openpyxl)
# ==========================================
def write_project_excel(filepath, project_name, buildings_data):
    """
    为单个项目写入 Excel 销控表，包含：
    - Tab 1: "销控汇总明细" (扁平表形式，包含所有楼栋的所有单位)
    - 之后的 Tabs: 每个楼栋单独一页，采用 "销控网格图 (Grid)" 形式展示
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------
    # Tab 1: 销控汇总明细
    # ----------------------------------------
    ws_detail = wb.active
    ws_detail.title = "销控汇总明细"
    
    font_title = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    font_data = Font(name='Microsoft YaHei', size=10)
    align_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 写标题
    ws_detail.merge_cells("A1:N1")
    ws_detail['A1'] = f"{project_name} - 一手销控汇总明细"
    ws_detail['A1'].font = font_title
    ws_detail['A1'].alignment = align_center
    ws_detail.row_dimensions[1].height = 40
    
    # 写表头 (增加了最高折扣、折实价、付款办法等字段)
    headers = [
        "楼栋", "楼层", "房号", "户型", "实用面积 (平方呎)", "销控状态", "成交日期", 
        "总价 (港币)", "实用呎价 (港币/呎)", "最高折扣", "折实总价 (港币)", "折实呎价 (港币/呎)", "付款办法", "是否招标"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=2, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_thin
    ws_detail.row_dimensions[2].height = 25

    # 填充明细数据
    row_idx = 3
    for bname, units in buildings_data.items():
        sorted_units = sorted(units, key=lambda x: (extract_floor_number(x['floor']), x['flat']), reverse=True)
        for u in sorted_units:
            ws_detail.cell(row=row_idx, column=1, value=bname)
            ws_detail.cell(row=row_idx, column=2, value=u['floor'])
            ws_detail.cell(row=row_idx, column=3, value=u['flat'])
            ws_detail.cell(row=row_idx, column=4, value=u['room_layout'])
            ws_detail.cell(row=row_idx, column=5, value=u['net_area'] if u['net_area'] > 0 else '暂无')
            ws_detail.cell(row=row_idx, column=6, value=u['status'])
            ws_detail.cell(row=row_idx, column=7, value=u['sold_date'])
            
            is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
            
            # 如果是招标且未售，总价显示招标
            if is_tender and u['status_raw'] != 'sold':
                ws_detail.cell(row=row_idx, column=8, value='招标单位')
                ws_detail.cell(row=row_idx, column=9, value='-')
                ws_detail.cell(row=row_idx, column=10, value='-')
                ws_detail.cell(row=row_idx, column=11, value='招标单位')
                ws_detail.cell(row=row_idx, column=12, value='-')
                ws_detail.cell(row=row_idx, column=13, value='-')
            else:
                ws_detail.cell(row=row_idx, column=8, value=u['price'] if u['price'] else '暂无')
                ws_detail.cell(row=row_idx, column=9, value=u['price_per_sq_ft'] if u['price_per_sq_ft'] else '暂无')
                ws_detail.cell(row=row_idx, column=10, value=u.get('discount_percent_str', '-'))
                ws_detail.cell(row=row_idx, column=11, value=u.get('discount_price') if u.get('discount_price') else '暂无')
                ws_detail.cell(row=row_idx, column=12, value=u.get('discount_price_per_sq_ft') if u.get('discount_price_per_sq_ft') else '暂无')
                ws_detail.cell(row=row_idx, column=13, value=u.get('payment_method', '-'))
                
            ws_detail.cell(row=row_idx, column=14, value='是' if is_tender else '否')
            
            for col in range(1, 15):
                c = ws_detail.cell(row=row_idx, column=col)
                c.font = font_data
                c.alignment = align_center
                c.border = border_thin
                
                # 数字格式化 (包含折实总价 11 和折实呎价 12)
                if col in [8, 9, 11, 12] and isinstance(c.value, (int, float)):
                    c.number_format = '$#,##0'
                elif col == 5 and isinstance(c.value, (int, float)):
                    c.number_format = '#,##0'

            row_idx += 1

    # 自动列宽
    for col in ws_detail.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_detail.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 设置打印自适应：单页宽度，高度自动（纵向打印）
    ws_detail.sheet_properties.pageSetUpPr.fitToPage = True
    ws_detail.page_setup.fitToWidth = 1
    ws_detail.page_setup.fitToHeight = 0
    ws_detail.page_setup.orientation = ws_detail.ORIENTATION_PORTRAIT
    ws_detail.page_setup.paperSize = ws_detail.PAPERSIZE_A4

    # 设置窄页边距 (0.25 英寸)
    ws_detail.page_margins.left = 0.25
    ws_detail.page_margins.right = 0.25
    ws_detail.page_margins.top = 0.25
    ws_detail.page_margins.bottom = 0.25
    ws_detail.page_margins.header = 0.1
    ws_detail.page_margins.footer = 0.1

    # ----------------------------------------
    # Tab 2+: 每个楼栋生成可视化销控网格 (Grid)
    # ----------------------------------------
    multi_buildings = {}
    villa_buildings = {}
    
    if buildings_data:
        for bname, units in buildings_data.items():
            # 判断是否为独栋 (如果名称含'屋'、属于 HOUSE 类型，或不含'座'且单位数 <= 2)
            has_house_type = any(u.get('unit_type') == 'HOUSE' or 'HOUSE' in str(u.get('bldg_type', '')) for u in units)
            is_villa = '屋' in bname or has_house_type or ('座' not in bname and len(units) <= 2)
            if is_villa:
                villa_buildings[bname] = units
            else:
                multi_buildings[bname] = units

    # 1. 多个单位 of 楼栋，每个楼栋放在单独的工作表
    for bname, units in multi_buildings.items():
        _write_building_grid(wb, bname, units, font_data, font_header, font_title, align_center, border_thin, fill_header)

    # 2. 独栋楼，集合放在一个工作表
    if villa_buildings:
        _write_villa_grid(wb, project_name, villa_buildings, font_data, align_center, border_thin)

    wb.save(filepath)

def update_global_excel_with_retry(filepath, new_units):
    """
    更新全局汇总 Excel，替换并追加重试项目的最新数据（采用极速读取与重写机制）。
    """
    if not os.path.exists(filepath):
        write_global_excel(filepath, new_units)
        return
        
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        replaced_pids = set(u['项目ID'] for u in new_units)
        
        # 1. 保留未被替换的行数据
        keep_rows = []
        max_r = ws.max_row
        
        # 从第3行开始读取
        for r in range(3, max_r + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 21)]
            # 检查空行
            if not any(row_vals):
                continue
            # 判断项目ID是否在需要替换的集合中
            cell_pid = row_vals[1] # 第二列是项目ID
            if cell_pid not in replaced_pids:
                keep_rows.append(row_vals)
                
        wb.close()
        
        # 2. 重新组合所有数据 (保留的行 + 新重试的行)
        combined_rows = []
        # 先放入原先保留的行
        for idx, row in enumerate(keep_rows, 1):
            row[0] = idx # 重新生成序号
            # 如果读出的历史数据少于 20 列，对其进行扩充 padding
            if len(row) < 20:
                row = row[:15] + ['-', '暂无', '暂无', '-'] + row[15:]
            combined_rows.append(row)
            
        # 再追加新重试的行
        start_idx = len(combined_rows) + 1
        for idx, d in enumerate(new_units, start_idx):
            combined_rows.append([
                idx,
                d['项目ID'],
                d['项目名称'],
                d['区域'],
                d['商圈'],
                d['开发商'],
                d['楼栋名称'],
                d['楼层'],
                d['房号'],
                d['户型'],
                d['实用面积'],
                d['销控状态'],
                d['成交日期'],
                d['总价'],
                d['呎价'],
                d.get('最高折扣', '-'),
                d.get('折实总价', '暂无'),
                d.get('折实呎价', '暂无'),
                d.get('付款办法', '-'),
                d['是否招标']
            ])
            
        # 3. 完全重建全局汇总表以获得极高写入速度
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "一手新盘销控汇总"
        
        # 写入表头 (和 write_global_excel 保持一致)
        headers = [
            "序号", "项目ID", "项目名称", "区域", "商圈", "开发商", 
            "楼栋名称", "楼层", "房号", "户型", "实用面积 (平方呎)", 
            "销控状态", "成交日期", "单位总价 (港币)", "实用呎价 (港币/呎)", 
            "最高折扣", "折实总价 (港币)", "折实呎价 (港币/呎)", "付款办法", "是否招标"
        ]
        
        # 标题行
        font_title = Font(name='Microsoft YaHei', size=14, bold=True, color='FFFFFF')
        fill_title = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        
        new_ws.merge_cells("A1:T1")
        title_cell = new_ws.cell(row=1, column=1, value="香港一手新盘销控汇总分析表")
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = align_center
        new_ws.row_dimensions[1].height = 40
        
        # 表头行
        font_header = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        
        new_ws.row_dimensions[2].height = 28
        for col_idx, text in enumerate(headers, 1):
            c = new_ws.cell(row=2, column=col_idx, value=text)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center
            
        # 写入所有行数据
        font_data = Font(name='Microsoft YaHei', size=10)
        border_thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        for r_idx, row_data in enumerate(combined_rows, 3):
            new_ws.row_dimensions[r_idx].height = 20
            for c_idx, val in enumerate(row_data, 1):
                c = new_ws.cell(row=r_idx, column=c_idx, value=val)
                c.font = font_data
                c.alignment = align_center
                c.border = border_thin
                # 格式化数值 (包含折实总价 17 和折实呎价 18)
                if c_idx in [14, 15, 17, 18] and isinstance(c.value, (int, float)):
                    c.number_format = '$#,##0'
                elif c_idx == 11 and isinstance(c.value, (int, float)):
                    c.number_format = '#,##0'
                    
        # 列宽自适应
        for col in new_ws.columns:
            max_len = 0
            for cell in col:
                if cell.row == 1:
                    continue
                val_str = str(cell.value or '')
                # 计算近似字符长度
                length = sum(2 if ord(char) > 256 else 1 for char in val_str)
                if length > max_len:
                    max_len = length
            col_letter = get_column_letter(col[0].column)
            new_ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        # 设置页面自适应打印
        new_ws.sheet_properties.pageSetUpPr.fitToPage = True
        new_ws.page_setup.orientation = new_ws.ORIENTATION_PORTRAIT
        new_ws.page_setup.paperSize = new_ws.PAPERSIZE_A4
        new_ws.page_setup.fitToWidth = 1
        new_ws.page_setup.fitToHeight = 0
        
        new_ws.page_margins.left = 0.25
        new_ws.page_margins.right = 0.25
        new_ws.page_margins.top = 0.25
        new_ws.page_margins.bottom = 0.25
        new_ws.page_margins.header = 0.1
        new_ws.page_margins.footer = 0.1
        
        new_wb.save(filepath)
        print(f"  [成功] 采用极速重写机制，已更新并追加 {len(new_units)} 条新重试数据到全局汇总表中。")
    except Exception as e:
        print(f"  [错误] 极速重载更新全局汇总 Excel 失败: {e}")

def write_global_excel(filepath, data_list):
    """写入全局汇总 Excel 数据（采用高效 append 与预设列宽）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一手新盘销控汇总"
    
    font_title = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 写标题
    ws.merge_cells("A1:T1")
    ws['A1'] = "香港港岛、九龙一手新盘单位销控明细汇总表"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 40

    # 写表头
    headers = [
        "序号", "项目ID", "项目名称", "区域", "商圈", "开发商", "楼栋名称", "楼层", "房号", "户型", 
        "实用面积 (平方呎)", "销控状态", "成交日期", "单位总价 (港币)", "实用呎价 (港币/呎)", "最高折扣", "折实总价 (港币)", "折实呎价 (港币/呎)", "付款办法", "是否招标"
    ]
    ws.append(headers)
    ws.row_dimensions[2].height = 25
    for col_idx in range(1, 21):
        c = ws.cell(row=2, column=col_idx)
        c.font = font_header
        c.alignment = align_center
        c.fill = fill_header
        c.border = border_thin

    # 高效批量 append 写入行
    for idx, d in enumerate(data_list, 1):
        row_vals = [
            idx, d['项目ID'], d['项目名称'], d['区域'], d['商圈'], d['开发商'],
            d['楼栋名称'], d['楼层'], d['房号'], d['户型'], d['实用面积'],
            d['销控状态'], d['成交日期'], d['总价'], d['呎价'],
            d.get('最高折扣', '-'), d.get('折实总价', '暂无'), d.get('折实呎价', '暂无'),
            d.get('付款办法', '-'), d['是否招标']
        ]
        ws.append(row_vals)

    # 预设最佳列宽（避免扫描数万行 * 20列 计120万个单元格，提速数十倍）
    widths = [6, 12, 18, 8, 12, 12, 10, 8, 8, 8, 14, 10, 12, 14, 14, 10, 14, 14, 16, 8]
    for c_idx, w in enumerate(widths, 1):
        col_letter = get_column_letter(c_idx)
        ws.column_dimensions[col_letter].width = w

    wb.save(filepath)

# ==========================================
# 5.4 历史成交数据沉淀与留存逻辑 (DB + Excel)
# ==========================================
def save_transactions_to_db(global_units):
    import sqlite3
    from datetime import datetime
    
    db_path = os.path.join(BASE_DIR, "成交历史数据库.db")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 1. 创建表结构
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS sold_history (
        region TEXT,             -- 区域
        district TEXT,           -- 商圈
        project_name TEXT,       -- 项目名称
        building_name TEXT,      -- 楼栋
        floor TEXT,              -- 楼层
        flat TEXT,               -- 房号
        layout TEXT,             -- 户型
        area TEXT,               -- 实用面积
        sold_date TEXT,          -- 成交日期 (YYYY-MM-DD)
        price TEXT,              -- 成交总价 / 登记价格 (港币)
        unit_price TEXT,         -- 成交呎价 (港币/呎)
        discount TEXT,           -- 最高折扣 (如果有)
        disc_price TEXT,         -- 折实价
        disc_unit_price TEXT,    -- 折实呎价
        payment TEXT,            -- 付款办法
        is_tender TEXT,          -- 是否招标 ('是'/'否')
        captured_at TEXT,        -- 捕获登记时间 (YYYY-MM-DD HH:MM:SS)
        PRIMARY KEY (project_name, building_name, floor, flat, sold_date)
    );
    """)
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_sold_history_unit_date ON sold_history (project_name, building_name, floor, flat, sold_date);")
    conn.commit()
    
    # 2. 批量构建写入数据（使用 executemany 极大提升性能）
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    records = []
    
    for u in global_units:
        if u.get('销控状态') == '已售' and u.get('成交日期') and u.get('成交日期') != '-':
            records.append((
                u.get('区域', ''),
                u.get('商圈', ''),
                u.get('项目名称', ''),
                u.get('楼栋名称', ''),
                u.get('楼层', ''),
                u.get('房号', ''),
                u.get('户型', ''),
                str(u.get('实用面积', '')),
                u.get('成交日期', ''),
                str(u.get('总价', '')),
                str(u.get('呎价', '')),
                u.get('最高折扣', ''),
                str(u.get('折实总价', '')),
                str(u.get('折实呎价', '')),
                u.get('付款办法', ''),
                u.get('是否招标', '否'),
                now_str
            ))
            
    if records:
        cursor.executemany("""
        INSERT OR IGNORE INTO sold_history (
            region, district, project_name, building_name, floor, flat,
            layout, area, sold_date, price, unit_price, discount,
            disc_price, disc_unit_price, payment, is_tender, captured_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, records)
        conn.commit()
        
    print(f"\n[成交数据留存] 成功沉淀/更新了成交纪录到本地数据库: {db_path}")
    
    # 3. 极速导出数据库成交历史 Excel
    try:
        export_history_to_excel(conn)
    except Exception as e:
        print(f"导出成交历史 Excel 时出错: {e}")
        
    conn.close()

def export_history_to_excel(conn):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    df = pd.read_sql_query("SELECT * FROM sold_history ORDER BY sold_date DESC, project_name ASC", conn)
    if df.empty:
        return
        
    # 重命名列头，使之符合规范
    df.columns = [
        '地区', '商圈', '项目名称', '楼栋', '楼层', '房号',
        '户型', '实用面积', '成交日期', '总价(港币)', '呎价(港币/呎)',
        '最高折扣', '折实总价(港币)', '折实呎价(港币/呎)', '付款办法', '是否招标', '首次登记时间'
    ]
    
    excel_path = os.path.join(BASE_DIR, "全局成交历史明细.xlsx")
    
    # 高性能 pandas + openpyxl 直接生成
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='成交历史数据')
        
        wb = writer.book
        ws = writer.sheets['成交历史数据']
        
        # 表头样式
        font_header = Font(name='Microsoft YaHei', size=10, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='004D00', end_color='004D00', fill_type='solid') # 深绿配色
        border_thin = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            
        # 设置预设最佳列宽（避免数万行全量单元格遍历，提升数十倍速度）
        widths = [8, 12, 18, 10, 8, 8, 8, 10, 12, 14, 14, 10, 14, 14, 16, 8, 20]
        for idx, w in enumerate(widths, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = w
            
        # 页面打印自适应设置（自适应一页宽，高度自动分页）
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        # 统一 0.25 英寸页边距
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        ws.page_margins.header = 0.1
        ws.page_margins.footer = 0.1
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        
    print(f"[成交数据留存] 成功将全部成交历史导出为 Excel 表: {excel_path}")

# ==========================================
# 6. 主程序逻辑
# ==========================================
def main():
    token = fetch_user_token()
    api_headers = {
        'User-Agent': random.choice(USER_AGENTS),
        'Authorization': f'Bearer {token}',
        'Origin': 'https://www.hkp.com.hk',
        'Referer': 'https://www.hkp.com.hk/'
    }

    # 6.1 获取项目列表 (如果是重试模式，直接从本地文件载入失败项目)
    is_retry = "--retry" in sys.argv
    failed_projects_file = os.path.join(BASE_DIR, "failed_projects.json")

    if is_retry:
        print("\n==========================================")
        print("      正在以 [重试模式] 重新运行爬取        ")
        print("==========================================")
        if not os.path.exists(failed_projects_file):
            print("提示: 未找到失败项目记录文件 failed_projects.json，无需重试。")
            return
        try:
            with open(failed_projects_file, 'r', encoding='utf-8') as f:
                filtered_projects = json.load(f)
            print(f"成功载入需要重试的项目数: {len(filtered_projects)}")
            if not filtered_projects:
                print("提示: 失败项目记录为空，无需重试。")
                return
        except Exception as e:
            print(f"错误: 载入失败项目记录文件异常: {e}")
            return
    else:
        print("\n[Step 1] 正在请求项目列表 API...")
        try:
            api_headers['User-Agent'] = random.choice(USER_AGENTS)
            r = requests.get("https://data.hkp.com.hk/search/v2/new-properties", params={'limit': 1000}, headers=api_headers, timeout=15)
            r.raise_for_status()
            projects_data = r.json()
        except Exception as e:
            print(f"错误: 无法请求项目列表 API: {e}", file=sys.stderr)
            return

        all_projects = projects_data.get('result', [])
        print(f"API 返回项目总数: {len(all_projects)}")

        # 过滤出港岛和九龙项目
        target_regions = ['港岛', '九龙']
        filtered_projects = []
        for p in all_projects:
            region_raw = p.get('region', {}).get('name', '')
            region_cn = t2s(region_raw)
            if region_cn in target_regions:
                filtered_projects.append({
                    'id': p.get('id'),
                    'name': t2s(p.get('name')),
                    'region': region_cn,
                    'district': t2s(p.get('district')),
                    'developer': t2s(p.get('developer', {}).get('name', ''))
                })

        print(f"港岛/九龙一手新盘数量: {len(filtered_projects)}")

    EXCLUDE_DISTRICTS = {"将军澳", "茶果岭、油塘及鲤鱼门", "长沙湾", "牛头角及九龙湾", "慈云山、钻石山及新蒲岗"}
    filtered_projects = [
        p for p in filtered_projects
        if not (p.get('region') == '九龙' and p.get('district') in EXCLUDE_DISTRICTS)
    ]
    print(f"过滤后执行抓取的港岛/九龙一手新盘数量: {len(filtered_projects)}")

    # 确保保存价单的主目录存在
    os.makedirs(BASE_DIR, exist_ok=True)

    # 扫描已有文件夹映射 (包含价单目录和原楼盘字典目录，复用文件夹名)
    existing_folders = scan_existing_folders()
    
    # 汇总数据列表，用于生成全局汇总表
    global_units = []
    
    # 用于记录和报告未成功抓取最新数据的项目
    restored_projects = []
    totally_failed_projects = []

    # 6.2 循环遍历每个项目
    for idx, proj in enumerate(filtered_projects):
        pid = proj['id']
        pname = proj['name']
        region = proj.get('region') or proj.get('region_name') or '港岛'
        district = proj.get('district') or '未知'
        
        # 检查本地是否已售罄，若售罄则不再爬取
        clean_proj_name = clean_name(pname)
        excel_filename = None
        if clean_proj_name in existing_folders:
            folder_name = existing_folders[clean_proj_name]
            excel_filename = get_safe_filename(os.path.join(BASE_DIR, folder_name), pname)
            
        if excel_filename and is_project_sold_out(excel_filename):
            print(f"\n({idx+1}/{len(filtered_projects)}) 正在处理: {pname} (ID: {pid}, 区域: {region}-{district})")
            print(f"  [已售罄] 该项目去化率已达 100%，无需重复爬取，直接沿用并合并本地数据。")
            load_existing_project_data(pname, pid, region, district, proj.get('developer', ''), existing_folders, global_units)
            continue
            
        print(f"\n({idx+1}/{len(filtered_projects)}) 正在处理: {pname} (ID: {pid}, 区域: {region}-{district})")

        # 6.2.1 预载全局历史成交记录以作日期备用补充
        tx_lookup = {}
        tx_data = []
        try:
            tx_url = "https://data.hkp.com.hk/search/v2/transactions"
            tx_params = {'phase_ids': pid, 'limit': 2000}
            api_headers['User-Agent'] = random.choice(USER_AGENTS)
            tx_res = requests.get(tx_url, params=tx_params, headers=api_headers, timeout=12)
            if tx_res.status_code == 200:
                tx_data = tx_res.json().get('result', [])
                for tx in tx_data:
                    b_info = tx.get('building', {})
                    tx_bname = t2s(b_info.get('name', '')).strip()
                    tx_floor_num = str(tx.get('floor') or '').strip()
                    if not tx_floor_num:
                        fl_obj = tx.get('floor_level') or {}
                        tx_floor_num = fl_obj.get('name') if isinstance(fl_obj, dict) else str(fl_obj or '')
                    tx_floor = str(tx_floor_num).strip()
                    tx_flat = str(tx.get('flat', '')).strip().upper()
                    tx_date_raw = tx.get('tx_date') or tx.get('contract_date')
                    if tx_date_raw and tx_flat:
                        tx_lookup[(tx_bname, tx_floor, tx_flat)] = parse_hkt_date(tx_date_raw)
                print(f"  成功载入已登记的一手成交纪录 {len(tx_lookup)} 条。")
        except Exception as e:
            print(f"  提示: 预载一手历史成交纪录异常 ({e})。将只使用销控接口默认日期。")

        # 6.2.2 获取项目详情（以提取楼栋列表）
        try:
            api_headers['User-Agent'] = random.choice(USER_AGENTS)
            detail_res = requests.get(f"https://data.hkp.com.hk/info/v1/new-properties/{pid}", headers=api_headers, timeout=10)
            if detail_res.status_code != 200:
                print(f"  警告: 无法获取项目详情 (Code: {detail_res.status_code})。")
                if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                    restored_projects.append(pname)
                else:
                    totally_failed_projects.append(pname)
                continue
            detail_data = detail_res.json()
            if 'result' in detail_data and isinstance(detail_data['result'], dict):
                detail_data = detail_data['result']
            max_discount_rate, max_discount_title, best_plan_bonuses = extract_max_discount_info(detail_data)
            print(f"  计算该楼盘最高折扣率: {int(round(max_discount_rate * 100))}% | 付款办法: {max_discount_title}")
            
            henderson_pdf_db = {}
            # 恒基兆业实际独立操盘白名单 (包括经确认由恒基主导操盘的联合开发项目)
            HENDERSON_WHITELIST = {
                "The Henley I", "The Henley II", "The Henley III", 
                "Henley Park", "利奥坊．凯岸", "利奥坊．曦岸", 
                "The Haddon", "必嘉坊．迎汇", "必嘉坊．曦汇", 
                "首汇", "南首", "壹沐第1期", "壹沐第2期", 
                "君誉峰", "君豪峰", "Woodis",
                "天泷", "首岸第1期", "首岸第3期", "首岸第4期"
            }
            is_henderson = pname in HENDERSON_WHITELIST
            if is_henderson:
                clean_proj_name = clean_name(pname)
                if clean_proj_name in existing_folders:
                    folder_name = existing_folders[clean_proj_name]
                else:
                    folder_name = f"{region}-{district}-{pname}"
                    folder_name = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '', folder_name)
                project_dir = os.path.join(BASE_DIR, folder_name)
                
                print(f"  [恒基系专属] 检测到恒基项目 {pname}，正在启动价单 PDF 真实折扣物理精算...")
                try:
                    henderson_pdf_db = get_henderson_pdf_data(detail_data, api_headers, project_dir, max_discount_rate, pname)
                    print(f"  [恒基系专属] 成功加载 PDF 精算，录入 {len(henderson_pdf_db)} 个单位解析数据。")
                except Exception as pdf_err:
                    print(f"  [恒基系专属] 物理扫描 PDF 异常: {pdf_err}")
        except Exception as e:
            print(f"  警告: 获取项目详情异常 ({e})。")
            if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                restored_projects.append(pname)
            else:
                totally_failed_projects.append(pname)
            continue

        # 6.2.2 收集所有可能的候选楼栋 ID 和名称 (结合 standard buildings, floorplan 以及交易记录以防错配)
        buildings_map = {}
        
        # 1) 标准 buildings
        for b in detail_data.get('buildings', []):
            bid = b.get('id')
            bname = t2s(b.get('name'))
            if bid:
                buildings_map[bid] = bname
            for sub_b in b.get('buildings', []):
                sub_bid = sub_b.get('id')
                sub_bname = t2s(sub_b.get('name'))
                if sub_bid:
                    buildings_map[sub_bid] = sub_bname

        # 2) floorplan (补充提取如 院墅A, 院墅B, 院墅C 等特定洋房/独立屋 ID)
        for fp in detail_data.get('floorplan', []):
            b_info = fp.get('building', {})
            bid = b_info.get('id')
            bname = t2s(b_info.get('name'))
            if bid and bname:
                buildings_map[bid] = bname

        # 如果提取到了具名洋房 (如 院墅A/B/C)，清理掉无具体单位的泛指 generic '独立屋' / '龍駒道3號（獨立屋）'
        specific_villas = [b for b in buildings_map.values() if '院墅' in b or '屋' in b]
        if len(specific_villas) > 1:
            generic_keys = [k for k, v in buildings_map.items() if ('獨立屋' in v or '独立屋' in v) and not ('院墅' in v)]
            for gk in generic_keys:
                del buildings_map[gk]

        # 3) 交易数据 (仅在标准 buildings 和 floorplan 均为空时作为后备，防止跨项目错配楼栋)
        if not buildings_map:
            for tx in tx_data:
                b_info = tx.get('building', {})
                bid = b_info.get('id')
                bname = t2s(b_info.get('name'))
                if bid:
                    buildings_map[bid] = bname

        if not buildings_map:
            print("  提示: 该项目没有关联楼栋信息。")
            if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                restored_projects.append(pname)
            else:
                totally_failed_projects.append(pname)
            continue

        print(f"  包含候选楼栋数: {len(buildings_map)}")
        
        # 存储当前项目的所有楼栋销控数据，用以生成项目独立明细表
        project_buildings_data = {}

        # 6.2.3 循环拉取每个楼栋的销控数据
        for bid, bname in buildings_map.items():
            print(f"    -> 楼栋: {bname} (ID: {bid})")

            try:
                api_headers['User-Agent'] = random.choice(USER_AGENTS)
                units_res = requests.get(f"https://data.hkp.com.hk/info/v1/new-property/transactions/buildings/{bid}", headers=api_headers, timeout=10)
                if units_res.status_code != 200:
                    print(f"      警告: 无法拉取该楼栋的单位明细。")
                    continue
                units_data = units_res.json()
            except Exception as e:
                print(f"      警告: 请求楼栋单位明细异常 ({e})。")
                continue

            units = units_data.get('data', [])
            print(f"      获取到单位数: {len(units)}")
            
            parsed_units = []
            for u in units:
                unit_id = u.get('unit_id')
                floor = u.get('floor')
                flat = u.get('flat') or u.get('flat_name') or '-'
                net_area = u.get('net_area', 0) or 0
                status_raw = u.get('status', 'pending')
                is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
                
                # 🚨 规则固化：待售单位保持 pending 待售；在售招标单位纠正为 sale 在售
                if status_raw == 'pending':
                    status_raw = 'pending'
                    status_cn = '待售'
                elif is_tender and status_raw != 'sold' and status_raw != 'stopped':
                    status_raw = 'sale'
                    status_cn = '在售'
                else:
                    status_cn = STATUS_MAP.get(status_raw, '待售')
                
                # 获取总价
                price = u.get('price')
                if price is not None:
                    try:
                        price = float(price)
                    except ValueError:
                        price = None
                
                # 计算呎价
                unit_price_net = u.get('unit_price_net')
                if unit_price_net is not None:
                    try:
                        price_per_sq_ft = int(float(unit_price_net))
                    except ValueError:
                        price_per_sq_ft = None
                else:
                    if price and net_area > 0:
                        price_per_sq_ft = int(round(price / net_area))
                    else:
                        price_per_sq_ft = None

                # 1. 提取户型 (n房)
                room_layout = '暂无'
                detail_list = u.get('detail', [])
                if detail_list:
                    room_type = detail_list[0].get('room_type')
                    if room_type is not None and str(room_type) != '':
                        if str(room_type) in ['0', 0]:
                            room_layout = '开放式'
                        else:
                            room_layout = f"{room_type}房"
                
                # 2. 提取成交日期 (YYYY-MM-DD) 及反向自动纠偏已售状态
                sold_date_raw = u.get('sold_date') or u.get('tx_date')
                sold_date = '-'
                
                norm_b = normalize_bname(bname)
                norm_floor = str(floor).strip()
                norm_flat = str(flat).strip().upper()
                mapped_date = tx_lookup.get((norm_b, norm_floor, norm_flat))

                if status_raw == 'sold':
                    if sold_date_raw:
                        sold_date = parse_hkt_date(sold_date_raw)
                    elif mapped_date:
                        sold_date = mapped_date
                else:
                    # 如果销控接口未标注为 sold，但成交库中已被登记了一手买卖成交记录，自动反向纠偏为已售！
                    if mapped_date or (sold_date_raw and str(sold_date_raw).strip() not in ['-', 'None', '']):
                        status_raw = 'sold'
                        status_cn = STATUS_MAP.get('sold', '已售')
                        sold_date = mapped_date or parse_hkt_date(sold_date_raw)

                # 计算最高折扣折实价与付款办法
                discount_percent_str = '-'
                discount_price = None
                discount_price_per_sq_ft = None
                payment_method = '-'
                
                if max_discount_rate > 0 and status_raw != 'pending' and not (is_tender and status_raw != 'sold'):
                    if status_raw == 'sold':
                        discount_price = price
                        discount_price_per_sq_ft = price_per_sq_ft
                        discount_percent_str = '-'
                        payment_method = '-'
                    else:
                        # 针对在售定价单位计算精准的付款折扣（回退默认项目最高值）
                        u_rate, u_title = calculate_unit_discount(u, detail_data.get('payment', []), max_discount_rate, max_discount_title)
                        # 强力的 恒基系 PDF 价格预加载覆写
                        if henderson_pdf_db:
                            norm_b = normalize_bname(bname).replace("座", "")
                            norm_floor = str(floor).strip()
                            norm_flat = str(flat).strip()
                            pdf_key = (norm_b, norm_floor, norm_flat)
                            if pdf_key in henderson_pdf_db:
                                if "original_price" in henderson_pdf_db[pdf_key]:
                                    price = henderson_pdf_db[pdf_key]["original_price"]
                                
                        if price:
                            discount_price = round(price * (1 - u_rate))
                            discount_percent_str = format_discount_rate_str(u_rate)
                            payment_method = u_title
                            if net_area > 0:
                                discount_price_per_sq_ft = int(round(discount_price / net_area))
                                
                        # 强力的 PDF 覆写代理
                        if henderson_pdf_db:
                            norm_b = normalize_bname(bname).replace("座", "")
                            norm_floor = str(floor).strip()
                            norm_flat = str(flat).strip()
                            pdf_key = (norm_b, norm_floor, norm_flat)
                            if pdf_key in henderson_pdf_db:
                                pdf_info = henderson_pdf_db[pdf_key]
                                
                                # 双保险：仅当 API 售价为空（招标），或 PDF 算出的价格比 API 常规售价更划算（便宜）时才覆写
                                pdf_discount_price = pdf_info.get("discount_price")
                                if pdf_discount_price is None:
                                    pdf_discount_price = pdf_info.get("original_price", 0)
                                    
                                if not price or discount_price is None or pdf_discount_price < discount_price:
                                    if "original_price" in pdf_info:
                                        price = pdf_info["original_price"]
                                        if net_area > 0:
                                            price_per_sq_ft = int(round(price / net_area))
                                            
                                    if pdf_info.get("is_compensation"):
                                        if price:
                                            c_rate = pdf_info["direct_rate"]
                                            contract_price = int(price * (1 - c_rate))
                                            contract_price = (contract_price // 100) * 100
                                            
                                            if "启德海湾 1" in pname:
                                                sd_amount = int(contract_price * 0.0375)
                                                payment_method = f"現金付款計劃 (減{c_rate * 100:.1f}%) + 代繳3.75%印花稅"
                                            else:
                                                sd_amount = calculate_scale2_stamp_duty(contract_price)
                                                payment_method = f"現金付款計劃 (減{c_rate * 100:.1f}%) + 代繳從價印花稅"
                                                
                                            discount_price = contract_price - sd_amount
                                            discount_price = int((discount_price // 100) * 100)
                                            
                                            total_benefit_pct = (price - discount_price) / price
                                            discount_percent_str = f"{total_benefit_pct * 100:.2f}%"
                                            if net_area > 0:
                                                discount_price_per_sq_ft = int(round(discount_price / net_area))
                                    else:
                                        discount_price = pdf_info["discount_price"]
                                        discount_percent_str = pdf_info["discount_percent_str"]
                                        payment_method = pdf_info["payment_method"]
                                        if net_area > 0:
                                            discount_price_per_sq_ft = int(round(discount_price / net_area))

                unit_info = {
                    'unit_id': unit_id,
                    'floor': floor,
                    'flat': flat,
                    'net_area': net_area,
                    'status_raw': status_raw,
                    'status': status_cn,
                    'price': price,
                    'price_per_sq_ft': price_per_sq_ft,
                    'is_tender': u.get('is_tender'),
                    'room_layout': room_layout,
                    'sold_date': sold_date,
                    'discount_percent_str': discount_percent_str,
                    'discount_price': discount_price,
                    'discount_price_per_sq_ft': discount_price_per_sq_ft,
                    'payment_method': payment_method
                }
                parsed_units.append(unit_info)

                # 添加到全局汇总列表
                global_units.append({
                    '项目ID': pid,
                    '项目名称': pname,
                    '区域': region,
                    '商圈': district,
                    '开发商': proj['developer'],
                    '楼栋名称': bname,
                    '楼层': floor,
                    '房号': flat,
                    '户型': room_layout,
                    '实用面积': net_area if net_area > 0 else '暂无',
                    '销控状态': status_cn,
                    '成交日期': sold_date,
                    '总价': '招标单位' if (is_tender and status_raw != 'sold') else (price if price else '暂无'),
                    '呎价': '-' if (is_tender and status_raw != 'sold') else (price_per_sq_ft if price_per_sq_ft else '暂无'),
                    '最高折扣': discount_percent_str,
                    '折实总价': '招标单位' if (is_tender and status_raw != 'sold') else (discount_price if discount_price else '暂无'),
                    '折实呎价': '-' if (is_tender and status_raw != 'sold') else (discount_price_per_sq_ft if discount_price_per_sq_ft else '暂无'),
                    '付款办法': payment_method,
                    '是否招标': '是' if is_tender else '否'
                })

            project_buildings_data[bname] = parsed_units
            
            # 随机限速，比原本固定 0.3s 更安全
            time.sleep(random.uniform(0.6, 1.8))

        # 5.2.3 如果有数据，生成该项目的独立明细表和文件夹
        if project_buildings_data:
            # 匹配或创建文件夹
            clean_proj_name = clean_name(pname)
            if clean_proj_name in existing_folders:
                folder_name = existing_folders[clean_proj_name]
            else:
                folder_name = f"{region}-{district}-{pname}"
                folder_name = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '', folder_name)
                
            project_dir = os.path.join(BASE_DIR, folder_name)
            os.makedirs(project_dir, exist_ok=True)
            
            excel_filename = get_safe_filename(project_dir, pname)
            write_project_excel(excel_filename, pname, project_buildings_data)
            print(f"  [完成] 写入项目销控表: {excel_filename}")
        else:
            print("  [提示] 抓取的新数据为空，尝试沿用历史数据。")
            if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                restored_projects.append(pname)
            else:
                totally_failed_projects.append(pname)

        # 随机限速，比原本固定 0.5s 更安全
        time.sleep(random.uniform(1.5, 3.5))

    # 5.3 写入或更新全局汇总表
    if global_units:
        global_excel_path = os.path.join(BASE_DIR, "香港一手新盘销控汇总.xlsx")
        if is_retry:
            update_global_excel_with_retry(global_excel_path, global_units)
        else:
            write_global_excel(global_excel_path, global_units)
            print(f"\n[完成] 写入全局汇总表: {global_excel_path} (共 {len(global_units)} 条单位数据)")
        
        # 沉淀并留存历史成交数据
        try:
            save_transactions_to_db(global_units)
        except Exception as e:
            print(f"[错误] 成交数据留存执行异常: {e}")
    else:
        print("\n[警告] 未成功抓取到任何单位数据，全局汇总表及成交历史未更新。")

    # 记录未抓取成功的项目到本地 json，供重试流程读取
    all_failed = restored_projects + totally_failed_projects
    if all_failed:
        failed_dicts = []
        for name in all_failed:
            for p in filtered_projects:
                if p['name'] == name:
                    failed_dicts.append(p)
                    break
        try:
            with open(failed_projects_file, 'w', encoding='utf-8') as f:
                json.dump(failed_dicts, f, ensure_ascii=False, indent=2)
            print(f"已更新失败项目列表: {failed_projects_file} (共 {len(failed_dicts)} 个项目)")
        except Exception as e:
            print(f"写入失败项目记录文件失败: {e}")
    else:
        if os.path.exists(failed_projects_file):
            try:
                os.remove(failed_projects_file)
                print(f"所有项目均已成功，已清理历史失败记录文件。")
            except Exception as e:
                pass

    # 5.4 打印并输出未抓取到数据的项目统计报告
    from datetime import datetime
    print("\n==========================================")
    print("      未抓取到最新数据的项目统计报告        ")
    print("==========================================")
    
    total_skipped = len(restored_projects) + len(totally_failed_projects)
    print(f"共有 {total_skipped} 个项目在本次抓取中未获取到最新数据：\n")
    
    if restored_projects:
        print(f"🔹 以下 {len(restored_projects)} 个项目拉取失败，但已【成功恢复并沿用】历史数据：")
        for pname in sorted(restored_projects):
            print(f"  - {pname}")
        print()
        
    if totally_failed_projects:
        print(f"❌ 以下 {len(totally_failed_projects)} 个项目拉取失败，且【无历史数据】可用：")
        for pname in sorted(totally_failed_projects):
            print(f"  - {pname}")
        print()
        
    # 将未抓取到的项目列表写入本地文件，方便查看
    report_path = os.path.join(BASE_DIR, "未抓取到最新数据项目列表.txt")
    try:
        with open(report_path, 'w', encoding='utf-8') as rf:
            rf.write("==========================================\n")
            rf.write("      未抓取到最新数据的项目统计报告        \n")
            rf.write(f"      报告时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            rf.write("==========================================\n\n")
            
            rf.write(f"共有 {total_skipped} 个项目未获取到最新数据。\n\n")
            
            rf.write(f"【已恢复并沿用历史数据的项目】({len(restored_projects)}个):\n")
            for pname in sorted(restored_projects):
                rf.write(f"  - {pname}\n")
            rf.write("\n")
            
            rf.write(f"【完全未获取到数据且无历史数据的项目】({len(totally_failed_projects)}个):\n")
            for pname in sorted(totally_failed_projects):
                rf.write(f"  - {pname}\n")
                
        print(f"未抓取数据项目报告已写入本地文件: {report_path}")
    except Exception as e:
        print(f"写入报告报告失败: {e}")
    print("==========================================\n")


def _write_building_grid(wb, bname, units, font_data, font_header, font_title, align_center, border_thin, fill_header):
    """为单个楼栋写入销控网格 Tab"""
    safe_tab = re.sub(r'[\\/?*\[\]:]', '', bname)[:31]
    ws = wb.create_sheet(title=safe_tab)

    # 统计面板数据
    total = len(units)
    cnt = {s: 0 for s in STATUS_MAP.keys()}
    for u in units:
        sr = u.get('status_raw', 'pending')
        if sr in cnt:
            cnt[sr] += 1

    # 汇总统计面板 (Row 1 标题，Row 2-3 数据)
    stat_labels = [
        ('总套数', str(total), 'FFFFFF', '000000', False),
        ('在售 (Sale)', str(cnt.get('sale', 0)), 'A9F5A9', '004D00', True),
        ('已定价未售', str(cnt.get('priced', 0)), 'A9D0F5', '002060', False),
        ('已售 (Sold)', str(cnt.get('sold', 0)), 'F5A9A9', '660000', False),
        ('暂停销售', str(cnt.get('stopped', 0)), 'FFC000', '000000', True),
        ('待售 (Pending)', str(cnt.get('pending', 0)), 'FFFFFF', '7F7F7F', False),
    ]

    # Row 1: 楼栋名称
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(stat_labels))
    ws['A1'] = f"{bname} 销控网格图"
    ws['A1'].font = Font(name='Microsoft YaHei', size=13, bold=True)
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 28

    # Row 2: 图例标签
    for col, (label, val, bg, fg, bold) in enumerate(stat_labels, 1):
        c2 = ws.cell(row=2, column=col, value=label)
        c2.font = Font(name='Microsoft YaHei', size=9, bold=bold, color=fg)
        c2.alignment = align_center
        c2.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        c2.border = border_thin
        # Row 3: 数值
        c3 = ws.cell(row=3, column=col, value=val)
        c3.font = Font(name='Microsoft YaHei', size=10, bold=True, color=fg)
        c3.alignment = align_center
        c3.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        c3.border = border_thin
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # 整理楼层、单位
    floors_set = set()
    flats_set = set()
    for u in units:
        floors_set.add(u['floor'])
        flats_set.add(u['flat'])

    floors_sorted = sorted(floors_set, key=extract_floor_number, reverse=True)
    flats_sorted  = sorted(flats_set)

    # 建立单位查找表 (floor, flat) -> unit
    unit_lookup = {(u['floor'], u['flat']): u for u in units}

    # 网格表头 (Row 4 = 单位列)
    ws.cell(row=4, column=1, value='楼层').font = Font(name='Microsoft YaHei', size=9, bold=True)
    ws.cell(row=4, column=1).alignment = align_center
    ws.cell(row=4, column=1).fill = fill_header
    ws.cell(row=4, column=1).border = border_thin
    ws.row_dimensions[4].height = 18

    for col_idx, flat in enumerate(flats_sorted, 2):
        c = ws.cell(row=4, column=col_idx, value=flat)
        c.font = Font(name='Microsoft YaHei', size=9, bold=True)
        c.alignment = align_center
        c.fill = fill_header
        c.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.column_dimensions['A'].width = 8

    # 数据行
    for row_idx, floor in enumerate(floors_sorted, 5):
        ws.cell(row=row_idx, column=1, value=floor)
        ws.cell(row=row_idx, column=1).font = Font(name='Microsoft YaHei', size=9, bold=True)
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=1).fill = fill_header
        ws.cell(row=row_idx, column=1).border = border_thin
        ws.row_dimensions[row_idx].height = 58

        for col_idx, flat in enumerate(flats_sorted, 2):
            u = unit_lookup.get((floor, flat))
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = border_thin
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if u is None:
                c.value = ''
                c.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                continue

            status_raw = u.get('status_raw', 'pending')
            is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
            
            # 🚨 规范固化：待售单位保持 pending (白底/-)，在售招标单位强制为 sale (绿底/招标单位)
            if status_raw == 'pending':
                effective_status = 'pending'
            elif is_tender and status_raw != 'sold' and status_raw != 'stopped':
                effective_status = 'sale'
            else:
                effective_status = status_raw

            color_cfg = COLORS.get(effective_status, COLORS['pending'])
            bg = color_cfg['bg']
            fg = color_cfg['fg']
            is_bold = effective_status in ('sale', 'stopped')

            # 4行文字内容
            line1 = f"{flat} | {u['net_area']}呎 ({u['room_layout']})"
            if effective_status == 'pending':
                line2 = "-"
                line3 = "-"
            elif is_tender and status_raw != 'sold':
                line2 = "招标单位"
                line3 = "-"
            else:
                if u['price']:
                    price_wan = round(u['price'] / 10000)
                    line2 = f"${price_wan}万"
                else:
                    line2 = "-"
                if u['price_per_sq_ft']:
                    line3 = f"${u['price_per_sq_ft']:,}/呎"
                else:
                    line3 = "-"

            if status_raw == 'sold' and u['sold_date'] and u['sold_date'] != '-':
                try:
                    parts = u['sold_date'].split('-')
                    yr = parts[0][-2:]
                    mo = parts[1]
                    dy = parts[2][:2]
                    line4 = f"({yr}年-{mo}月-{dy}日)"
                except:
                    line4 = f"({u['sold_date']})"
            else:
                status_display = STATUS_MAP.get(effective_status, '待售')
                line4 = f"({status_display})"

            c.value = f"{line1}\n{line2}\n{line3}\n{line4}"
            c.font = Font(name='Microsoft YaHei', size=8, bold=is_bold, color=fg)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

    # 打印设置
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    for margin in ['left', 'right', 'top', 'bottom']:
        setattr(ws.page_margins, margin, 0.25)
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def _write_villa_grid(wb, project_name, buildings_data, font_data, align_center, border_thin):
    """
    独立屋项目整合为一个单一的工作表，避免分割成无数个 1x1 标签页。
    采用 5 行文字的网格卡片式布局。
    """
    ws_grid = wb.create_sheet(title="独立屋销控表")
    
    font_title  = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 展平所有独立屋单位并打上楼栋标记
    all_units = []
    for bname, units in buildings_data.items():
        for u in units:
            u_copy = u.copy()
            u_copy['_bname'] = bname
            all_units.append(u_copy)
            
    # 基于命名模式分析其网格排布 (行列定位)
    parentheses_matches = sum(1 for u in all_units if re.search(r'^.+?\s*[\(\（].+?[\)\）]$', u['_bname']))
    space_matches = sum(1 for u in all_units if ' ' in u['_bname'].strip())
    
    has_parentheses = parentheses_matches >= len(all_units) * 0.7
    has_space = space_matches >= len(all_units) * 0.7
    
    # 预设分配网格坐标
    for idx, u in enumerate(all_units):
        bname = u['_bname']
        if has_parentheses:
            m = re.search(r'^(.+?)\s*[\(\（](.+?)[\)\）]$', bname)
            if m:
                u['_grid_row'] = m.group(1).strip()
                u['_grid_col'] = m.group(2).strip()
            else:
                u['_grid_row'] = "其他"
                u['_grid_col'] = bname
        elif has_space:
            parts = bname.rsplit(' ', 1)
            u['_grid_row'] = parts[0].strip()
            u['_grid_col'] = parts[1].strip()
        else:
            # 默认横排扁平
            if len(all_units) <= 8:
                u['_grid_row'] = "独立屋"
                u['_grid_col'] = bname
            else:
                # 超过8个，默认后面按 cols_per_row = 6 排布分配
                pass
                
    # 按照大楼/别墅名进行自然排序
    def natural_key(x):
        return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', x['_bname'])]
        
    all_units = sorted(all_units, key=natural_key)
    
    # 如果是超过8个的扁平排布，在此根据自然排序顺序分配行列坐标
    if not has_parentheses and not has_space and len(all_units) > 8:
        cols_per_row = 6
        for idx, u in enumerate(all_units):
            row_num = idx // cols_per_row + 1
            col_num = idx % cols_per_row + 1
            u['_grid_row'] = f"第 {row_num} 行"
            u['_grid_col'] = f"位置 {col_num}"

    # 整理出确切的网格行、列头
    floors = sorted(list(set(u['_grid_row'] for u in all_units)))
    
    # 列头自然排序
    def col_natural_key(text):
        return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', text)]
    flats = sorted(list(set(u['_grid_col'] for u in all_units)), key=col_natural_key)
    
    unit_map = {(u['_grid_row'], u['_grid_col']): u for u in all_units}
    
    # 统计数量
    total_units = len(all_units)
    status_counts = {'sale': 0, 'sold': 0, 'priced': 0, 'stopped': 0, 'pending': 0}
    for u in all_units:
        st = u.get('status_raw', 'pending')
        status_counts[st] = status_counts.get(st, 0) + 1

    # 设置标题
    ws_grid.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(flats) + 1)
    title_cell = ws_grid.cell(row=1, column=1, value=f"{project_name} - 独立屋销控网格图")
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws_grid.row_dimensions[1].height = 40

    # 展示统计面板数据
    ws_grid.row_dimensions[2].height = 20
    ws_grid.row_dimensions[3].height = 25
    
    stat_cols = [
        ("总套数", total_units, 'pending'),
        ("在售 (Sale)", status_counts['sale'], 'sale'),
        ("已定价未售", status_counts['priced'], 'priced'),
        ("已售 (Sold)", status_counts['sold'], 'sold'),
        ("暂停销售", status_counts['stopped'], 'stopped'),
        ("待售 (Pending)", status_counts['pending'], 'pending')
    ]
    
    for stat_idx, (label, val, fill_key) in enumerate(stat_cols, 1):
        cell_lbl = ws_grid.cell(row=2, column=stat_idx, value=label)
        cell_lbl.font = Font(name='Microsoft YaHei', size=9, bold=(fill_key in ['sale', 'stopped']), color=COLORS[fill_key]['fg'])
        cell_lbl.alignment = align_center
        cell_lbl.fill = FILLS[fill_key]
        cell_lbl.border = border_thin
        
        cell_val = ws_grid.cell(row=3, column=stat_idx, value=f"{val} 套")
        cell_val.font = Font(name='Microsoft YaHei', size=11, bold=True, color=COLORS[fill_key]['fg'])
        cell_val.alignment = align_center
        cell_val.fill = FILLS[fill_key]
        cell_val.border = border_thin

    # 空一行后写表头
    header_row = 5
    ws_grid.row_dimensions[header_row].height = 25
    cell_cross = ws_grid.cell(row=header_row, column=1, value="类型/位置" if not has_parentheses else "类型/屋号")
    cell_cross.font = font_header
    cell_cross.alignment = align_center
    cell_cross.border = border_thin
    cell_cross.fill = fill_header

    for col_idx, flat_name in enumerate(flats, 2):
        cell_flat = ws_grid.cell(row=header_row, column=col_idx, value=flat_name)
        cell_flat.font = font_header
        cell_flat.alignment = align_center
        cell_flat.border = border_thin
        cell_flat.fill = fill_header

    # 写入网格数据
    curr_row = header_row + 1
    for f_name in floors:
        ws_grid.row_dimensions[curr_row].height = 80 # 提高以容纳5行文字
        
        cell_floor = ws_grid.cell(row=curr_row, column=1, value=f_name)
        cell_floor.font = font_header
        cell_floor.alignment = align_center
        cell_floor.border = border_thin
        cell_floor.fill = fill_header
        
        for col_idx, flat_name in enumerate(flats, 2):
            u = unit_map.get((f_name, flat_name))
            cell_unit = ws_grid.cell(row=curr_row, column=col_idx)
            cell_unit.border = border_thin
            cell_unit.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if u:
                status_raw = u['status_raw']
                status_cn = u['status']
                price = u['price']
                net_area = u['net_area']
                price_per_sq_ft = u['price_per_sq_ft']
                is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
                sold_date = u.get('sold_date')
                room_layout = u.get('room_layout', '暂无')
                
                # 1. 洋房名称/编号
                line1 = u['_bname']
                
                # 2. 实用面积与户型房数
                area_str = f"{net_area}呎" if net_area > 0 else "暂无"
                layout_str = f" ({room_layout})" if room_layout != '暂无' else ""
                line2 = f"{area_str}{layout_str}"
                
                # 3. 销售状态 (已售写具体登记年份-月份)
                if status_raw == 'sold' and sold_date and sold_date != '-':
                    try:
                        parts = sold_date.split('-')
                        yr = parts[0][-2:]
                        mo = parts[1]
                        dy = parts[2][:2]
                        line3 = f"({yr}年-{mo}月-{dy}日)"
                    except:
                        line3 = f"({sold_date})"
                else:
                    line3 = f"({status_cn})"
                    
                # 4. 价格 (过亿单位显示 $X.XX亿)
                if is_tender and status_raw != 'sold':
                    line4 = "招标项目" if status_raw == 'sale' else "招标单位"
                else:
                    if price and price > 0:
                        if price >= 100000000:
                            line4 = f"${price/100000000:.2f}亿"
                        elif price >= 10000:
                            line4 = f"${price/10000:,.0f}万"
                        else:
                            line4 = f"${price:,}"
                    else:
                        line4 = "暂无价格"
                        
                # 5. 实用呎价
                if is_tender and status_raw != 'sold':
                    line5 = "-"
                elif price_per_sq_ft and price_per_sq_ft > 0:
                    line5 = f"${price_per_sq_ft:,}/呎"
                else:
                    line5 = "暂无呎价"
                    
                cell_text = f"{line1}\n{line2}\n{line3}\n{line4}\n{line5}"
                cell_unit.value = cell_text
                
                cell_unit.font = Font(name='Microsoft YaHei', size=8, bold=(status_raw in ['sale', 'stopped']), color=COLORS.get(status_raw, COLORS['pending'])['fg'])
                cell_unit.fill = FILLS.get(status_raw, FILLS['pending'])
            else:
                cell_unit.value = ""
                cell_unit.fill = PatternFill(start_color='E5E5E5', end_color='E5E5E5', fill_type='solid')
                
        curr_row += 1

    # 设置列宽
    ws_grid.column_dimensions['A'].width = 12
    for col_idx in range(2, len(flats) + 2):
        col_letter = get_column_letter(col_idx)
        ws_grid.column_dimensions[col_letter].width = 18

    # 启用单页打印自适应设置
    ws_grid.sheet_properties.pageSetUpPr.fitToPage = True
    ws_grid.page_setup.fitToWidth = 1
    ws_grid.page_setup.fitToHeight = 1
    ws_grid.page_setup.orientation = ws_grid.ORIENTATION_LANDSCAPE
    ws_grid.page_setup.paperSize = ws_grid.PAPERSIZE_A4

    # 设置窄页边距 (0.25 英寸)
    for margin in ['left', 'right', 'top', 'bottom']:
        setattr(ws_grid.page_margins, margin, 0.25)
    ws_grid.page_margins.header = 0.1
    ws_grid.page_margins.footer = 0.1


if __name__ == '__main__':
    main()


