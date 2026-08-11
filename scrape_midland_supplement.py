#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
美联独有新盘补充抓取脚本 (scrape_midland_supplement.py) v2.0 动态对比版

功能：
  1. 实时拉取美联（midland.com.hk）和香港置业（hkp.com.hk）的全量项目列表
  2. 自动对比，找出美联独有但 HKP 没有的港岛/九龙一手新盘
  3. 对差集项目进行销控数据抓取，生成 Excel 销控明细表
  4. 每次运行都自动发现新上的盘，无需手动维护项目列表

数据源：https://data.midland.com.hk / https://data.hkp.com.hk
"""

import os
import re
import sys
import json
import time
import random
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# 1. 基础配置
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
zh2hans_path = os.path.join(os.path.dirname(BASE_DIR), "楼盘字典", "zh2hans.json")

if os.path.exists(zh2hans_path):
    with open(zh2hans_path, 'r', encoding='utf-8') as f:
        zh2hans_dict = json.load(f)
    char_map = {k: v for k, v in zh2hans_dict.items() if len(k) == 1 and len(v) == 1}
    TRADITIONAL_CHARS = "".join(char_map.keys())
    SIMPLIFIED_CHARS  = "".join(char_map.values())
else:
    TRADITIONAL_CHARS = "港島九龍啟德灣仔堅尼地城筲箕灣紅磡鰂魚涌鴨脷洲黃竹坑壽臣山山頂淺水灣深水灣渣甸山薄扶林香港仔上環中環土瓜灣北角摩星嶺黃泥涌"
    SIMPLIFIED_CHARS  = "港岛九龙启德湾仔坚尼地城筲箕湾红磡鲗鱼涌鸭脷洲黄竹坑寿臣山山顶浅水湾深水湾渣甸山薄扶林香港仔上环中环土瓜湾北角摩星岭黄泥涌"

TRANS_MAP = str.maketrans(TRADITIONAL_CHARS, SIMPLIFIED_CHARS)

def t2s(text):
    if not text:
        return ""
    if isinstance(text, dict):
        return {k: t2s(v) for k, v in text.items()}
    if isinstance(text, list):
        return [t2s(x) for x in text]
    return str(text).translate(TRANS_MAP)

def clean_name(name):
    if not name:
        return ""
    name = t2s(name).lower()
    name = re.sub(r'[\s\-\(\)\（\）\.\,\，\。]', '', name)
    return name

def get_safe_filename(project_dir, pname):
    safe_pname = re.sub(r'[\/\\\:\*\?\"<>|]', '', pname)
    return os.path.join(project_dir, f"{safe_pname}_销控明细表.xlsx")

def normalize_bname(name):
    if not name:
        return ""
    name = t2s(name).strip()
    name = re.sub(r'^第', '', name)
    return name

def extract_floor_number(floor_str):
    if not floor_str:
        return -99
    floor_str = str(floor_str).upper().strip()
    match = re.search(r'(\d+)', floor_str)
    if match:
        val = int(match.group(1))
        if 'R' in floor_str:
            val += 0.5
        return val
    if 'G' in floor_str or 'UG' in floor_str:
        return 0
    if 'B' in floor_str:
        return -1
    return -99

# ==========================================
# 2. 状态映射与颜色配置
# ==========================================
STATUS_MAP = {
    'pending': '待售',
    'priced': '已定价未售',
    'sale': '在售',
    'sold': '已售',
    'stopped': '暂停销售'
}

COLORS = {
    'sale':    {'bg': 'A9F5A9', 'fg': '004D00'},
    'sold':    {'bg': 'F5A9A9', 'fg': '660000'},
    'priced':  {'bg': 'A9D0F5', 'fg': '002060'},
    'stopped': {'bg': 'FFC000', 'fg': '000000'},
    'pending': {'bg': 'FFFFFF', 'fg': '7F7F7F'}
}

FILLS = {k: PatternFill(start_color=v['bg'], end_color=v['bg'], fill_type='solid') for k, v in COLORS.items()}

# ==========================================
# 3. 区域代码映射（美联 API 区域代码）
# ==========================================
MIDLAND_REGIONS = {
    'HKI': ('港岛', ['Hong Kong Island', 'HKI']),
    'KLN': ('九龙', ['Kowloon', 'KLN']),
}

# 美联独有且不再出现在 active 列表中的历史新盘（这些项目依然在美联有详情和销控页）
STATIC_MIDLAND_PROJECTS = {
    'E000001984': {'name': '御海园', 'region': '港岛', 'district': '坚尼地城及摩星岭'},
    'E000008181': {'name': 'Shouson Peak', 'region': '港岛', 'district': '寿臣山及浅水湾'},
    'E000000456': {'name': '寿臣山道东1号', 'region': '港岛', 'district': '寿臣山及浅水湾'},
    'E000015288': {'name': 'Twelve Peaks', 'region': '港岛', 'district': '山顶区'},
    'E000016272': {'name': '大潭道45号', 'region': '港岛', 'district': '大潭及石澳'},
    'E000016270': {'name': '77/79 Peak Road', 'region': '港岛', 'district': '山顶区'},
    'E000016346': {'name': '浅水湾108', 'region': '港岛', 'district': '寿臣山及浅水湾'},
    'E000015301': {'name': '远晴', 'region': '港岛', 'district': '筲箕湾'},
    'E000015268': {'name': '339 Tai Hang Road', 'region': '港岛', 'district': '跑马地'},
    'E000015314': {'name': '喇沙汇', 'region': '九龙', 'district': '九龙塘'},
    'E000015455': {'name': '皇廷汇', 'region': '九龙', 'district': '九龙塘'},
    'E000004402': {'name': '天玺', 'region': '九龙', 'district': '九龙站'},
    'E000017140': {'name': '本木', 'region': '九龙', 'district': '尖沙咀'},
    'E000008203': {'name': '耀爵台', 'region': '九龙', 'district': '何文田'},
    'E000015299': {'name': '宾吉道3号', 'region': '港岛', 'district': '山顶区'}
}


# 港岛已知商圈关键词（含这些词语的就是港岛）
HKI_DISTRICT_KEYWORDS = [
    '北角', '半山', '坚尼地城', '摩星岭', '大潭', '石澳', '寿臣山', '浅水湾', '深水湾',
    '山顶', '赤柱', '铜锣湾', '湾仔', '筲箕湾', '薄扶林', '营盘', '上环', '中环',
    '香港仔', '鸭脷洲', '黄竹坑', '黄泥涌', '渣甸山', '鲗鱼涌', '柴湾', '岛山',
    '西区', '东区', '南区'
]
# 九龙已知商圈关键词
KLN_DISTRICT_KEYWORDS = [
    '九龙塘', '西南九龙', '尖沙咀', '何文田', '启德', '红磡', '马头角', '土瓜湾',
    '长沙湾', '深水埗', '旺角', '油麻地', '太子', '牛头角', '九龙湾', '观塘',
    '茶果岭', '油塘', '鲤鱼门', '慈云山', '钻石山', '荔枝角', '石硖尾', '新蒲岗'
]

ALL_VALID_KEYWORDS = HKI_DISTRICT_KEYWORDS + KLN_DISTRICT_KEYWORDS


# ==========================================
# 4. Token 获取（优先动态获取，失败时使用备用静态 Token）
# ==========================================
MIDLAND_BUILD_TOKEN = (
    "eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJndWlkIjoibXItMjAyMC0xMC0yOC1LdUlFbUR5bnlfWDV0NVVDMkZwSExpdFBodGxLNWRhblV5TFo4WXY0bVJIZEhJY3lrSW1fRmYzS0NYWjlkcENiOEp0Y3hLUy1JSSIsImlhdCI6MTYwMzg3ODk2MiwiaXNzIjoiZGF0YS5taWRsYW5kLmNvbS5oayJ9."
    "O2ASPk6CIDbK4cexR1FOy-TwOOvc-o1KoGANo_ckoOVvyB7OI-sif3VwlAEHZ79dGuC1bWa-oYsbelfRnpX5HZx7cM667gjd7sXycp5j9pCbGwPJ8dzOpwHfnsMIiTnTeL0_D0_PV8-PK7nsrH8iDM7cumD-F6qDl0THXNUahqXycf6MipIAv1CENyMKI3m9wWsoWsnw3T0RGJQ4YtNZ0itackeQbnsFCwm5wxJ2nDNb_yUmN0dA8cApHNaJb8f8IZ91HOf_IXQjwy_LIm4A7VlnOKVx2tvEvK4fOdPVLPbGyHeNINESOhmqCq-Z6ZbvvAel6IySf268NuV_Olk3DUQsBeMRSXrfNEHMGKMpCawtmtZgoRlFnSBYuzgJ1-MkyEgA7Sap6fM6RRhWKi7pST0RiVYhIeQypYaKIMvbsVyvfGGrxptjJzMt8eBUxYK6mZHl2q3N2RBCmc30x7XLsgpQNdW3sQcvA9AsBPWDATFBdVy8l6phVYypD8bn-EOfEZalLle95M1Y80FWo_ek6_uUSMOmLo0YXz6-GgJaJ6GkMf1Syx90bsGjgWHhgGJvLebbwx3Fl7ecVv-afdc6oNSDjQkZID609kaxE6LqEEWQoxGOYrOZDRAhA_dTYgxmjVqhqiBDQOf4ooMvkAVwGL17c9REkNInowi9JX7RHyw"
)

MIDLAND_API_BASE = "https://data.midland.com.hk"
USER_AGENTS = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:125.0) Gecko/20100101 Firefox/125.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 17_4_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Mobile/15E148 Safari/604.1'
]

HEADERS_UA = {
    'User-Agent': USER_AGENTS[0]
}

def fetch_midland_token():
    """
    尝试从 www.midland.com.hk 动态获取最新 Token。
    若失败，则使用备用静态 Build Token。
    """
    print("正在获取美联 Token...")
    # 尝试从页面提取动态 token
    for page_url in [
        "https://www.midland.com.hk/zh-cn/new-property/",
        "https://www.midland.com.hk/zh-hk/new-property/",
    ]:
        try:
            headers = {'User-Agent': random.choice(USER_AGENTS)}
            r = requests.get(page_url, headers=headers, timeout=15)
            if r.status_code == 200:
                # 查找 BUILD_TOKEN 或 userToken
                m = re.search(r'"BUILD_TOKEN"\s*:\s*"([^"]+)"', r.text)
                if m:
                    token = m.group(1)
                    print(f"  成功从页面提取到 BUILD_TOKEN（长度：{len(token)}）")
                    return token
                m2 = re.search(r'"userToken"\s*:\s*"([^"]+)"', r.text)
                if m2:
                    token = m2.group(1)
                    print(f"  成功从页面提取到 userToken（长度：{len(token)}）")
                    return token
        except Exception as e:
            print(f"  警告: 动态获取 Token 失败 ({e})")

    print(f"  使用备用静态 Build Token（长度：{len(MIDLAND_BUILD_TOKEN)}）")
    return MIDLAND_BUILD_TOKEN


# ==========================================
# 5. Excel 写入（完整复刻 scrape_hkp_sales_control.py 的格式）
# ==========================================
def write_project_excel(filepath, project_name, buildings_data):
    """
    为单个项目写入 Excel 销控表，包含：
    - Tab 1: "销控汇总明细"
    - Tab 2+: 每个楼栋的"销控网格图 (Grid)"
    """
    wb = openpyxl.Workbook()

    # ---- Tab 1: 销控汇总明细 ----
    ws_detail = wb.active
    ws_detail.title = "销控汇总明细"

    font_title  = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    font_data   = Font(name='Microsoft YaHei', size=10)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin  = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    ws_detail.merge_cells("A1:J1")
    ws_detail['A1'] = f"{project_name} - 一手销控汇总明细"
    ws_detail['A1'].font = font_title
    ws_detail['A1'].alignment = align_center
    ws_detail.row_dimensions[1].height = 40

    headers = ["楼栋", "楼层", "房号", "户型", "实用面积 (平方呎)", "销控状态", "成交日期", "总价 (港币)", "实用呎价 (港币/呎)", "是否招标"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=2, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_thin
    ws_detail.row_dimensions[2].height = 25

    row_idx = 3
    for bname, units in buildings_data.items():
        sorted_units = sorted(units, key=lambda x: (extract_floor_number(x['floor']), x['flat']), reverse=True)
        for u in sorted_units:
            ws_detail.cell(row=row_idx, column=1, value=bname)
            ws_detail.cell(row=row_idx, column=2, value=u['floor'])
            ws_detail.cell(row=row_idx, column=3, value=u['flat'])
            ws_detail.cell(row=row_idx, column=4, value=u['room_layout'])
            ws_detail.cell(row=row_idx, column=5, value=u['net_area'] if u['net_area'] > 0 else '暂无')
            ws_detail.cell(row=row_idx, column=6, value=u['status'])
            ws_detail.cell(row=row_idx, column=7, value=u['sold_date'])

            is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
            if is_tender and u['status_raw'] != 'sold':
                ws_detail.cell(row=row_idx, column=8, value='招标单位')
                ws_detail.cell(row=row_idx, column=9, value='-')
            else:
                ws_detail.cell(row=row_idx, column=8, value=u['price'] if u['price'] else '暂无')
                ws_detail.cell(row=row_idx, column=9, value=u['price_per_sq_ft'] if u['price_per_sq_ft'] else '暂无')
            ws_detail.cell(row=row_idx, column=10, value='是' if is_tender else '否')

            for col in range(1, 11):
                c = ws_detail.cell(row=row_idx, column=col)
                c.font = font_data
                c.alignment = align_center
                c.border = border_thin
                if col == 8 and isinstance(c.value, (int, float)):
                    c.number_format = '$#,##0'
                elif col == 9 and isinstance(c.value, (int, float)):
                    c.number_format = '$#,##0'
                elif col == 5 and isinstance(c.value, (int, float)):
                    c.number_format = '#,##0'
            row_idx += 1

    for col in ws_detail.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_detail.column_dimensions[col_letter].width = max(max_len + 3, 12)

    ws_detail.sheet_properties.pageSetUpPr.fitToPage = True
    ws_detail.page_setup.fitToWidth = 1
    ws_detail.page_setup.fitToHeight = 0
    ws_detail.page_setup.orientation = ws_detail.ORIENTATION_PORTRAIT
    ws_detail.page_setup.paperSize = ws_detail.PAPERSIZE_A4
    for margin in ['left', 'right', 'top', 'bottom']:
        setattr(ws_detail.page_margins, margin, 0.25)
    ws_detail.page_margins.header = 0.1
    ws_detail.page_margins.footer = 0.1

    multi_buildings = {}
    villa_buildings = {}
    
    if buildings_data:
        for bname, units in buildings_data.items():
            # 判断是否为独栋 (如果该楼栋单位数 <= 2 或是 HOUSE 类型)
            has_house_type = any(u.get('unit_type') == 'HOUSE' for u in units)
            if len(units) <= 2 or has_house_type:
                villa_buildings[bname] = units
            else:
                multi_buildings[bname] = units

    # 1. 多个单位的楼栋，每个楼栋放在单独的工作表
    for bname, units in multi_buildings.items():
        _write_building_grid(wb, bname, units, font_data, font_header, font_title, align_center, border_thin, fill_header)

    # 2. 独栋楼，集合放在一个工作表
    if villa_buildings:
        _write_villa_grid(wb, project_name, villa_buildings, font_data, align_center, border_thin)

    wb.save(filepath)


def _write_building_grid(wb, bname, units, font_data, font_header, font_title, align_center, border_thin, fill_header):
    """为单个楼栋写入销控网格 Tab"""
    safe_tab = re.sub(r'[\\/?*\[\]:]', '', bname)[:31]
    ws = wb.create_sheet(title=safe_tab)

    # 统计面板数据
    total = len(units)
    cnt = {s: 0 for s in STATUS_MAP.keys()}
    for u in units:
        sr = u.get('status_raw', 'pending')
        if sr in cnt:
            cnt[sr] += 1

    # 汇总统计面板 (Row 1 标题，Row 2-3 数据)
    stat_labels = [
        ('总套数', str(total), 'FFFFFF', '000000', False),
        ('在售 (Sale)', str(cnt.get('sale', 0)), 'A9F5A9', '004D00', True),
        ('已定价未售', str(cnt.get('priced', 0)), 'A9D0F5', '002060', False),
        ('已售 (Sold)', str(cnt.get('sold', 0)), 'F5A9A9', '660000', False),
        ('暂停销售', str(cnt.get('stopped', 0)), 'FFC000', '000000', True),
        ('待售 (Pending)', str(cnt.get('pending', 0)), 'FFFFFF', '7F7F7F', False),
    ]

    # Row 1: 楼栋名称
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(stat_labels))
    ws['A1'] = f"{bname} 销控网格图"
    ws['A1'].font = Font(name='Microsoft YaHei', size=13, bold=True)
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 28

    # Row 2: 图例标签
    for col, (label, val, bg, fg, bold) in enumerate(stat_labels, 1):
        c2 = ws.cell(row=2, column=col, value=label)
        c2.font = Font(name='Microsoft YaHei', size=9, bold=bold, color=fg)
        c2.alignment = align_center
        c2.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        c2.border = border_thin
        # Row 3: 数值
        c3 = ws.cell(row=3, column=col, value=val)
        c3.font = Font(name='Microsoft YaHei', size=10, bold=True, color=fg)
        c3.alignment = align_center
        c3.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        c3.border = border_thin
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # 整理楼层、单位
    floors_set = set()
    flats_set = set()
    for u in units:
        floors_set.add(u['floor'])
        flats_set.add(u['flat'])

    floors_sorted = sorted(floors_set, key=extract_floor_number, reverse=True)
    flats_sorted  = sorted(flats_set)

    # 建立单位查找表 (floor, flat) -> unit
    unit_lookup = {(u['floor'], u['flat']): u for u in units}

    # 网格表头 (Row 4 = 单位列)
    ws.cell(row=4, column=1, value='楼层').font = Font(name='Microsoft YaHei', size=9, bold=True)
    ws.cell(row=4, column=1).alignment = align_center
    ws.cell(row=4, column=1).fill = fill_header
    ws.cell(row=4, column=1).border = border_thin
    ws.row_dimensions[4].height = 18

    for col_idx, flat in enumerate(flats_sorted, 2):
        c = ws.cell(row=4, column=col_idx, value=flat)
        c.font = Font(name='Microsoft YaHei', size=9, bold=True)
        c.alignment = align_center
        c.fill = fill_header
        c.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.column_dimensions['A'].width = 8

    # 数据行
    for row_idx, floor in enumerate(floors_sorted, 5):
        ws.cell(row=row_idx, column=1, value=floor)
        ws.cell(row=row_idx, column=1).font = Font(name='Microsoft YaHei', size=9, bold=True)
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=1).fill = fill_header
        ws.cell(row=row_idx, column=1).border = border_thin
        ws.row_dimensions[row_idx].height = 58

        for col_idx, flat in enumerate(flats_sorted, 2):
            u = unit_lookup.get((floor, flat))
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = border_thin
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if u is None:
                c.value = ''
                c.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                continue

            status_raw = u.get('status_raw', 'pending')
            color_cfg = COLORS.get(status_raw, COLORS['pending'])
            bg = color_cfg['bg']
            fg = color_cfg['fg']
            is_bold = status_raw in ('sale', 'stopped')

            # 4行文字内容
            line1 = f"{flat} | {u['net_area']}呎 ({u['room_layout']})"
            is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
            if is_tender and status_raw != 'sold':
                line2 = "招标单位"
                line3 = "-"
            else:
                if u['price']:
                    price_wan = round(u['price'] / 10000)
                    line2 = f"${price_wan}万"
                else:
                    line2 = "-"
                if u['price_per_sq_ft']:
                    line3 = f"${u['price_per_sq_ft']:,}/呎"
                else:
                    line3 = "-"

            if status_raw == 'sold' and u['sold_date'] and u['sold_date'] != '-':
                try:
                    parts = u['sold_date'].split('-')
                    yr = parts[0][-2:]
                    mo = parts[1]
                    dy = parts[2][:2] if len(parts) >= 3 else ""
                    if dy:
                        line4 = f"({yr}年-{mo}月-{dy}日)"
                    else:
                        line4 = f"({yr}年-{mo}月)"
                except:
                    line4 = f"({u['sold_date']})"
            else:
                status_display = STATUS_MAP.get(status_raw, '待售')
                line4 = f"({status_display})"

            c.value = f"{line1}\n{line2}\n{line3}\n{line4}"
            c.font = Font(name='Microsoft YaHei', size=8, bold=is_bold, color=fg)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

    # 打印设置
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    for margin in ['left', 'right', 'top', 'bottom']:
        setattr(ws.page_margins, margin, 0.25)
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def _write_villa_grid(wb, project_name, buildings_data, font_data, align_center, border_thin):
    """
    独立屋项目整合为一个单一的工作表，避免分割成无数个 1x1 标签页。
    采用 5 行文字的网格卡片式布局（与 scrape_hkp_sales_control.py 一致）。
    """
    ws_grid = wb.create_sheet(title="独立屋销控表")
    
    font_title  = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 展平所有独立屋单位并打上楼栋标记
    all_units = []
    for bname, units in buildings_data.items():
        for u in units:
            u_copy = u.copy()
            u_copy['_bname'] = bname
            all_units.append(u_copy)
            
    # 基于命名模式分析其网格排布 (行列定位)
    parentheses_matches = sum(1 for u in all_units if re.search(r'^.+?\s*[\(\（].+?[\)\）]$', u['_bname']))
    space_matches = sum(1 for u in all_units if ' ' in u['_bname'].strip())
    
    has_parentheses = parentheses_matches >= len(all_units) * 0.7
    has_space = space_matches >= len(all_units) * 0.7
    
    # 预设分配网格坐标
    for idx, u in enumerate(all_units):
        bname = u['_bname']
        if has_parentheses:
            m = re.search(r'^(.+?)\s*[\(\（](.+?)[\)\）]$', bname)
            if m:
                u['_grid_row'] = m.group(1).strip()
                u['_grid_col'] = m.group(2).strip()
            else:
                u['_grid_row'] = "其他"
                u['_grid_col'] = bname
        elif has_space:
            parts = bname.rsplit(' ', 1)
            u['_grid_row'] = parts[0].strip()
            u['_grid_col'] = parts[1].strip()
        else:
            # 默认横排扁平
            if len(all_units) <= 8:
                u['_grid_row'] = "独立屋"
                u['_grid_col'] = bname
            else:
                # 超过8个，默认后面按 cols_per_row = 6 排布分配
                pass
                
    # 按照大楼/别墅名进行自然排序
    def natural_key(x):
        return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', x['_bname'])]
        
    all_units = sorted(all_units, key=natural_key)
    
    # 如果是超过8个的扁平排布，在此根据自然排序顺序分配行列坐标
    if not has_parentheses and not has_space and len(all_units) > 8:
        cols_per_row = 6
        for idx, u in enumerate(all_units):
            row_num = idx // cols_per_row + 1
            col_num = idx % cols_per_row + 1
            u['_grid_row'] = f"第 {row_num} 行"
            u['_grid_col'] = f"位置 {col_num}"

    # 整理出确切的网格行、列头
    floors = sorted(list(set(u['_grid_row'] for u in all_units)))
    
    # 列头自然排序
    def col_natural_key(text):
        return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', text)]
    flats = sorted(list(set(u['_grid_col'] for u in all_units)), key=col_natural_key)
    
    unit_map = {(u['_grid_row'], u['_grid_col']): u for u in all_units}
    
    # 统计数量
    total_units = len(all_units)
    status_counts = {'sale': 0, 'sold': 0, 'priced': 0, 'stopped': 0, 'pending': 0}
    for u in all_units:
        st = u.get('status_raw', 'pending')
        status_counts[st] = status_counts.get(st, 0) + 1

    # 设置标题
    ws_grid.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(flats) + 1)
    title_cell = ws_grid.cell(row=1, column=1, value=f"{project_name} - 独立屋销控网格图")
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws_grid.row_dimensions[1].height = 40

    # 展示统计面板数据
    ws_grid.row_dimensions[2].height = 20
    ws_grid.row_dimensions[3].height = 25
    
    stat_cols = [
        ("总套数", total_units, 'pending'),
        ("在售 (Sale)", status_counts['sale'], 'sale'),
        ("已定价未售", status_counts['priced'], 'priced'),
        ("已售 (Sold)", status_counts['sold'], 'sold'),
        ("暂停销售", status_counts['stopped'], 'stopped'),
        ("待售 (Pending)", status_counts['pending'], 'pending')
    ]
    
    for stat_idx, (label, val, fill_key) in enumerate(stat_cols, 1):
        cell_lbl = ws_grid.cell(row=2, column=stat_idx, value=label)
        cell_lbl.font = Font(name='Microsoft YaHei', size=9, bold=(fill_key in ['sale', 'stopped']), color=COLORS[fill_key]['fg'])
        cell_lbl.alignment = align_center
        cell_lbl.fill = FILLS[fill_key]
        cell_lbl.border = border_thin
        
        cell_val = ws_grid.cell(row=3, column=stat_idx, value=f"{val} 套")
        cell_val.font = Font(name='Microsoft YaHei', size=11, bold=True, color=COLORS[fill_key]['fg'])
        cell_val.alignment = align_center
        cell_val.fill = FILLS[fill_key]
        cell_val.border = border_thin

    # 空一行后写表头
    header_row = 5
    ws_grid.row_dimensions[header_row].height = 25
    cell_cross = ws_grid.cell(row=header_row, column=1, value="类型/位置" if not has_parentheses else "类型/屋号")
    cell_cross.font = font_header
    cell_cross.alignment = align_center
    cell_cross.border = border_thin
    cell_cross.fill = fill_header

    for col_idx, flat_name in enumerate(flats, 2):
        cell_flat = ws_grid.cell(row=header_row, column=col_idx, value=flat_name)
        cell_flat.font = font_header
        cell_flat.alignment = align_center
        cell_flat.border = border_thin
        cell_flat.fill = fill_header

    # 写入网格数据
    curr_row = header_row + 1
    for f_name in floors:
        ws_grid.row_dimensions[curr_row].height = 80 # 提高以容纳5行文字
        
        cell_floor = ws_grid.cell(row=curr_row, column=1, value=f_name)
        cell_floor.font = font_header
        cell_floor.alignment = align_center
        cell_floor.border = border_thin
        cell_floor.fill = fill_header
        
        for col_idx, flat_name in enumerate(flats, 2):
            u = unit_map.get((f_name, flat_name))
            cell_unit = ws_grid.cell(row=curr_row, column=col_idx)
            cell_unit.border = border_thin
            cell_unit.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if u:
                status_raw = u['status_raw']
                status_cn = u['status']
                price = u['price']
                net_area = u['net_area']
                price_per_sq_ft = u['price_per_sq_ft']
                is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
                sold_date = u.get('sold_date')
                room_layout = u.get('room_layout', '暂无')
                
                # 1. 洋房名称/编号
                line1 = u['_bname']
                
                # 2. 实用面积与户型房数
                area_str = f"{net_area}呎" if net_area > 0 else "暂无"
                layout_str = f" ({room_layout})" if room_layout != '暂无' else ""
                line2 = f"{area_str}{layout_str}"
                
                # 3. 销售状态 (已售写具体登记年份-月份)
                if status_raw == 'sold' and sold_date and sold_date != '-':
                    try:
                        parts = sold_date.split('-')
                        yr = parts[0][-2:]
                        mo = parts[1]
                        dy = parts[2][:2] if len(parts) >= 3 else ""
                        if dy:
                            line3 = f"({yr}年-{mo}月-{dy}日)"
                        else:
                            line3 = f"({yr}年-{mo}月)"
                    except:
                        line3 = f"({sold_date})"
                else:
                    line3 = f"({status_cn})"
                    
                # 4. 价格 (过亿单位显示 $X.XX亿)
                if is_tender and status_raw != 'sold':
                    line4 = "招标项目" if status_raw == 'sale' else "招标单位"
                else:
                    if price and price > 0:
                        if price >= 100000000:
                            line4 = f"${price/100000000:.2f}亿"
                        elif price >= 10000:
                            line4 = f"${price/10000:,.0f}万"
                        else:
                            line4 = f"${price:,}"
                    else:
                        line4 = "暂无价格"
                        
                # 5. 实用呎价
                if is_tender and status_raw != 'sold':
                    line5 = "-"
                elif price_per_sq_ft and price_per_sq_ft > 0:
                    line5 = f"${price_per_sq_ft:,}/呎"
                else:
                    line5 = "暂无呎价"
                    
                cell_text = f"{line1}\n{line2}\n{line3}\n{line4}\n{line5}"
                cell_unit.value = cell_text
                
                cell_unit.font = Font(name='Microsoft YaHei', size=8, bold=(status_raw in ['sale', 'stopped']), color=COLORS.get(status_raw, COLORS['pending'])['fg'])
                cell_unit.fill = FILLS.get(status_raw, FILLS['pending'])
            else:
                cell_unit.value = ""
                cell_unit.fill = PatternFill(start_color='E5E5E5', end_color='E5E5E5', fill_type='solid')
                
        curr_row += 1

    # 设置列宽
    ws_grid.column_dimensions['A'].width = 12
    for col_idx in range(2, len(flats) + 2):
        col_letter = get_column_letter(col_idx)
        ws_grid.column_dimensions[col_letter].width = 18

    # 启用单页打印自适应设置
    ws_grid.sheet_properties.pageSetUpPr.fitToPage = True
    ws_grid.page_setup.fitToWidth = 1
    ws_grid.page_setup.fitToHeight = 1
    ws_grid.page_setup.orientation = ws_grid.ORIENTATION_LANDSCAPE
    ws_grid.page_setup.paperSize = ws_grid.PAPERSIZE_A4

    # 设置窄页边距 (0.25 英寸)
    for margin in ['left', 'right', 'top', 'bottom']:
        setattr(ws_grid.page_margins, margin, 0.25)
    ws_grid.page_margins.header = 0.1
    ws_grid.page_margins.footer = 0.1


def is_project_sold_out(excel_filename):
    """检查本地 excel 文件是否记录为 100% 售出，且所有已售单位均有成交日期。"""
    if not excel_filename or not os.path.exists(excel_filename):
        return False
    try:
        wb = openpyxl.load_workbook(excel_filename, data_only=True, read_only=True)
        if "销控汇总明细" not in wb.sheetnames:
            return False
        ws = wb["销控汇总明细"]
        total = 0
        sold = 0
        missing_date = 0
        for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
            if not row or row[0] is None:
                break
            total += 1
            status_val = str(row[5]).strip() if len(row) >= 6 and row[5] is not None else ""
            date_val = str(row[6]).strip() if len(row) >= 7 and row[6] is not None else "-"
            if status_val == "已售":
                sold += 1
                if date_val == '-' or not date_val:
                    missing_date += 1
        wb.close()
        return total > 0 and sold == total and missing_date == 0
    except Exception as e:
        print(f"  警告: 检查本地售罄状态异常 ({e})。")
    return False


def load_existing_project_data(pname, pid, region, district, developer, existing_folders, global_units):
    """从本地已存 excel 中加载数据加入 global_units"""
    clean_proj = clean_name(pname)
    if clean_proj not in existing_folders:
        return False
    folder_name = existing_folders[clean_proj]
    excel_path = get_safe_filename(os.path.join(BASE_DIR, folder_name), pname)
    if not os.path.exists(excel_path):
        return False
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["销控汇总明细"]
        for r in range(3, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=col).value for col in range(1, 11)]
            if not any(row_vals) or not row_vals[0]:
                break
            global_units.append({
                '项目ID': pid,
                '项目名称': pname,
                '区域': region,
                '商圈': district,
                '开发商': developer,
                '楼栋名称': row_vals[0],
                '楼层': row_vals[1],
                '房号': row_vals[2],
                '户型': row_vals[3],
                '实用面积': row_vals[4],
                '销控状态': row_vals[5],
                '成交日期': row_vals[6] if row_vals[6] else '-',
                '总价': row_vals[7],
                '呎价': row_vals[8],
                '是否招标': row_vals[9]
            })
        wb.close()
        return True
    except Exception as e:
        print(f"  警告: 载入本地备份数据异常 ({e})。")
    return False



# ==========================================
# 6. 主程序逻辑
# ==========================================
def is_hki_kln_district(district_str):
    """判断商圈是否属于港岛或九龙（过滤新界项目）"""
    if not district_str:
        return False
    for kw in ALL_VALID_KEYWORDS:
        if kw in district_str:
            return True
    return False


def fetch_midland_projects(api_headers):
    """
    实时从美联 API 获取港岛+九龙全量项目列表。
    使用 region=HKI/KLN 参数过滤，并以商圈关键词做二次过滤确保准确。
    返回 dict: {project_id: {'name':简体名, 'region':区域, 'district':商圈, 'id':pid}}
    """
    projects = {}
    region_map = {'HKI': '港岛', 'KLN': '九龙'}
    for region_code, region_cn in region_map.items():
        offset = 0
        limit = 200
        while True:
            try:
                api_headers['User-Agent'] = random.choice(USER_AGENTS)
                r = requests.get(
                    f"{MIDLAND_API_BASE}/search/v2/new-properties",
                    params={'region': region_code, 'limit': limit, 'offset': offset,
                            'language': 'zh-cn', 'order_by': 'order'},
                    headers=api_headers, timeout=15
                )
                if r.status_code != 200:
                    print(f"  美联 {region_cn} 项目列表获取失败: HTTP {r.status_code}")
                    break
                data = r.json()
                items = data.get('result', [])
                if not items:
                    break
                for p in items:
                    pid = p.get('id') or p.get('phase_id') or p.get('project_id')
                    name_raw = (p.get('name') or p.get('name_sc') or
                                p.get('project_name') or '')
                    name = t2s(name_raw)
                    # 提取商圈
                    dist_raw = p.get('district') or p.get('area') or ''
                    if isinstance(dist_raw, dict):
                        dist_raw = dist_raw.get('name_sc') or dist_raw.get('name') or ''
                    elif isinstance(p.get('location'), dict):
                        dist_raw = p['location'].get('district', '') or dist_raw
                    district = t2s(str(dist_raw or ''))

                    if pid and name:
                        if district:
                            # 有商圈信息：用商圈判断是否港岛/九龙，过滤新界
                            if not is_hki_kln_district(district):
                                continue  # 跳过新界项目
                            # 根据商圈关键词校正 region
                            has_hki = any(kw in district for kw in HKI_DISTRICT_KEYWORDS)
                            has_kln = any(kw in district for kw in KLN_DISTRICT_KEYWORDS)
                            if has_hki and not has_kln:
                                actual_region = '港岛'
                            elif has_kln and not has_hki:
                                actual_region = '九龙'
                            else:
                                actual_region = region_cn  # 默认信任 region_code
                        else:
                            # 商圈为空：直接信任 region_code 参数
                            actual_region = region_cn
                        projects[pid] = {'id': pid, 'name': name,
                                         'region': actual_region, 'district': district}
                offset += limit
                total = data.get('total', 0)
                if total == 0 or offset >= total:
                    break
                time.sleep(random.uniform(0.4, 1.2))
            except Exception as e:
                print(f"  美联项目列表请求异常: {e}")
                break
    return projects


def fetch_hkp_projects():
    """
    实时从 HKP API 获取港岛+九龙全量项目列表。
    返回 set: {project_id} 以及 dict: {clean_name: pid}
    """
    # 先从页面获取 HKP token
    hkp_token = None
    for ua in random.sample(USER_AGENTS, 3):
        try:
            r = requests.get('https://www.hkp.com.hk/zh-hk/list/new-property/',
                             headers={'User-Agent': ua}, timeout=15)
            if r.status_code == 200:
                m = re.search(r'"userToken"\s*:\s*"([^"]+)"', r.text)
                if m:
                    hkp_token = m.group(1)
                    break
        except Exception:
            pass

    if not hkp_token:
        print("  警告: 无法获取 HKP Token，HKP 项目列表将为空")
        return set(), {}

    hkp_headers = {
        'Authorization': f'Bearer {hkp_token}',
        'User-Agent': random.choice(USER_AGENTS),
        'Origin': 'https://www.hkp.com.hk',
        'Referer': 'https://www.hkp.com.hk/'
    }

    pid_set = set()
    name_map = {}  # clean_name -> pid
    for region_code in ['HKI', 'KLN']:
        offset = 0
        limit = 200
        while True:
            try:
                hkp_headers['User-Agent'] = random.choice(USER_AGENTS)
                r = requests.get(
                    'https://data.hkp.com.hk/search/v2/new-properties',
                    params={'region': region_code, 'limit': limit, 'offset': offset,
                            'language': 'zh-hk', 'order_by': 'order'},
                    headers=hkp_headers, timeout=15
                )
                if r.status_code != 200:
                    break
                data = r.json()
                items = data.get('result', [])
                if not items:
                    break
                for p in items:
                    pid = p.get('id') or p.get('phase_id') or p.get('project_id')
                    name_raw = (p.get('name') or p.get('name_sc') or
                                p.get('project_name') or '')
                    name = t2s(name_raw)
                    if pid:
                        pid_set.add(pid)
                    if name:
                        name_map[clean_name(name)] = pid
                offset += limit
                if offset >= data.get('total', 0):
                    break
                time.sleep(random.uniform(0.4, 1.2))
            except Exception as e:
                print(f"  HKP 项目列表请求异常: {e}")
                break
    return pid_set, name_map


def find_midland_exclusive(midland_projects, hkp_pid_set, hkp_name_map):
    """
    对比美联和 HKP 项目列表，返回美联独有的项目列表。
    双重匹配机制：
    1. 项目 ID 直接匹配
    2. 清洗后项目名称匹配
    """
    exclusive = []
    for pid, proj in midland_projects.items():
        # 1. ID 直接匹配
        if pid in hkp_pid_set:
            continue
        cn = clean_name(proj['name'])
        # 2. 名称与 HKP 项目列表匹配
        if cn in hkp_name_map:
            continue
        exclusive.append(proj)
    return exclusive


def main():
    print("=" * 50)
    print("  美联独有新盘补充抓取脚本 v2.0 (动态对比版)")
    print("=" * 50)

    # 步骤 1: 获取美联 Token
    token = fetch_midland_token()
    api_headers = {
        'User-Agent': random.choice(USER_AGENTS),
        'Authorization': f'Bearer {token}',
        'Origin': 'https://www.midland.com.hk',
        'Referer': 'https://www.midland.com.hk/'
    }

    # 步骤 2: 实时拉取美联全量项目列表
    print("\n[1/4] 正在拉取美联项目列表...")
    midland_projects = fetch_midland_projects(api_headers)
    print(f"  美联港岛+九龙共 {len(midland_projects)} 个项目")

    # 步骤 3: 实时拉取 HKP 全量项目列表
    print("\n[2/4] 正在拉取 HKP 项目列表...")
    hkp_pid_set, hkp_name_map = fetch_hkp_projects()
    print(f"  HKP 港岛+九龙共 {len(hkp_pid_set)} 个项目")

    # 步骤 4: 对比找出差集并合并静态独有项目
    print("\n[3/4] 正在对比差集并合并静态独有项目...")
    exclusive_projects = find_midland_exclusive(midland_projects, hkp_pid_set, hkp_name_map)
    
    # 合并静态列表中的独有项目
    scraped_ids = {p['id'] for p in exclusive_projects}
    for spid, sproj in STATIC_MIDLAND_PROJECTS.items():
        if spid not in scraped_ids:
            exclusive_projects.append({
                'id': spid,
                'name': sproj['name'],
                'region': sproj['region'],
                'district': sproj['district']
            })

    print(f"  美联独有项目 (含静态追加): {len(exclusive_projects)} 个")
    for p in exclusive_projects:
        print(f"    - [{p['id']}] {p['name']} ({p['region']}-{p['district']})")

    if not exclusive_projects:
        print("\n  两家平台项目完全一致，无需补充抓取。")
        return

    # 步骤 5: 扫描已有文件夹
    existing_folders = {}
    for d in os.listdir(BASE_DIR):
        path = os.path.join(BASE_DIR, d)
        if os.path.isdir(path) and d.startswith(('港岛-', '九龙-')):
            parts = d.split('-', 2)
            if len(parts) >= 3:
                key = clean_name(parts[2])
                existing_folders[key] = d

    # 预载本地数据库历史成交作为备份
    db_tx_lookup = {}
    db_path = os.path.join(BASE_DIR, "成交历史数据库.db")
    if os.path.exists(db_path):
        try:
            import sqlite3
            conn_db = sqlite3.connect(db_path)
            c_db = conn_db.cursor()
            c_db.execute("SELECT project_name, building_name, floor, flat, sold_date FROM sold_history WHERE sold_date IS NOT NULL AND sold_date != '-' AND sold_date != ''")
            for r_db in c_db.fetchall():
                pn_db, bn_db, fl_db, ft_db, sd_db = r_db
                cp_db = clean_name(pn_db)
                nb_db = normalize_bname(bn_db)
                db_tx_lookup[(cp_db, nb_db, str(fl_db).strip(), str(ft_db).strip())] = sd_db[:10]
            conn_db.close()
            print(f"  成功从本地数据库预载成交历史留存记录 {len(db_tx_lookup)} 条。")
        except Exception as e_db:
            print(f"  预载本地数据库成交历史记录提示: {e_db}")

    # 步骤 6: 准备全局汇总数据
    global_units = []
    failed_projects = []
    success_count = 0
    total = len(exclusive_projects)
    print(f"\n[4/4] 开始抓取 {total} 个美联独有项目...")

    # 6.4 逐项目抓取
    for idx, proj in enumerate(exclusive_projects):
        pid    = proj['id']
        pname  = proj['name']
        region = proj['region']
        district = proj['district']

        print(f"\n({idx+1}/{total}) 正在处理: {pname} (ID: {pid})")

        # 6.4.1 预载历史成交记录
        tx_lookup = {}
        try:
            api_headers['User-Agent'] = random.choice(USER_AGENTS)
            tx_res = requests.get(
                f"{MIDLAND_API_BASE}/search/v2/transactions",
                params={'phase_ids': pid, 'limit': 2000},
                headers=api_headers, timeout=12
            )
            if tx_res.status_code == 200:
                tx_data = tx_res.json().get('result', [])
                for tx in tx_data:
                    tx_bname = normalize_bname(tx.get('building', {}).get('name'))
                    tx_floor = str(tx.get('floor', '')).strip()
                    tx_flat  = str(tx.get('flat', '')).strip()
                    tx_date_raw = tx.get('tx_date')
                    if tx_date_raw:
                        tx_lookup[(tx_bname, tx_floor, tx_flat)] = tx_date_raw[:10]
                print(f"  已载入历史成交纪录 {len(tx_lookup)} 条。")
        except Exception as e:
            print(f"  提示: 预载历史成交纪录失败 ({e})。")

        # 6.4.2 获取项目详情（楼栋列表）
        try:
            api_headers['User-Agent'] = random.choice(USER_AGENTS)
            detail_res = requests.get(
                f"{MIDLAND_API_BASE}/info/v1/new-properties/{pid}",
                headers=api_headers, timeout=12
            )
            if detail_res.status_code != 200:
                print(f"  警告: 无法获取项目详情 (HTTP {detail_res.status_code})，跳过。")
                failed_projects.append(pname)
                continue
            detail_data = detail_res.json()
        except Exception as e:
            print(f"  警告: 请求项目详情异常 ({e})，跳过。")
            failed_projects.append(pname)
            continue

        # 检查本地是否已售罄，若售罄则直接加载本地备份，不再爬取
        developer = t2s(detail_data.get('developer', ''))
        clean_proj_name = clean_name(pname)
        excel_filename = None
        if clean_proj_name in existing_folders:
            folder_name = existing_folders[clean_proj_name]
            excel_filename = get_safe_filename(os.path.join(BASE_DIR, folder_name), pname)

        if excel_filename and is_project_sold_out(excel_filename):
            print(f"  [已售罄] 该项目去化率已达 100%，无需重复爬取，直接沿用并合并本地数据。")
            if load_existing_project_data(pname, pid, region, district, developer, existing_folders, global_units):
                success_count += 1
            else:
                failed_projects.append(pname)
            continue

        # 收集楼栋信息
        buildings_map = {}
        for b in detail_data.get('buildings', []):
            bid = b.get('id')
            bname = t2s(b.get('name', ''))
            if bid:
                buildings_map[bid] = bname
            for sub_b in b.get('buildings', []):
                sub_bid = sub_b.get('id')
                sub_bname = t2s(sub_b.get('name', ''))
                if sub_bid:
                    buildings_map[sub_bid] = sub_bname

        # 从 floorplan 补充楼栋
        for fp in detail_data.get('floorplan', []):
            b_info = fp.get('building', {})
            bid = b_info.get('id')
            bname = t2s(b_info.get('name', ''))
            if bid and bid not in buildings_map:
                buildings_map[bid] = bname

        # 从历史成交补充
        for tx in tx_lookup.keys():
            pass  # tx_lookup key is (bname, floor, flat) — buildings by ID sourced above

        sell_status = detail_data.get('sell_status', '')
        print(f"  楼栋数: {len(buildings_map)} | 售卖状态: {sell_status}")

        if not buildings_map:
            print(f"  提示: [{sell_status}] 该项目暂无楼栋/单位数据（可能未开售）。")
            if sell_status not in ('coming_soon',):
                failed_projects.append(pname)
            else:
                print(f"  [跳过] 尚未开售项目，无销控数据可抓取。")
            continue

        # 6.4.3 逐楼栋抓取单位销控数据
        project_buildings_data = {}

        for bid, bname in buildings_map.items():
            print(f"    -> 楼栋: {bname} (ID: {bid})")
            try:
                api_headers['User-Agent'] = random.choice(USER_AGENTS)
                units_res = requests.get(
                    f"{MIDLAND_API_BASE}/info/v1/new-property/transactions/buildings/{bid}",
                    headers=api_headers, timeout=12
                )
                if units_res.status_code != 200:
                    print(f"      警告: HTTP {units_res.status_code}，跳过该楼栋。")
                    continue
                units_data = units_res.json()
            except Exception as e:
                print(f"      警告: 请求单位数据异常 ({e})，跳过该楼栋。")
                continue

            units = units_data.get('data', [])
            print(f"      获取到单位数: {len(units)}")

            parsed_units = []
            for u in units:
                floor       = u.get('floor')
                flat        = u.get('flat') or u.get('flat_name') or '-'
                net_area    = u.get('net_area', 0) or 0
                status_raw  = u.get('status', 'pending')
                status_cn   = STATUS_MAP.get(status_raw, '待售')
                is_tender   = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']

                price = u.get('price')
                if price is not None:
                    try:
                        price = float(price)
                    except ValueError:
                        price = None

                unit_price_net = u.get('unit_price_net')
                if unit_price_net is not None:
                    try:
                        price_per_sq_ft = int(float(unit_price_net))
                    except ValueError:
                        price_per_sq_ft = None
                else:
                    price_per_sq_ft = int(round(price / net_area)) if (price and net_area > 0) else None

                # 户型提取
                room_layout = '暂无'
                detail_list = u.get('detail', [])
                if detail_list:
                    room_type = detail_list[0].get('room_type')
                    if room_type is not None and str(room_type) != '':
                        if str(room_type) in ['0', 0]:
                            room_layout = '开放式'
                        else:
                            room_layout = f"{room_type}房"

                # 成交日期
                sold_date_raw = u.get('sold_date') or u.get('tx_date')
                sold_date = '-'
                if status_raw == 'sold':
                    if sold_date_raw:
                        sold_date = sold_date_raw[:10]
                    else:
                        norm_b = normalize_bname(bname)
                        norm_floor = str(floor).strip()
                        norm_flat = str(flat).strip()
                        
                        # 1) 接口历史成交精确匹配
                        mapped_date = tx_lookup.get((norm_b, norm_floor, norm_flat))
                        
                        # 2) 接口历史成交去除"第"/"座"模糊匹配
                        if not mapped_date:
                            clean_b = norm_b.replace('第', '').replace('座', '').strip()
                            for (tb, tf, tfl), d in tx_lookup.items():
                                if tf == norm_floor and tfl == norm_flat:
                                    clean_tb = tb.replace('第', '').replace('座', '').strip()
                                    if clean_tb == clean_b or not clean_b or not clean_tb:
                                        mapped_date = d
                                        break
                                        
                        # 3) 本地历史数据库沉淀补全
                        if not mapped_date and db_tx_lookup:
                            cpname = clean_name(pname)
                            mapped_date = db_tx_lookup.get((cpname, norm_b, norm_floor, norm_flat))
                            if not mapped_date:
                                clean_b = norm_b.replace('第', '').replace('座', '').strip()
                                for (cp, tb, tf, tfl), d in db_tx_lookup.items():
                                    if cp == cpname and tf == norm_floor and tfl == norm_flat:
                                        clean_tb = tb.replace('第', '').replace('座', '').strip()
                                        if clean_tb == clean_b or not clean_b or not clean_tb:
                                            mapped_date = d
                                            break
                                            
                        if mapped_date:
                            sold_date = mapped_date

                unit_info = {
                    'floor': floor, 'flat': flat, 'net_area': net_area,
                    'status_raw': status_raw, 'status': status_cn,
                    'price': price, 'price_per_sq_ft': price_per_sq_ft,
                    'is_tender': u.get('is_tender'),
                    'room_layout': room_layout, 'sold_date': sold_date,
                    'unit_type': u.get('unit_type', '')
                }
                parsed_units.append(unit_info)

                # 汇总全局数据
                global_units.append({
                    '项目ID':   pid,
                    '项目名称': pname,
                    '区域':     region,
                    '商圈':     district,
                    '开发商':   t2s(detail_data.get('developer', '')),
                    '楼栋名称': bname,
                    '楼层':     floor,
                    '房号':     flat,
                    '户型':     room_layout,
                    '实用面积': net_area if net_area > 0 else '暂无',
                    '销控状态': status_cn,
                    '成交日期': sold_date,
                    '总价':    '招标单位' if (is_tender and status_raw != 'sold') else (price if price else '暂无'),
                    '呎价':    '-' if (is_tender and status_raw != 'sold') else (price_per_sq_ft if price_per_sq_ft else '暂无'),
                    '是否招标': '是' if is_tender else '否'
                })

            if parsed_units:
                project_buildings_data[bname] = parsed_units

            # 随机限速，比原本固定 0.3s 更安全
            time.sleep(random.uniform(0.6, 1.8))

        # 6.4.4 写入 Excel
        if project_buildings_data:
            clean_proj = clean_name(pname)
            if clean_proj in existing_folders:
                folder_name = existing_folders[clean_proj]
            else:
                folder_name = f"{region}-{district}-{pname}"
                folder_name = re.sub(r'[\/\\\:\*\?\"<>|]', '', folder_name)

            project_dir = os.path.join(BASE_DIR, folder_name)
            os.makedirs(project_dir, exist_ok=True)

            excel_path = get_safe_filename(project_dir, pname)
            write_project_excel(excel_path, pname, project_buildings_data)
            print(f"  [完成] 写入销控表: {excel_path}")
            success_count += 1
        else:
            print(f"  [提示] 未获取到有效单位数据，跳过 {pname}。")
            failed_projects.append(pname)

        # 随机限速，比原本固定 0.5s 更安全
        time.sleep(random.uniform(1.5, 3.5))

    # 6.5 将本次抓取的新项目数据合并更新到全局汇总表
    print(f"\n{'='*50}")
    print(f"本次成功抓取: {success_count}/{total} 个项目")
    if failed_projects:
        print(f"失败/跳过项目: {failed_projects}")

    if global_units:
        global_excel_path = os.path.join(BASE_DIR, "香港一手新盘销控汇总.xlsx")
        print(f"\n正在将 {len(global_units)} 条新数据合并到全局汇总表...")
        _append_to_global_excel(global_excel_path, global_units)
        print(f"[完成] 全局汇总表已更新: {global_excel_path}")
    else:
        print("\n[提示] 无新单位数据，全局汇总表未更新。")

    print("=" * 50)
    print("补充抓取完成！")


def _append_to_global_excel(global_path, new_units):
    """
    将新抓取的数据追加到全局汇总表。
    若文件存在，则在末尾追加行；若不存在，则新建。
    """
    if not os.path.exists(global_path):
        print("  全局汇总表不存在，将新建文件。")
        _write_global_excel_new(global_path, new_units)
        return

    try:
        wb = openpyxl.load_workbook(global_path)
    except Exception as e:
        print(f"  警告: 无法打开全局汇总表 ({e})，将新建。")
        _write_global_excel_new(global_path, new_units)
        return

    ws = wb.active
    font_data   = Font(name='Microsoft YaHei', size=9)
    align_ctr   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )

    # 找到第一个空行
    last_row = ws.max_row
    # 检查是否已有这些项目，避免重复
    existing_keys = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[1]:  # 项目名称在第2列
            existing_keys.add((str(row[1]), str(row[7] or ''), str(row[8] or '')))  # (pname, floor, flat)

    added = 0
    for u in new_units:
        key = (str(u['项目名称']), str(u['楼层'] or ''), str(u['房号'] or ''))
        if key in existing_keys:
            continue  # 已存在，跳过

        last_row += 1
        row_vals = [
            u['区域'], u['项目名称'], u['商圈'], u['开发商'],
            u['楼栋名称'], u['楼层'], u['房号'], u['户型'],
            u['实用面积'], u['销控状态'], u['成交日期'],
            u['总价'], u['呎价'], u['是否招标']
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws.cell(row=last_row, column=col_idx, value=val)
            c.font = font_data
            c.alignment = align_ctr
            c.border = border_thin

        existing_keys.add(key)
        added += 1

    wb.save(global_path)
    print(f"  成功追加 {added} 条新单位数据到全局汇总表。")


def _write_global_excel_new(global_path, units):
    """新建全局汇总表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全局销控汇总"
    headers = ["区域", "项目名称", "商圈", "开发商", "楼栋名称", "楼层", "房号", "户型",
               "实用面积 (平方呎)", "销控状态", "成交日期", "总价 (港币)", "实用呎价 (港币/呎)", "是否招标"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for row_idx, u in enumerate(units, 2):
        vals = [u['区域'], u['项目名称'], u['商圈'], u['开发商'], u['楼栋名称'],
                u['楼层'], u['房号'], u['户型'], u['实用面积'], u['销控状态'],
                u['成交日期'], u['总价'], u['呎价'], u['是否招标']]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=c, value=v)
    wb.save(global_path)


if __name__ == '__main__':
    main()
