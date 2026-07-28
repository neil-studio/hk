#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
import json
import time
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# 1. 简繁转换辅助函数与映射
# ==========================================
# 输出文件基础目录定位到“价单”文件夹
BASE_DIR = "/Users/nb/google/Antigravity/工作/运营/价单"
# 简繁转换对照文件仍引用“楼盘字典”下的原文件
zh2hans_path = "/Users/nb/google/Antigravity/工作/运营/楼盘字典/zh2hans.json"

if os.path.exists(zh2hans_path):
    with open(zh2hans_path, 'r', encoding='utf-8') as f:
        zh2hans_dict = json.load(f)
    # 只提取单字符映射，防止包含多字词汇导致的错位对齐映射
    char_map = {k: v for k, v in zh2hans_dict.items() if len(k) == 1 and len(v) == 1}
    TRADITIONAL_CHARS = "".join(char_map.keys())
    SIMPLIFIED_CHARS  = "".join(char_map.values())
else:
    # 备用简繁字符映射
    TRADITIONAL_CHARS = "港島九龍啟德灣仔堅尼地城筲箕灣紅磡鰂魚涌鴨脷洲黃竹坑東半山西半山中半山西九龍東九龍壽臣山山頂淺水灣深水灣跑马地大坑道司徒拔道渣甸山薄扶林香港仔上環中環土瓜灣小西湾西营盘铜锣湾天后鲗鱼涌康怡太古湾仔跑马地大坑渣甸山北角半山石塘咀坚尼地城摩星岭薄扶林香港仔黄竹坑深水湾浅水湾赤柱大潭石澳峯滙"
    SIMPLIFIED_CHARS  = "港岛九龙启德湾仔坚尼地城筲箕湾红磡鲗鱼涌鸭脷洲黄竹坑东半山西半山中半山西九龙东九龙寿臣山山顶浅水湾深水湾跑马地大坑道司徒拔道渣甸山薄扶林香港仔上环中环土瓜湾小西湾西营盘铜锣湾天后鲗鱼涌康怡太古湾仔跑马地大坑渣甸山北角半山石塘咀坚尼地城摩星岭薄扶林香港仔黄竹坑深水湾浅水湾赤柱大潭石澳峰汇"

TRANS_MAP = str.maketrans(TRADITIONAL_CHARS, SIMPLIFIED_CHARS)

def t2s(text):
    if not text:
        return ""
    if isinstance(text, dict):
        return {k: t2s(v) for k, v in text.items()}
    if isinstance(text, list):
        return [t2s(x) for x in text]
    return str(text).translate(TRANS_MAP)

def clean_name(name):
    """标准化名称比对，去除空格、特殊字符并转为小写简体"""
    if not name:
        return ""
    name = t2s(name).lower()
    name = re.sub(r'[\s\-\(\)\（\）\.\,\，\。]', '', name)
    return name

def get_safe_filename(project_dir, pname):
    """生成去除特殊字符的安全 Excel 文件名"""
    safe_pname = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '', pname)
    return os.path.join(project_dir, f"{safe_pname}_销控明细表.xlsx")

def normalize_bname(name):
    """标准化楼栋名称以进行比对映射"""
    if not name:
        return ""
    name = t2s(name).strip()
    name = re.sub(r'^第', '', name)
    return name

# ==========================================
# 2. 状态映射与颜色配置
# ==========================================
STATUS_MAP = {
    'pending': '待售',
    'priced': '已定价未售',
    'sale': '在售',
    'sold': '已售',
    'stopped': '暂停销售'
}

COLORS = {
    'sale': {'bg': 'A9F5A9', 'fg': '004D00'},
    'sold': {'bg': 'F5A9A9', 'fg': '660000'},
    'priced': {'bg': 'A9D0F5', 'fg': '002060'},
    'stopped': {'bg': 'FFC000', 'fg': '000000'},
    'pending': {'bg': 'FFFFFF', 'fg': '7F7F7F'}
}

# 销控单元格颜色填充配置 (Openpyxl Fills)
FILLS = {k: PatternFill(start_color=v['bg'], end_color=v['bg'], fill_type='solid') for k, v in COLORS.items()}

# ==========================================
# 3. 网络请求基础配置与 Token 获取
# ==========================================
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

def fetch_user_token():
    """从主列表页的 __NEXT_DATA__ 中解析最新的 Bearer Token"""
    url = "https://www.hkp.com.hk/zh-hk/list/new-property/"
    print("正在从主页获取最新 Token...")
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        match = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>', r.text)
        if match:
            data = json.loads(match.group(1))
            token = data.get('props', {}).get('pageProps', {}).get('userToken')
            if token:
                print(f"Token 获取成功! 长度: {len(token)}")
                return token
        raise ValueError("__NEXT_DATA__ 中未找到 userToken")
    except Exception as e:
        print(f"错误: 无法获取 Token: {e}", file=sys.stderr)
        sys.exit(1)

# ==========================================
# 4. 全局目录扫描与映射
# ==========================================
def scan_existing_folders():
    """扫描已有文件夹（包含价单和原楼盘字典目录），以防止命名拼写差异导致创建重复文件夹"""
    mapping = {}
    folders_to_scan = [BASE_DIR, "/Users/nb/google/Antigravity/工作/运营/楼盘字典"]
    for folder in folders_to_scan:
        if not os.path.exists(folder):
            continue
        for d in os.listdir(folder):
            path = os.path.join(folder, d)
            if os.path.isdir(path):
                parts = d.split('-')
                if len(parts) >= 3:
                    proj_name = parts[-1]
                    mapping[clean_name(proj_name)] = d
    return mapping

def load_existing_project_data(pname, pid, region, district, developer, existing_folders, global_units):
    """
    如果抓取失败，尝试从已有的本地 Excel 文件中加载数据并合并到 global_units 中。
    """
    clean_proj_name = clean_name(pname)
    if clean_proj_name not in existing_folders:
        return False
        
    folder_name = existing_folders[clean_proj_name]
    project_dir = os.path.join(BASE_DIR, folder_name)
    excel_filename = get_safe_filename(project_dir, pname)
    
    if not os.path.exists(excel_filename):
        return False
        
    print(f"  [历史沿用] 正在从本地历史数据加载: {excel_filename}")
    try:
        # 使用 read_only=True 加速读取
        wb = openpyxl.load_workbook(excel_filename, read_only=True, data_only=True)
        if "销控汇总明细" in wb.sheetnames:
            ws = wb["销控汇总明细"]
            count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                # 检查是否是空行 (第一列"楼栋"为空则视为结束)
                if not row or row[0] is None:
                    break
                
                global_units.append({
                    '项目ID': pid,
                    '项目名称': pname,
                    '区域': region,
                    '商圈': district,
                    '开发商': developer,
                    '楼栋名称': row[0],
                    '楼层': row[1],
                    '房号': row[2],
                    '户型': row[3],
                    '实用面积': row[4],
                    '销控状态': row[5],
                    '成交日期': row[6],
                    '总价': row[7],
                    '呎价': row[8],
                    '是否招标': row[9]
                })
                count += 1
            print(f"  [成功] 从本地历史数据恢复了 {count} 条单位数据。")
            wb.close()
            return True
        wb.close()
    except Exception as e:
        print(f"  [警告] 读取本地历史 Excel 失败: {e}")
    return False

def is_project_sold_out(excel_filename):
    """
    检查已有的本地 Excel 文件是否显示该项目已售罄 (100%)。
    """
    if not os.path.exists(excel_filename):
        return False
    try:
        wb = openpyxl.load_workbook(excel_filename, read_only=True, data_only=True)
        if "销控汇总明细" in wb.sheetnames:
            ws = wb["销控汇总明细"]
            total = 0
            sold = 0
            for row in ws.iter_rows(min_row=3, max_col=6, values_only=True):
                if not row or row[0] is None:
                    break
                total += 1
                status_str = str(row[5]).strip() if len(row) >= 6 and row[5] is not None else ""
                if status_str == '已售' or ('售' in status_str and '未售' not in status_str and '待售' not in status_str and '暂停' not in status_str):
                    sold += 1
            wb.close()
            return total > 0 and sold == total
        wb.close()
    except Exception as e:
        print(f"  [警告] 读取本地 Excel 判断是否售罄时失败: {e}")
    return False


def extract_floor_number(floor_str):
    """从楼层字符串提取用于数字排序的数值"""
    if not floor_str:
        return -99
    floor_str = str(floor_str).upper().strip()
    match = re.search(r'(\d+)', floor_str)
    if match:
        val = int(match.group(1))
        if 'R' in floor_str:
            val += 0.5
        return val
    if 'G' in floor_str or 'UG' in floor_str:
        return 0
    if 'B' in floor_str:
        return -1
    return -99

# ==========================================
# 5. Excel 写入辅助逻辑 (使用 Openpyxl)
# ==========================================
def write_project_excel(filepath, project_name, buildings_data):
    """
    为单个项目写入 Excel 销控表，包含：
    - Tab 1: "销控汇总明细" (扁平表形式，包含所有楼栋的所有单位)
    - 之后的 Tabs: 每个楼栋单独一页，采用 "销控网格图 (Grid)" 形式展示
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------
    # Tab 1: 销控汇总明细
    # ----------------------------------------
    ws_detail = wb.active
    ws_detail.title = "销控汇总明细"
    
    font_title = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    font_data = Font(name='Microsoft YaHei', size=10)
    align_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 写标题
    ws_detail.merge_cells("A1:J1")
    ws_detail['A1'] = f"{project_name} - 一手销控汇总明细"
    ws_detail['A1'].font = font_title
    ws_detail['A1'].alignment = align_center
    ws_detail.row_dimensions[1].height = 40
    
    # 写表头 (包含了户型和成交日期)
    headers = ["楼栋", "楼层", "房号", "户型", "实用面积 (平方呎)", "销控状态", "成交日期", "总价 (港币)", "实用呎价 (港币/呎)", "是否招标"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=2, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_thin
    ws_detail.row_dimensions[2].height = 25

    # 填充明细数据
    row_idx = 3
    for bname, units in buildings_data.items():
        sorted_units = sorted(units, key=lambda x: (extract_floor_number(x['floor']), x['flat']), reverse=True)
        for u in sorted_units:
            ws_detail.cell(row=row_idx, column=1, value=bname)
            ws_detail.cell(row=row_idx, column=2, value=u['floor'])
            ws_detail.cell(row=row_idx, column=3, value=u['flat'])
            ws_detail.cell(row=row_idx, column=4, value=u['room_layout'])
            ws_detail.cell(row=row_idx, column=5, value=u['net_area'] if u['net_area'] > 0 else '暂无')
            ws_detail.cell(row=row_idx, column=6, value=u['status'])
            ws_detail.cell(row=row_idx, column=7, value=u['sold_date'])
            
            is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
            
            # 如果是招标且未售，总价显示招标
            if is_tender and u['status_raw'] != 'sold':
                ws_detail.cell(row=row_idx, column=8, value='招标单位')
                ws_detail.cell(row=row_idx, column=9, value='-')
            else:
                ws_detail.cell(row=row_idx, column=8, value=u['price'] if u['price'] else '暂无')
                ws_detail.cell(row=row_idx, column=9, value=u['price_per_sq_ft'] if u['price_per_sq_ft'] else '暂无')
                
            ws_detail.cell(row=row_idx, column=10, value='是' if is_tender else '否')
            
            for col in range(1, 11):
                c = ws_detail.cell(row=row_idx, column=col)
                c.font = font_data
                c.alignment = align_center
                c.border = border_thin
                
                # 数字格式化
                if col == 8 and isinstance(c.value, (int, float)):
                    c.number_format = '$#,##0'
                elif col == 9 and isinstance(c.value, (int, float)):
                    c.number_format = '$#,##0'
                elif col == 5 and isinstance(c.value, (int, float)):
                    c.number_format = '#,##0'

            row_idx += 1

    # 自动列宽
    for col in ws_detail.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_detail.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 设置打印自适应：单页宽度，高度自动（纵向打印）
    ws_detail.sheet_properties.pageSetUpPr.fitToPage = True
    ws_detail.page_setup.fitToWidth = 1
    ws_detail.page_setup.fitToHeight = 0
    ws_detail.page_setup.orientation = ws_detail.ORIENTATION_PORTRAIT
    ws_detail.page_setup.paperSize = ws_detail.PAPERSIZE_A4

    # 设置窄页边距 (0.25 英寸)
    ws_detail.page_margins.left = 0.25
    ws_detail.page_margins.right = 0.25
    ws_detail.page_margins.top = 0.25
    ws_detail.page_margins.bottom = 0.25
    ws_detail.page_margins.header = 0.1
    ws_detail.page_margins.footer = 0.1

    # ----------------------------------------
    # Tab 2+: 每个楼栋生成可视化销控网格 (Grid)
    # ----------------------------------------
    multi_buildings = {}
    villa_buildings = {}
    
    if buildings_data:
        for bname, units in buildings_data.items():
            # 判断是否为独栋 (如果该楼栋单位数 <= 2 或是 HOUSE 类型)
            has_house_type = any(u.get('unit_type') == 'HOUSE' for u in units)
            if len(units) <= 2 or has_house_type:
                villa_buildings[bname] = units
            else:
                multi_buildings[bname] = units

    # 1. 多个单位 of 楼栋，每个楼栋放在单独的工作表
    for bname, units in multi_buildings.items():
        _write_building_grid(wb, bname, units, font_data, font_header, font_title, align_center, border_thin, fill_header)

    # 2. 独栋楼，集合放在一个工作表
    if villa_buildings:
        _write_villa_grid(wb, project_name, villa_buildings, font_data, align_center, border_thin)

    wb.save(filepath)

def update_global_excel_with_retry(filepath, new_units):
    """
    更新全局汇总 Excel，替换并追加重试项目的最新数据（采用极速读取与重写机制）。
    """
    if not os.path.exists(filepath):
        write_global_excel(filepath, new_units)
        return
        
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        replaced_pids = set(u['项目ID'] for u in new_units)
        
        # 1. 保留未被替换的行数据
        keep_rows = []
        max_r = ws.max_row
        
        # 从第3行开始读取
        for r in range(3, max_r + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 17)]
            # 检查空行
            if not any(row_vals):
                continue
            # 判断项目ID是否在需要替换的集合中
            cell_pid = row_vals[1] # 第二列是项目ID
            if cell_pid not in replaced_pids:
                keep_rows.append(row_vals)
                
        wb.close()
        
        # 2. 重新组合所有数据 (保留的行 + 新重试的行)
        combined_rows = []
        # 先放入原先保留的行
        for idx, row in enumerate(keep_rows, 1):
            row[0] = idx # 重新生成序号
            combined_rows.append(row)
            
        # 再追加新重试的行
        start_idx = len(combined_rows) + 1
        for idx, d in enumerate(new_units, start_idx):
            combined_rows.append([
                idx,
                d['项目ID'],
                d['项目名称'],
                d['区域'],
                d['商圈'],
                d['开发商'],
                d['楼栋名称'],
                d['楼层'],
                d['房号'],
                d['户型'],
                d['实用面积'],
                d['销控状态'],
                d['成交日期'],
                d['总价'],
                d['呎价'],
                d['是否招标']
            ])
            
        # 3. 完全重建全局汇总表以获得极高写入速度
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "一手新盘销控汇总"
        
        # 写入表头 (和 write_global_excel 保持一致)
        headers = [
            "序号", "项目ID", "项目名称", "区域", "商圈", "开发商", 
            "楼栋名称", "楼层", "房号", "户型", "实用面积 (平方呎)", 
            "销控状态", "成交日期", "总价 (港币)", "实用呎价 (港币/呎)", "是否招标"
        ]
        
        # 标题行
        font_title = Font(name='Microsoft YaHei', size=14, bold=True, color='FFFFFF')
        fill_title = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        
        new_ws.merge_cells("A1:P1")
        title_cell = new_ws.cell(row=1, column=1, value="香港一手新盘销控汇总分析表")
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = align_center
        new_ws.row_dimensions[1].height = 40
        
        # 表头行
        font_header = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        
        new_ws.row_dimensions[2].height = 28
        for col_idx, text in enumerate(headers, 1):
            c = new_ws.cell(row=2, column=col_idx, value=text)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center
            
        # 写入所有行数据
        font_data = Font(name='Microsoft YaHei', size=10)
        border_thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        for r_idx, row_data in enumerate(combined_rows, 3):
            new_ws.row_dimensions[r_idx].height = 20
            for c_idx, val in enumerate(row_data, 1):
                c = new_ws.cell(row=r_idx, column=c_idx, value=val)
                c.font = font_data
                c.alignment = align_center
                c.border = border_thin
                
                # 格式化数值
                if c_idx in [14, 15] and isinstance(c.value, (int, float)):
                    c.number_format = '$#,##0'
                elif c_idx == 11 and isinstance(c.value, (int, float)):
                    c.number_format = '#,##0'
                    
        # 列宽自适应
        for col in new_ws.columns:
            max_len = 0
            for cell in col:
                if cell.row == 1:
                    continue
                val_str = str(cell.value or '')
                # 计算近似字符长度
                length = sum(2 if ord(char) > 256 else 1 for char in val_str)
                if length > max_len:
                    max_len = length
            col_letter = get_column_letter(col[0].column)
            new_ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        # 设置页面自适应打印
        new_ws.sheet_properties.pageSetUpPr.fitToPage = True
        new_ws.page_setup.orientation = new_ws.ORIENTATION_PORTRAIT
        new_ws.page_setup.paperSize = new_ws.PAPERSIZE_A4
        new_ws.page_setup.fitToWidth = 1
        new_ws.page_setup.fitToHeight = 0
        
        new_ws.page_margins.left = 0.25
        new_ws.page_margins.right = 0.25
        new_ws.page_margins.top = 0.25
        new_ws.page_margins.bottom = 0.25
        new_ws.page_margins.header = 0.1
        new_ws.page_margins.footer = 0.1
        
        new_wb.save(filepath)
        print(f"  [成功] 采用极速重写机制，已更新并追加 {len(new_units)} 条新重试数据到全局汇总表中。")
    except Exception as e:
        print(f"  [错误] 极速重载更新全局汇总 Excel 失败: {e}")

def write_global_excel(filepath, data_list):
    """写入全局汇总 Excel 数据"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一手新盘销控汇总"
    
    font_title = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    font_data = Font(name='Microsoft YaHei', size=10)
    align_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 写标题
    ws.merge_cells("A1:P1")
    ws['A1'] = "香港港岛、九龙一手新盘单位销控明细汇总表"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 40

    # 写表头
    headers = ["序号", "项目ID", "项目名称", "区域", "商圈", "开发商", "楼栋名称", "楼层", "房号", "户型", "实用面积 (平方呎)", "销控状态", "成交日期", "单位总价 (港币)", "实用呎价 (港币/呎)", "是否招标"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_thin
    ws.row_dimensions[2].height = 25

    # 填充数据
    row_idx = 3
    for idx, d in enumerate(data_list, 1):
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=d['项目ID'])
        ws.cell(row=row_idx, column=3, value=d['项目名称'])
        ws.cell(row=row_idx, column=4, value=d['区域'])
        ws.cell(row=row_idx, column=5, value=d['商圈'])
        ws.cell(row=row_idx, column=6, value=d['开发商'])
        ws.cell(row=row_idx, column=7, value=d['楼栋名称'])
        ws.cell(row=row_idx, column=8, value=d['楼层'])
        ws.cell(row=row_idx, column=9, value=d['房号'])
        ws.cell(row=row_idx, column=10, value=d['户型'])
        ws.cell(row=row_idx, column=11, value=d['实用面积'])
        ws.cell(row=row_idx, column=12, value=d['销控状态'])
        ws.cell(row=row_idx, column=13, value=d['成交日期'])
        ws.cell(row=row_idx, column=14, value=d['总价'])
        ws.cell(row=row_idx, column=15, value=d['呎价'])
        ws.cell(row=row_idx, column=16, value=d['是否招标'])

        # 数据行格式化与样式
        for col in range(1, 17):
            c = ws.cell(row=row_idx, column=col)
            c.font = font_data
            c.alignment = align_center
            c.border = border_thin
            
            # 数字格式
            if col == 14 and isinstance(c.value, (int, float)):
                c.number_format = '$#,##0'
            elif col == 15 and isinstance(c.value, (int, float)):
                c.number_format = '$#,##0'
            elif col == 11 and isinstance(c.value, (int, float)):
                c.number_format = '#,##0'

        row_idx += 1

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filepath)

# ==========================================
# 6. 主程序逻辑
# ==========================================
def main():
    token = fetch_user_token()
    api_headers = {
        'User-Agent': HEADERS['User-Agent'],
        'Authorization': f'Bearer {token}',
        'Origin': 'https://www.hkp.com.hk',
        'Referer': 'https://www.hkp.com.hk/'
    }

    # 6.1 获取项目列表 (如果是重试模式，直接从本地文件载入失败项目)
    is_retry = "--retry" in sys.argv
    failed_projects_file = os.path.join(BASE_DIR, "failed_projects.json")

    if is_retry:
        print("\n==========================================")
        print("      正在以 [重试模式] 重新运行爬取        ")
        print("==========================================")
        if not os.path.exists(failed_projects_file):
            print("提示: 未找到失败项目记录文件 failed_projects.json，无需重试。")
            return
        try:
            with open(failed_projects_file, 'r', encoding='utf-8') as f:
                filtered_projects = json.load(f)
            print(f"成功载入需要重试的项目数: {len(filtered_projects)}")
            if not filtered_projects:
                print("提示: 失败项目记录为空，无需重试。")
                return
        except Exception as e:
            print(f"错误: 载入失败项目记录文件异常: {e}")
            return
    else:
        print("\n[Step 1] 正在请求项目列表 API...")
        try:
            r = requests.get("https://data.hkp.com.hk/search/v2/new-properties", params={'limit': 1000}, headers=api_headers, timeout=15)
            r.raise_for_status()
            projects_data = r.json()
        except Exception as e:
            print(f"错误: 无法请求项目列表 API: {e}", file=sys.stderr)
            return

        all_projects = projects_data.get('result', [])
        print(f"API 返回项目总数: {len(all_projects)}")

        # 过滤出港岛和九龙项目
        target_regions = ['港岛', '九龙']
        filtered_projects = []
        for p in all_projects:
            region_raw = p.get('region', {}).get('name', '')
            region_cn = t2s(region_raw)
            if region_cn in target_regions:
                filtered_projects.append({
                    'id': p.get('id'),
                    'name': t2s(p.get('name')),
                    'region': region_cn,
                    'district': t2s(p.get('district')),
                    'developer': t2s(p.get('developer', {}).get('name', ''))
                })

        print(f"港岛/九龙一手新盘数量: {len(filtered_projects)}")

    # 确保保存价单的主目录存在
    os.makedirs(BASE_DIR, exist_ok=True)

    # 扫描已有文件夹映射 (包含价单目录和原楼盘字典目录，复用文件夹名)
    existing_folders = scan_existing_folders()
    
    # 汇总数据列表，用于生成全局汇总表
    global_units = []
    
    # 用于记录和报告未成功抓取最新数据的项目
    restored_projects = []
    totally_failed_projects = []

    # 6.2 循环遍历每个项目
    for idx, proj in enumerate(filtered_projects):
        pid = proj['id']
        pname = proj['name']
        region = proj['region']
        district = proj['district']
        
        # 检查本地是否已售罄，若售罄则不再爬取
        clean_proj_name = clean_name(pname)
        excel_filename = None
        if clean_proj_name in existing_folders:
            folder_name = existing_folders[clean_proj_name]
            excel_filename = get_safe_filename(os.path.join(BASE_DIR, folder_name), pname)
            
        if excel_filename and is_project_sold_out(excel_filename):
            print(f"\n({idx+1}/{len(filtered_projects)}) 正在处理: {pname} (ID: {pid}, 区域: {region}-{district})")
            print(f"  [已售罄] 该项目去化率已达 100%，无需重复爬取，直接沿用并合并本地数据。")
            load_existing_project_data(pname, pid, region, district, proj.get('developer', ''), existing_folders, global_units)
            continue
            
        print(f"\n({idx+1}/{len(filtered_projects)}) 正在处理: {pname} (ID: {pid}, 区域: {region}-{district})")

        # 6.2.1 预载全局历史成交记录以作日期备用补充
        tx_lookup = {}
        try:
            tx_url = "https://data.hkp.com.hk/search/v2/transactions"
            tx_params = {'phase_ids': pid, 'limit': 2000}
            tx_res = requests.get(tx_url, params=tx_params, headers=api_headers, timeout=12)
            if tx_res.status_code == 200:
                tx_data = tx_res.json().get('result', [])
                for tx in tx_data:
                    tx_bname = normalize_bname(tx.get('building', {}).get('name'))
                    tx_floor = str(tx.get('floor', '')).strip()
                    tx_flat = str(tx.get('flat', '')).strip()
                    tx_date_raw = tx.get('tx_date')
                    if tx_date_raw:
                        tx_lookup[(tx_bname, tx_floor, tx_flat)] = tx_date_raw[:10]
                print(f"  成功载入已登记的一手成交纪录 {len(tx_lookup)} 条。")
        except Exception as e:
            print(f"  提示: 预载一手历史成交纪录异常 ({e})。将只使用销控接口默认日期。")

        # 6.2.2 获取项目详情（以提取楼栋列表）
        try:
            detail_res = requests.get(f"https://data.hkp.com.hk/info/v1/new-properties/{pid}", headers=api_headers, timeout=10)
            if detail_res.status_code != 200:
                print(f"  警告: 无法获取项目详情 (Code: {detail_res.status_code})。")
                if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                    restored_projects.append(pname)
                else:
                    totally_failed_projects.append(pname)
                continue
            detail_data = detail_res.json()
        except Exception as e:
            print(f"  警告: 获取项目详情异常 ({e})。")
            if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                restored_projects.append(pname)
            else:
                totally_failed_projects.append(pname)
            continue

        # 6.2.2 收集所有可能的候选楼栋 ID 和名称 (结合 standard buildings, floorplan 以及交易记录以防错配)
        buildings_map = {}
        
        # 1) 标准 buildings
        for b in detail_data.get('buildings', []):
            bid = b.get('id')
            bname = t2s(b.get('name'))
            if bid:
                buildings_map[bid] = bname
            for sub_b in b.get('buildings', []):
                sub_bid = sub_b.get('id')
                sub_bname = t2s(sub_b.get('name'))
                if sub_bid:
                    buildings_map[sub_bid] = sub_bname

        # 2) floorplan
        for fp in detail_data.get('floorplan', []):
            b_info = fp.get('building', {})
            bid = b_info.get('id')
            bname = t2s(b_info.get('name'))
            if bid:
                buildings_map[bid] = bname

        # 3) 交易数据
        for tx in tx_data:
            b_info = tx.get('building', {})
            bid = b_info.get('id')
            bname = t2s(b_info.get('name'))
            if bid:
                # 交易数据中的楼栋名称通常非常准确
                buildings_map[bid] = bname

        if not buildings_map:
            print("  提示: 该项目没有关联楼栋信息。")
            if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                restored_projects.append(pname)
            else:
                totally_failed_projects.append(pname)
            continue

        print(f"  包含候选楼栋数: {len(buildings_map)}")
        
        # 存储当前项目的所有楼栋销控数据，用以生成项目独立明细表
        project_buildings_data = {}

        # 6.2.3 循环拉取每个楼栋的销控数据
        for bid, bname in buildings_map.items():
            print(f"    -> 楼栋: {bname} (ID: {bid})")

            try:
                units_res = requests.get(f"https://data.hkp.com.hk/info/v1/new-property/transactions/buildings/{bid}", headers=api_headers, timeout=10)
                if units_res.status_code != 200:
                    print(f"      警告: 无法拉取该楼栋的单位明细。")
                    continue
                units_data = units_res.json()
            except Exception as e:
                print(f"      警告: 请求楼栋单位明细异常 ({e})。")
                continue

            units = units_data.get('data', [])
            print(f"      获取到单位数: {len(units)}")
            
            parsed_units = []
            for u in units:
                unit_id = u.get('unit_id')
                floor = u.get('floor')
                flat = u.get('flat') or u.get('flat_name') or '-'
                net_area = u.get('net_area', 0) or 0
                status_raw = u.get('status', 'pending')
                status_cn = STATUS_MAP.get(status_raw, '待售')
                is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
                
                # 获取总价
                price = u.get('price')
                if price is not None:
                    try:
                        price = float(price)
                    except ValueError:
                        price = None
                
                # 计算呎价
                unit_price_net = u.get('unit_price_net')
                if unit_price_net is not None:
                    try:
                        price_per_sq_ft = int(float(unit_price_net))
                    except ValueError:
                        price_per_sq_ft = None
                else:
                    if price and net_area > 0:
                        price_per_sq_ft = int(round(price / net_area))
                    else:
                        price_per_sq_ft = None

                # 1. 提取户型 (n房)
                room_layout = '暂无'
                detail_list = u.get('detail', [])
                if detail_list:
                    room_type = detail_list[0].get('room_type')
                    if room_type is not None and str(room_type) != '':
                        if str(room_type) in ['0', 0]:
                            room_layout = '开放式'
                        else:
                            room_layout = f"{room_type}房"
                
                # 2. 提取成交日期 (YYYY-MM-DD)，若销控接口返回 None，则使用历史登记记录作为补充
                sold_date_raw = u.get('sold_date') or u.get('tx_date')
                sold_date = '-'
                if status_raw == 'sold':
                    if sold_date_raw:
                        sold_date = sold_date_raw[:10]
                    else:
                        # 尝试从预载的成交记录中查找并映射
                        norm_b = normalize_bname(bname)
                        norm_floor = str(floor).strip()
                        norm_flat = str(flat).strip()
                        mapped_date = tx_lookup.get((norm_b, norm_floor, norm_flat))
                        if mapped_date:
                            sold_date = mapped_date

                unit_info = {
                    'unit_id': unit_id,
                    'floor': floor,
                    'flat': flat,
                    'net_area': net_area,
                    'status_raw': status_raw,
                    'status': status_cn,
                    'price': price,
                    'price_per_sq_ft': price_per_sq_ft,
                    'is_tender': u.get('is_tender'),
                    'room_layout': room_layout,
                    'sold_date': sold_date
                }
                parsed_units.append(unit_info)

                # 添加到全局汇总列表
                global_units.append({
                    '项目ID': pid,
                    '项目名称': pname,
                    '区域': region,
                    '商圈': district,
                    '开发商': proj['developer'],
                    '楼栋名称': bname,
                    '楼层': floor,
                    '房号': flat,
                    '户型': room_layout,
                    '实用面积': net_area if net_area > 0 else '暂无',
                    '销控状态': status_cn,
                    '成交日期': sold_date,
                    '总价': '招标单位' if (is_tender and status_raw != 'sold') else (price if price else '暂无'),
                    '呎价': '-' if (is_tender and status_raw != 'sold') else (price_per_sq_ft if price_per_sq_ft else '暂无'),
                    '是否招标': '是' if is_tender else '否'
                })

            project_buildings_data[bname] = parsed_units
            
            # 短暂限速
            time.sleep(0.3)

        # 5.2.3 如果有数据，生成该项目的独立明细表和文件夹
        if project_buildings_data:
            # 匹配或创建文件夹
            clean_proj_name = clean_name(pname)
            if clean_proj_name in existing_folders:
                folder_name = existing_folders[clean_proj_name]
            else:
                folder_name = f"{region}-{district}-{pname}"
                folder_name = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '', folder_name)
                
            project_dir = os.path.join(BASE_DIR, folder_name)
            os.makedirs(project_dir, exist_ok=True)
            
            excel_filename = get_safe_filename(project_dir, pname)
            write_project_excel(excel_filename, pname, project_buildings_data)
            print(f"  [完成] 写入项目销控表: {excel_filename}")
        else:
            print("  [提示] 抓取的新数据为空，尝试沿用历史数据。")
            if load_existing_project_data(pname, pid, region, district, proj['developer'], existing_folders, global_units):
                restored_projects.append(pname)
            else:
                totally_failed_projects.append(pname)

        # 限速
        time.sleep(0.5)

    # 5.3 写入或更新全局汇总表
    if global_units:
        global_excel_path = os.path.join(BASE_DIR, "香港一手新盘销控汇总.xlsx")
        if is_retry:
            update_global_excel_with_retry(global_excel_path, global_units)
        else:
            write_global_excel(global_excel_path, global_units)
            print(f"\n[完成] 写入全局汇总表: {global_excel_path} (共 {len(global_units)} 条单位数据)")
    else:
        print("\n[警告] 未成功抓取到任何单位数据，全局汇总表未更新。")

    # 记录未抓取成功的项目到本地 json，供重试流程读取
    all_failed = restored_projects + totally_failed_projects
    if all_failed:
        failed_dicts = []
        for name in all_failed:
            for p in filtered_projects:
                if p['name'] == name:
                    failed_dicts.append(p)
                    break
        try:
            with open(failed_projects_file, 'w', encoding='utf-8') as f:
                json.dump(failed_dicts, f, ensure_ascii=False, indent=2)
            print(f"已更新失败项目列表: {failed_projects_file} (共 {len(failed_dicts)} 个项目)")
        except Exception as e:
            print(f"写入失败项目记录文件失败: {e}")
    else:
        if os.path.exists(failed_projects_file):
            try:
                os.remove(failed_projects_file)
                print(f"所有项目均已成功，已清理历史失败记录文件。")
            except Exception as e:
                pass

    # 5.4 打印并输出未抓取到数据的项目统计报告
    from datetime import datetime
    print("\n==========================================")
    print("      未抓取到最新数据的项目统计报告        ")
    print("==========================================")
    
    total_skipped = len(restored_projects) + len(totally_failed_projects)
    print(f"共有 {total_skipped} 个项目在本次抓取中未获取到最新数据：\n")
    
    if restored_projects:
        print(f"🔹 以下 {len(restored_projects)} 个项目拉取失败，但已【成功恢复并沿用】历史数据：")
        for pname in sorted(restored_projects):
            print(f"  - {pname}")
        print()
        
    if totally_failed_projects:
        print(f"❌ 以下 {len(totally_failed_projects)} 个项目拉取失败，且【无历史数据】可用：")
        for pname in sorted(totally_failed_projects):
            print(f"  - {pname}")
        print()
        
    # 将未抓取到的项目列表写入本地文件，方便查看
    report_path = os.path.join(BASE_DIR, "未抓取到最新数据项目列表.txt")
    try:
        with open(report_path, 'w', encoding='utf-8') as rf:
            rf.write("==========================================\n")
            rf.write("      未抓取到最新数据的项目统计报告        \n")
            rf.write(f"      报告时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            rf.write("==========================================\n\n")
            
            rf.write(f"共有 {total_skipped} 个项目未获取到最新数据。\n\n")
            
            rf.write(f"【已恢复并沿用历史数据的项目】({len(restored_projects)}个):\n")
            for pname in sorted(restored_projects):
                rf.write(f"  - {pname}\n")
            rf.write("\n")
            
            rf.write(f"【完全未获取到数据且无历史数据的项目】({len(totally_failed_projects)}个):\n")
            for pname in sorted(totally_failed_projects):
                rf.write(f"  - {pname}\n")
                
        print(f"未抓取数据项目报告已写入本地文件: {report_path}")
    except Exception as e:
        print(f"写入报告报告失败: {e}")
    print("==========================================\n")

if __name__ == '__main__':
    main()


def _write_building_grid(wb, bname, units, font_data, font_header, font_title, align_center, border_thin, fill_header):
    """为单个楼栋写入销控网格 Tab"""
    safe_tab = re.sub(r'[\\/?*\[\]:]', '', bname)[:31]
    ws = wb.create_sheet(title=safe_tab)

    # 统计面板数据
    total = len(units)
    cnt = {s: 0 for s in STATUS_MAP.keys()}
    for u in units:
        sr = u.get('status_raw', 'pending')
        if sr in cnt:
            cnt[sr] += 1

    # 汇总统计面板 (Row 1 标题，Row 2-3 数据)
    stat_labels = [
        ('总套数', str(total), 'FFFFFF', '000000', False),
        ('在售 (Sale)', str(cnt.get('sale', 0)), 'A9F5A9', '004D00', True),
        ('已定价未售', str(cnt.get('priced', 0)), 'A9D0F5', '002060', False),
        ('已售 (Sold)', str(cnt.get('sold', 0)), 'F5A9A9', '660000', False),
        ('暂停销售', str(cnt.get('stopped', 0)), 'FFC000', '000000', True),
        ('待售 (Pending)', str(cnt.get('pending', 0)), 'FFFFFF', '7F7F7F', False),
    ]

    # Row 1: 楼栋名称
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(stat_labels))
    ws['A1'] = f"{bname} 销控网格图"
    ws['A1'].font = Font(name='Microsoft YaHei', size=13, bold=True)
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 28

    # Row 2: 图例标签
    for col, (label, val, bg, fg, bold) in enumerate(stat_labels, 1):
        c2 = ws.cell(row=2, column=col, value=label)
        c2.font = Font(name='Microsoft YaHei', size=9, bold=bold, color=fg)
        c2.alignment = align_center
        c2.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        c2.border = border_thin
        # Row 3: 数值
        c3 = ws.cell(row=3, column=col, value=val)
        c3.font = Font(name='Microsoft YaHei', size=10, bold=True, color=fg)
        c3.alignment = align_center
        c3.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        c3.border = border_thin
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # 整理楼层、单位
    floors_set = set()
    flats_set = set()
    for u in units:
        floors_set.add(u['floor'])
        flats_set.add(u['flat'])

    floors_sorted = sorted(floors_set, key=extract_floor_number, reverse=True)
    flats_sorted  = sorted(flats_set)

    # 建立单位查找表 (floor, flat) -> unit
    unit_lookup = {(u['floor'], u['flat']): u for u in units}

    # 网格表头 (Row 4 = 单位列)
    ws.cell(row=4, column=1, value='楼层').font = Font(name='Microsoft YaHei', size=9, bold=True)
    ws.cell(row=4, column=1).alignment = align_center
    ws.cell(row=4, column=1).fill = fill_header
    ws.cell(row=4, column=1).border = border_thin
    ws.row_dimensions[4].height = 18

    for col_idx, flat in enumerate(flats_sorted, 2):
        c = ws.cell(row=4, column=col_idx, value=flat)
        c.font = Font(name='Microsoft YaHei', size=9, bold=True)
        c.alignment = align_center
        c.fill = fill_header
        c.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.column_dimensions['A'].width = 8

    # 数据行
    for row_idx, floor in enumerate(floors_sorted, 5):
        ws.cell(row=row_idx, column=1, value=floor)
        ws.cell(row=row_idx, column=1).font = Font(name='Microsoft YaHei', size=9, bold=True)
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=1).fill = fill_header
        ws.cell(row=row_idx, column=1).border = border_thin
        ws.row_dimensions[row_idx].height = 58

        for col_idx, flat in enumerate(flats_sorted, 2):
            u = unit_lookup.get((floor, flat))
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = border_thin
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if u is None:
                c.value = ''
                c.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                continue

            status_raw = u.get('status_raw', 'pending')
            color_cfg = COLORS.get(status_raw, COLORS['pending'])
            bg = color_cfg['bg']
            fg = color_cfg['fg']
            is_bold = status_raw in ('sale', 'stopped')

            # 4行文字内容
            line1 = f"{flat} | {u['net_area']}呎 ({u['room_layout']})"
            is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
            if is_tender and status_raw != 'sold':
                line2 = "招标单位"
                line3 = "-"
            else:
                if u['price']:
                    price_wan = round(u['price'] / 10000)
                    line2 = f"${price_wan}万"
                else:
                    line2 = "-"
                if u['price_per_sq_ft']:
                    line3 = f"${u['price_per_sq_ft']:,}/呎"
                else:
                    line3 = "-"

            if status_raw == 'sold' and u['sold_date'] and u['sold_date'] != '-':
                try:
                    parts = u['sold_date'].split('-')
                    yr = parts[0][-2:]
                    mo = parts[1]
                    line4 = f"({yr}年-{mo}月)"
                except:
                    line4 = f"({u['sold_date']})"
            else:
                status_display = STATUS_MAP.get(status_raw, '待售')
                line4 = f"({status_display})"

            c.value = f"{line1}\n{line2}\n{line3}\n{line4}"
            c.font = Font(name='Microsoft YaHei', size=8, bold=is_bold, color=fg)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

    # 打印设置
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    for margin in ['left', 'right', 'top', 'bottom']:
        setattr(ws.page_margins, margin, 0.25)
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def _write_villa_grid(wb, project_name, buildings_data, font_data, align_center, border_thin):
    """
    独立屋项目整合为一个单一的工作表，避免分割成无数个 1x1 标签页。
    采用 5 行文字的网格卡片式布局。
    """
    ws_grid = wb.create_sheet(title="独立屋销控表")
    
    font_title  = Font(name='Microsoft YaHei', size=14, bold=True)
    font_header = Font(name='Microsoft YaHei', size=10, bold=True)
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 展平所有独立屋单位并打上楼栋标记
    all_units = []
    for bname, units in buildings_data.items():
        for u in units:
            u_copy = u.copy()
            u_copy['_bname'] = bname
            all_units.append(u_copy)
            
    # 基于命名模式分析其网格排布 (行列定位)
    parentheses_matches = sum(1 for u in all_units if re.search(r'^.+?\s*[\(\（].+?[\)\）]$', u['_bname']))
    space_matches = sum(1 for u in all_units if ' ' in u['_bname'].strip())
    
    has_parentheses = parentheses_matches >= len(all_units) * 0.7
    has_space = space_matches >= len(all_units) * 0.7
    
    # 预设分配网格坐标
    for idx, u in enumerate(all_units):
        bname = u['_bname']
        if has_parentheses:
            m = re.search(r'^(.+?)\s*[\(\（](.+?)[\)\）]$', bname)
            if m:
                u['_grid_row'] = m.group(1).strip()
                u['_grid_col'] = m.group(2).strip()
            else:
                u['_grid_row'] = "其他"
                u['_grid_col'] = bname
        elif has_space:
            parts = bname.rsplit(' ', 1)
            u['_grid_row'] = parts[0].strip()
            u['_grid_col'] = parts[1].strip()
        else:
            # 默认横排扁平
            if len(all_units) <= 8:
                u['_grid_row'] = "独立屋"
                u['_grid_col'] = bname
            else:
                # 超过8个，默认后面按 cols_per_row = 6 排布分配
                pass
                
    # 按照大楼/别墅名进行自然排序
    def natural_key(x):
        return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', x['_bname'])]
        
    all_units = sorted(all_units, key=natural_key)
    
    # 如果是超过8个的扁平排布，在此根据自然排序顺序分配行列坐标
    if not has_parentheses and not has_space and len(all_units) > 8:
        cols_per_row = 6
        for idx, u in enumerate(all_units):
            row_num = idx // cols_per_row + 1
            col_num = idx % cols_per_row + 1
            u['_grid_row'] = f"第 {row_num} 行"
            u['_grid_col'] = f"位置 {col_num}"

    # 整理出确切的网格行、列头
    floors = sorted(list(set(u['_grid_row'] for u in all_units)))
    
    # 列头自然排序
    def col_natural_key(text):
        return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', text)]
    flats = sorted(list(set(u['_grid_col'] for u in all_units)), key=col_natural_key)
    
    unit_map = {(u['_grid_row'], u['_grid_col']): u for u in all_units}
    
    # 统计数量
    total_units = len(all_units)
    status_counts = {'sale': 0, 'sold': 0, 'priced': 0, 'stopped': 0, 'pending': 0}
    for u in all_units:
        st = u.get('status_raw', 'pending')
        status_counts[st] = status_counts.get(st, 0) + 1

    # 设置标题
    ws_grid.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(flats) + 1)
    title_cell = ws_grid.cell(row=1, column=1, value=f"{project_name} - 独立屋销控网格图")
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws_grid.row_dimensions[1].height = 40

    # 展示统计面板数据
    ws_grid.row_dimensions[2].height = 20
    ws_grid.row_dimensions[3].height = 25
    
    stat_cols = [
        ("总套数", total_units, 'pending'),
        ("在售 (Sale)", status_counts['sale'], 'sale'),
        ("已定价未售", status_counts['priced'], 'priced'),
        ("已售 (Sold)", status_counts['sold'], 'sold'),
        ("暂停销售", status_counts['stopped'], 'stopped'),
        ("待售 (Pending)", status_counts['pending'], 'pending')
    ]
    
    for stat_idx, (label, val, fill_key) in enumerate(stat_cols, 1):
        cell_lbl = ws_grid.cell(row=2, column=stat_idx, value=label)
        cell_lbl.font = Font(name='Microsoft YaHei', size=9, bold=(fill_key in ['sale', 'stopped']), color=COLORS[fill_key]['fg'])
        cell_lbl.alignment = align_center
        cell_lbl.fill = FILLS[fill_key]
        cell_lbl.border = border_thin
        
        cell_val = ws_grid.cell(row=3, column=stat_idx, value=f"{val} 套")
        cell_val.font = Font(name='Microsoft YaHei', size=11, bold=True, color=COLORS[fill_key]['fg'])
        cell_val.alignment = align_center
        cell_val.fill = FILLS[fill_key]
        cell_val.border = border_thin

    # 空一行后写表头
    header_row = 5
    ws_grid.row_dimensions[header_row].height = 25
    cell_cross = ws_grid.cell(row=header_row, column=1, value="类型/位置" if not has_parentheses else "类型/屋号")
    cell_cross.font = font_header
    cell_cross.alignment = align_center
    cell_cross.border = border_thin
    cell_cross.fill = fill_header

    for col_idx, flat_name in enumerate(flats, 2):
        cell_flat = ws_grid.cell(row=header_row, column=col_idx, value=flat_name)
        cell_flat.font = font_header
        cell_flat.alignment = align_center
        cell_flat.border = border_thin
        cell_flat.fill = fill_header

    # 写入网格数据
    curr_row = header_row + 1
    for f_name in floors:
        ws_grid.row_dimensions[curr_row].height = 80 # 提高以容纳5行文字
        
        cell_floor = ws_grid.cell(row=curr_row, column=1, value=f_name)
        cell_floor.font = font_header
        cell_floor.alignment = align_center
        cell_floor.border = border_thin
        cell_floor.fill = fill_header
        
        for col_idx, flat_name in enumerate(flats, 2):
            u = unit_map.get((f_name, flat_name))
            cell_unit = ws_grid.cell(row=curr_row, column=col_idx)
            cell_unit.border = border_thin
            cell_unit.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if u:
                status_raw = u['status_raw']
                status_cn = u['status']
                price = u['price']
                net_area = u['net_area']
                price_per_sq_ft = u['price_per_sq_ft']
                is_tender = u.get('is_tender') is True or str(u.get('is_tender')).lower() in ['true', '1']
                sold_date = u.get('sold_date')
                room_layout = u.get('room_layout', '暂无')
                
                # 1. 洋房名称/编号
                line1 = u['_bname']
                
                # 2. 实用面积与户型房数
                area_str = f"{net_area}呎" if net_area > 0 else "暂无"
                layout_str = f" ({room_layout})" if room_layout != '暂无' else ""
                line2 = f"{area_str}{layout_str}"
                
                # 3. 销售状态 (已售写具体登记年份-月份)
                if status_raw == 'sold' and sold_date and sold_date != '-':
                    try:
                        parts = sold_date.split('-')
                        yr = parts[0][-2:]
                        mo = parts[1]
                        line3 = f"({yr}年-{mo}月)"
                    except:
                        line3 = f"({sold_date})"
                else:
                    line3 = f"({status_cn})"
                    
                # 4. 价格 (过亿单位显示 $X.XX亿)
                if is_tender and status_raw != 'sold':
                    line4 = "招标项目" if status_raw == 'sale' else "招标单位"
                else:
                    if price and price > 0:
                        if price >= 100000000:
                            line4 = f"${price/100000000:.2f}亿"
                        elif price >= 10000:
                            line4 = f"${price/10000:,.0f}万"
                        else:
                            line4 = f"${price:,}"
                    else:
                        line4 = "暂无价格"
                        
                # 5. 实用呎价
                if is_tender and status_raw != 'sold':
                    line5 = "-"
                elif price_per_sq_ft and price_per_sq_ft > 0:
                    line5 = f"${price_per_sq_ft:,}/呎"
                else:
                    line5 = "暂无呎价"
                    
                cell_text = f"{line1}\n{line2}\n{line3}\n{line4}\n{line5}"
                cell_unit.value = cell_text
                
                cell_unit.font = Font(name='Microsoft YaHei', size=8, bold=(status_raw in ['sale', 'stopped']), color=COLORS.get(status_raw, COLORS['pending'])['fg'])
                cell_unit.fill = FILLS.get(status_raw, FILLS['pending'])
            else:
                cell_unit.value = ""
                cell_unit.fill = PatternFill(start_color='E5E5E5', end_color='E5E5E5', fill_type='solid')
                
        curr_row += 1

    # 设置列宽
    ws_grid.column_dimensions['A'].width = 12
    for col_idx in range(2, len(flats) + 2):
        col_letter = get_column_letter(col_idx)
        ws_grid.column_dimensions[col_letter].width = 18

    # 启用单页打印自适应设置
    ws_grid.sheet_properties.pageSetUpPr.fitToPage = True
    ws_grid.page_setup.fitToWidth = 1
    ws_grid.page_setup.fitToHeight = 1
    ws_grid.page_setup.orientation = ws_grid.ORIENTATION_LANDSCAPE
    ws_grid.page_setup.paperSize = ws_grid.PAPERSIZE_A4

    # 设置窄页边距 (0.25 英寸)
    for margin in ['left', 'right', 'top', 'bottom']:
        setattr(ws_grid.page_margins, margin, 0.25)
    ws_grid.page_margins.header = 0.1
    ws_grid.page_margins.footer = 0.1

