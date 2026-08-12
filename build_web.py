#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import json
import glob
import shutil
import urllib.parse
from datetime import datetime
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WEB_DIR = os.path.join(BASE_DIR, "web")
FILES_DIR = os.path.join(WEB_DIR, "files")
ROOT_FILES_DIR = os.path.join(BASE_DIR, "files")

def ensure_dirs():
    """确保网页输出目录和文件存储目录存在，并清空旧的 files 目录以防遗留售罄项目"""
    os.makedirs(WEB_DIR, exist_ok=True)
    for d in [FILES_DIR, ROOT_FILES_DIR]:
        if os.path.exists(d):
            shutil.rmtree(d)
        os.makedirs(d, exist_ok=True)

def parse_project_stats(file_path):
    """
    通过只读模式打开 Excel，读取'销控汇总明细'表，计算项目销控状态统计及最低在售价。
    """
    stats = {
        'total': 0,
        'sold': 0,
        'sale': 0,
        'priced': 0,
        'stopped': 0,
        'pending': 0,
        'sold_rate': 0.0,
        'min_price': None
    }
    
    if not os.path.exists(file_path):
        return stats
        
    try:
        # 使用 read_only=True 极速读取
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if "销控汇总明细" in wb.sheetnames:
            ws = wb["销控汇总明细"]
            row_count = 0
            min_p_val = float('inf')
            for row in ws.iter_rows(min_row=3, values_only=True):
                # 检查是否是空行 (第一列"楼栋"为空则视为结束)
                if not row or row[0] is None:
                    break
                
                status_val = row[5] if len(row) >= 6 else None
                if status_val:
                    status_str = str(status_val).strip()
                    row_count += 1
                    
                    if status_str == '已售':
                        stats['sold'] += 1
                    elif status_str in ('在售', 'sale'):
                        stats['sale'] += 1
                    elif status_str in ('已定价未售', 'priced'):
                        stats['priced'] += 1
                    elif status_str in ('暂停销售', 'stopped'):
                        stats['stopped'] += 1
                    elif status_str in ('待售', 'pending'):
                        stats['pending'] += 1
                    else:
                        if '已售' in status_str:
                            stats['sold'] += 1
                        else:
                            stats['pending'] += 1

                    if status_str in ('在售', '已定价未售', 'sale', 'priced'):
                        disc_price = row[10] if len(row) >= 11 else (row[7] if len(row) >= 8 else None)
                        if disc_price:
                            try:
                                p_num = float(str(disc_price).replace(',', '').replace('$', ''))
                                if p_num > 0 and p_num < min_p_val:
                                    min_p_val = p_num
                            except: pass
            
            stats['total'] = row_count
            if stats['total'] > 0:
                stats['sold_rate'] = round((stats['sold'] / stats['total']) * 100, 1)
            if min_p_val != float('inf'):
                stats['min_price'] = int(min_p_val)
        wb.close()
    except Exception as e:
        print(f"警告: 读取 Excel {file_path} 统计信息失败: {e}")
        
    return stats

def fetch_hkp_status_map():
    """从 HKP 接口拉取项目级销售状态 (sell_status & sell_status_detail) 及基础元数据"""
    status_map = {}
    try:
        import sys, requests, urllib3
        urllib3.disable_warnings()
        sys.path.append(BASE_DIR)
        import scrape_hkp_sales_control as scraper
        
        token = scraper.fetch_user_token()
        headers = {'Authorization': f'Bearer {token}'}
        r = requests.get('https://data.hkp.com.hk/search/v2/new-properties?limit=1000', headers=headers, verify=False, timeout=10)
        if r.status_code == 200:
            for p in r.json().get('result', []):
                pname_raw = p.get('name', '')
                pname = scraper.t2s(pname_raw) if pname_raw else ''
                st = p.get('sell_status', 'on_sale')
                st_detail = p.get('sell_status_detail', {})
                st_cn = st_detail.get('name') if isinstance(st_detail, dict) else '出售中'
                reg_obj = p.get('region', {})
                reg_name = reg_obj.get('name') if isinstance(reg_obj, dict) else '九龙'
                reg_cn = scraper.t2s(reg_name) if reg_name else '九龙'
                dist_name = scraper.t2s(p.get('district', ''))
                tot_unit = p.get('total_unit', 0)
                tot_sold = p.get('total_sold', 0)
                tot_sale = p.get('total_sale', 0)

                item_info = {
                    'name': pname,
                    'sell_status': st,
                    'sell_status_cn': st_cn,
                    'region': reg_cn,
                    'district': dist_name,
                    'total_unit': tot_unit,
                    'total_sold': tot_sold,
                    'total_sale': tot_sale,
                    'developer': scraper.t2s(p.get('developer', {}).get('name', '')) if isinstance(p.get('developer'), dict) else ''
                }
                # 仅保留简体中文规范名，避免生成繁简重复项目
                if pname and pname not in status_map:
                    status_map[pname] = item_info
    except Exception as e:
        print(f"提示: 获取 HKP 状态映射失败: {e}，将使用项目统计保底。")
    return status_map

def load_custom_districts():
    custom_dist_file = os.path.join(BASE_DIR, "custom_districts.json")
    if os.path.exists(custom_dist_file):
        try:
            with open(custom_dist_file, 'r', encoding='utf-8') as f:
                d = json.load(f)
                print(f"成功载入 {len(d)} 条用户自定义商圈规则。")
                return d
        except Exception as e:
            print(f"读取 custom_districts.json 失败: {e}")
    return {}

def build_excel_unit_layout_map():
    """扫描全量项目 Excel 销控明细表，构建 (项目名称, 楼栋, 楼层, 房号) -> 精确户型 映射字典"""
    unit_map = {}
    area_map = {}
    excel_files = glob.glob(os.path.join(BASE_DIR, "*/*_销控明细表.xlsx"))
    for fpath in excel_files:
        folder = os.path.dirname(fpath)
        folder_name = os.path.basename(folder)
        pname = folder_name.split('-')[-1].strip() if '-' in folder_name else folder_name
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fpath, data_only=True)
            if '销控汇总明细' in wb.sheetnames:
                ws = wb['销控汇总明细']
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or len(row) < 5: continue
                    bname, floor, flat, layout, area = str(row[0] or '').strip(), str(row[1] or '').strip(), str(row[2] or '').strip(), str(row[3] or '').strip(), row[4]
                    if pname and layout and layout not in ('None', '', '暂无', '-'):
                        if bname and floor and flat:
                            unit_map[(pname, bname, floor, flat)] = layout
                        try:
                            area_int = int(float(str(area)))
                            if area_int > 0:
                                area_map[(pname, area_int)] = layout
                        except: pass
        except Exception:
            pass
    return unit_map, area_map

def load_real_history_analytics(custom_districts=None, valid_new_project_names=None):
    """从 SQLite 数据库 成交历史数据库.db 抽取 3.7万+ 条真实历史成交数据并按项目、年、月、周聚合 (仅限一手新房白名单)"""
    db_path = os.path.join(BASE_DIR, "成交历史数据库.db")
    if not os.path.exists(db_path):
        print("提示: 未找到 成交历史数据库.db，无法生成真实成交分析。")
        return {}

    if custom_districts is None:
        custom_districts = load_custom_districts()

    try:
        import sqlite3
        from datetime import datetime
        excel_unit_map, excel_area_map = build_excel_unit_layout_map()
        print(f"成功构建 Excel 官方房源户型词典，包含 {len(excel_unit_map)} 套物理房源户型。")

        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute('SELECT project_name, district, region, sold_date, price, disc_price, unit_price, disc_unit_price, building_name, area, floor, flat, layout FROM sold_history WHERE sold_date IS NOT NULL AND sold_date != ""')
        rows = c.fetchall()
        conn.close()

        print(f"成功从 SQLite 抽取 {len(rows)} 条真实成交历史记录。")
        projects_analytics = {}

        def parse_num(val):
            if not val:
                return 0.0
            val_str = str(val).strip()
            if val_str in ('-', '暂无', 'null', 'None', '', '招标单位'):
                return 0.0
            try:
                return float(val_str.replace(',', ''))
            except:
                return 0.0

        def strip_phase(n):
            s = str(n).strip()
            s = re.sub(r'\(第.*?\)', '', s)
            s = re.sub(r'第\s*[0-9A-Za-z\-]+期.*$', '', s)
            s = re.sub(r'Phase\s*[0-9A-Za-z\-]+.*$', '', s, flags=re.IGNORECASE)
            s = re.sub(r'第[I|V|X|A-Z0-9]+期', '', s)
            s = re.sub(r'第[0-9A-Za-z]+$', '', s)
            s = re.sub(r'[0-9]+[a-zA-Z]+$', '', s)
            s = re.sub(r'\s+[0-9]+$', '', s)
            s = re.sub(r'\s+I{1,3}$', '', s)
            s = re.sub(r'\s+II$', '', s)
            s = re.sub(r'\s+III$', '', s)
            s = re.sub(r'\(.*?\)', '', s)
            return s.strip()

        for row in rows:
            pname, dist, reg, sdate, price, dprice, uprice, duprice, bname, area, fl, ft, layout = row
            pname_str = str(pname).strip() if pname else ''
            if not pname_str: continue

            if valid_new_project_names and pname_str not in valid_new_project_names:
                continue

            # 真实成交价与实用呎价 (含保底倒算)
            f_price = parse_num(dprice) or parse_num(price)
            if 10 <= f_price < 100000:
                f_price = f_price * 10000
            f_uprice = parse_num(duprice) or parse_num(uprice)
            f_area = parse_num(area)

            if f_uprice == 0 and f_price > 0 and f_area > 0:
                f_uprice = round(f_price / f_area)

            pname = pname.strip()
            clean_pname = strip_phase(pname)
            target_keys = [pname]
            if '21 Borrett' in pname or '波老道' in pname or '应天' in pname:
                target_keys.append('21 Borrett Road')

            sdate_clean = str(sdate).strip()
            try:
                year = sdate_clean[:4]
                month = sdate_clean[:7]
                dt_parts = [int(p) for p in sdate_clean[:10].split('-')]
                dt_obj = datetime(dt_parts[0], dt_parts[1], dt_parts[2])
                iso_yr, iso_wk, _ = dt_obj.isocalendar()
                week = f"{iso_yr}-W{iso_wk:02d}"
            except:
                continue

            for tkey in target_keys:
                if tkey not in projects_analytics:
                    user_cd = (custom_districts or {}).get(tkey) or (custom_districts or {}).get(clean_pname) or (custom_districts or {}).get(pname) or {}
                    p_reg = user_cd.get('region') or reg
                    p_dist = user_cd.get('district') or dist

                    projects_analytics[tkey] = {
                        'region': p_reg,
                        'district': p_dist,
                        'yearly': {},
                        'monthly': {},
                        'weekly': {},
                        'layouts': {'开放式': 0, '1房': 0, '2房': 0, '3房': 0, '4房+': 0},
                        'price_ranges': {'500万下': 0, '500-1000万': 0, '1000-2000万': 0, '2000-5000万': 0, '5000万+': 0}
                    }

                pa = projects_analytics[tkey]

                # 统计年度
                if year not in pa['yearly']:
                    pa['yearly'][year] = {'volume': 0, 'total_price': 0.0, 'total_uprice': 0.0, 'uprices': []}
                pa['yearly'][year]['volume'] += 1
                pa['yearly'][year]['total_price'] += f_price
                if f_uprice > 0:
                    pa['yearly'][year]['total_uprice'] += f_uprice
                    pa['yearly'][year]['uprices'].append(f_uprice)

                # 统计月度
                if month not in pa['monthly']:
                    pa['monthly'][month] = {'volume': 0, 'total_price': 0.0, 'total_uprice': 0.0, 'uprices': []}
                pa['monthly'][month]['volume'] += 1
                pa['monthly'][month]['total_price'] += f_price
                if f_uprice > 0:
                    pa['monthly'][month]['total_uprice'] += f_uprice
                    pa['monthly'][month]['uprices'].append(f_uprice)

                # 统计周度
                if week:
                    if week not in pa['weekly']:
                        pa['weekly'][week] = {'volume': 0, 'total_price': 0.0, 'total_uprice': 0.0, 'uprices': []}
                    pa['weekly'][week]['volume'] += 1
                    pa['weekly'][week]['total_price'] += f_price
                    if f_uprice > 0:
                        pa['weekly'][week]['total_uprice'] += f_uprice
                        pa['weekly'][week]['uprices'].append(f_uprice)

                # 四级精准联动对号入座户型
                target_layout = None
                fl_str = str(fl).strip() if fl else ''
                ft_str = str(ft).strip() if ft else ''
                b_str = str(bname).strip() if bname else ''

                # 1级：Excel 房号精确匹配
                if b_str and fl_str and ft_str:
                    target_layout = excel_unit_map.get((pname, b_str, fl_str, ft_str)) or excel_unit_map.get((clean_pname, b_str, fl_str, ft_str))

                # 2级：Excel 同盘同面积匹配
                if not target_layout and f_area > 0:
                    area_key = int(round(f_area))
                    target_layout = excel_area_map.get((pname, area_key)) or excel_area_map.get((clean_pname, area_key))

                # 3级：网签原生户型文本
                if not target_layout and layout and str(layout).strip() not in ('None', '', '-', 'null'):
                    target_layout = str(layout).strip()

                l_str = str(target_layout).strip() if target_layout else ''
                if '开放式' in l_str or '开放' in l_str:
                    pa['layouts']['开放式'] += 1
                elif '1房' in l_str or '一房' in l_str:
                    pa['layouts']['1房'] += 1
                elif '2房' in l_str or '两房' in l_str or '二房' in l_str:
                    pa['layouts']['2房'] += 1
                elif '3房' in l_str or '三房' in l_str:
                    pa['layouts']['3房'] += 1
                elif '4房' in l_str or '四房' in l_str or '5房' in l_str or '6房' in l_str:
                    pa['layouts']['4房+'] += 1
                else:
                    # 4级：未匹配保底
                    if f_area > 0:
                        if f_area < 280: pa['layouts']['开放式'] += 1
                        elif f_area < 400: pa['layouts']['1房'] += 1
                        elif f_area < 600: pa['layouts']['2房'] += 1
                        elif f_area < 900: pa['layouts']['3房'] += 1
                        else: pa['layouts']['4房+'] += 1
                    else:
                        pa['layouts']['2房'] += 1

                # 统计总价区间
                if f_price < 5000000: pa['price_ranges']['500万下'] += 1
                elif f_price < 10000000: pa['price_ranges']['500-1000万'] += 1
                elif f_price < 20000000: pa['price_ranges']['1000-2000万'] += 1
                elif f_price < 50000000: pa['price_ranges']['2000-5000万'] += 1
                else: pa['price_ranges']['5000万+'] += 1

        # 计算平均与最大最小值
        for pname, pa in projects_analytics.items():
            for t_dict in [pa['yearly'], pa['monthly'], pa['weekly']]:
                for key, val in t_dict.items():
                    vol = val['volume']
                    val['avg_price'] = round(val['total_price'] / vol, 2) if vol > 0 else 0
                    u_list = val['uprices']
                    val['avg_uprice'] = round(sum(u_list) / len(u_list), 1) if u_list else 0
                    val['min_uprice'] = round(min(u_list), 1) if u_list else 0
                    val['max_uprice'] = round(max(u_list), 1) if u_list else 0
                    del val['total_price']
                    del val['total_uprice']
                    del val['uprices']

        return projects_analytics
    except Exception as e:
        print(f"处理真实成交分析失败: {e}")
        return {}

def build_leaderboard_data(valid_new_project_names=None):
    """从 SQLite 数据库 成交历史数据库.db 生成一手新房动态热销榜单数据 (保留独立期数、仅限一手新盘、剔除无价单位)"""
    db_path = os.path.join(BASE_DIR, "成交历史数据库.db")
    if not os.path.exists(db_path):
        return {}

    custom_dist_file = os.path.join(BASE_DIR, "custom_districts.json")
    custom_districts = {}
    if os.path.exists(custom_dist_file):
        try:
            with open(custom_dist_file, 'r', encoding='utf-8') as f:
                custom_districts = json.load(f)
        except:
            pass

    try:
        import sqlite3
        from datetime import datetime, timedelta
        conn = sqlite3.connect(db_path)
        c = conn.cursor()

        def query_rankings_in_memory(where_sql, region_filter=None, price_min=None, price_max=None, limit=10):
            c.execute(f'''
                SELECT project_name, region, district, price, unit_price, disc_price, disc_unit_price, sold_date, area
                FROM sold_history
                WHERE sold_date IS NOT NULL AND sold_date != '' AND {where_sql}
            ''')
            rows = c.fetchall()
            grouped = {}
            seen_transactions = set()

            for pname, reg, dist, price, uprice, dprice, duprice, sdate, area in rows:
                pname_str = str(pname).strip() if pname else ''
                if not pname_str: continue

                # 核心拦截：仅保留一手新盘项目，严防二手盘混入榜单
                if valid_new_project_names and pname_str not in valid_new_project_names:
                    continue
                
                clean_name = re.sub(r'\(第.*?\)', '', pname_str).strip()
                user_cd = custom_districts.get(pname_str) or custom_districts.get(clean_name) or {}
                reg_val = user_cd.get('region') or reg or '九龙'
                dist_val = user_cd.get('district') or dist or ''
                
                if region_filter and reg_val != region_filter:
                    continue
                    
                final_p = 0
                for pval in [dprice, price]:
                    try:
                        val = float(str(pval).replace(',', ''))
                        if val >= 10:
                            if val < 100000: val = val * 10000 # 智能纠正万元为元
                            final_p = val
                            break
                    except: pass
                    
                if price_min is not None and final_p < price_min: continue
                if price_max is not None and final_p > price_max: continue
                
                final_u = 0
                for uval in [duprice, uprice]:
                    try:
                        val = float(str(uval).replace(',', ''))
                        if val > 1000:
                            final_u = val
                            break
                    except: pass
                
                # 智能防重合指纹比对 (同项目+同日期+同面积+同金额去重)
                dedup_key = (pname_str, str(sdate).strip(), str(area).strip() if area else '', round(final_p / 10000) if final_p > 0 else 0)
                if dedup_key in seen_transactions:
                    continue
                seen_transactions.add(dedup_key)
                    
                if pname_str not in grouped:
                    grouped[pname_str] = {'region': reg_val, 'district': dist_val, 'count': 0, 'prices': [], 'sqfts': []}
                    
                grouped[pname_str]['count'] += 1
                if final_p > 0: grouped[pname_str]['prices'].append(final_p)
                if final_u > 0: grouped[pname_str]['sqfts'].append(final_u)

            ranked = sorted(grouped.items(), key=lambda x: x[1]['count'], reverse=True)[:limit]
            res = []
            for name, info in ranked:
                avg_p = (sum(info['prices'])/len(info['prices'])/10000) if info['prices'] else 0
                avg_u = (sum(info['sqfts'])/len(info['sqfts'])) if info['sqfts'] else 0
                res.append({
                    'project_name': name,
                    'region': info['region'],
                    'district': info['district'],
                    'volume': info['count'],
                    'avg_price_wan': round(avg_p, 1),
                    'avg_sqft': round(avg_u)
                })
            return res

        def get_category_bundle(where_prefix):
            return {
                'overall': query_rankings_in_memory(f"{where_prefix}", limit=10),
                'region_hk': query_rankings_in_memory(f"{where_prefix}", region_filter='港岛', limit=10),
                'region_kl': query_rankings_in_memory(f"{where_prefix}", region_filter='九龙', limit=10),
                'price_500_2000m': query_rankings_in_memory(f"{where_prefix}", price_min=5000000, price_max=20000000, limit=10),
                'price_2000_5000m': query_rankings_in_memory(f"{where_prefix}", price_min=20000000, price_max=50000000, limit=10),
                'price_5000_10000m': query_rankings_in_memory(f"{where_prefix}", price_min=50000000, price_max=100000000, limit=10),
                'price_10000m_above': query_rankings_in_memory(f"{where_prefix}", price_min=100000000, limit=10)
            }

        # 1. 提取可用年份
        c.execute("SELECT DISTINCT substr(sold_date, 1, 4) FROM sold_history WHERE sold_date IS NOT NULL AND sold_date != '' ORDER BY sold_date DESC LIMIT 5")
        years = [r[0] for r in c.fetchall() if r[0] and len(r[0]) == 4]
        
        years_list = []
        yearly_map = {}
        for y in years:
            label = f"{y}年度累计" if y == "2026" else f"{y}全年度"
            years_list.append({'val': y, 'label': label})
            yearly_map[y] = get_category_bundle(f"substr(sold_date, 1, 4) = '{y}'")

        # 2. 提取可用月份
        c.execute("SELECT DISTINCT substr(sold_date, 1, 7) FROM sold_history WHERE sold_date IS NOT NULL AND sold_date != '' ORDER BY sold_date DESC LIMIT 12")
        months = [r[0] for r in c.fetchall() if r[0] and len(r[0]) == 7]

        months_list = []
        monthly_map = {}
        for idx, m in enumerate(months):
            parts = m.split('-')
            m_label = f"{parts[0]}年{int(parts[1])}月" + (" (最新)" if idx == 0 else "")
            months_list.append({'val': m, 'label': m_label})
            monthly_map[m] = get_category_bundle(f"substr(sold_date, 1, 7) = '{m}'")

        # 3. 提取可用周度 (规范 ISO 周自然周区间)
        def get_week_start_end(year, week):
            first_day = datetime(year, 1, 4)
            first_monday = first_day - timedelta(days=first_day.weekday())
            week_monday = first_monday + timedelta(weeks=week - 1)
            week_sunday = week_monday + timedelta(days=6)
            return week_monday.strftime('%Y-%m-%d'), week_sunday.strftime('%Y-%m-%d')

        c.execute("SELECT sold_date FROM sold_history WHERE sold_date >= '2026-01-01' AND sold_date IS NOT NULL AND sold_date != '' ORDER BY sold_date DESC")
        date_rows = c.fetchall()
        week_map_raw = {}
        for r in date_rows:
            dstr = str(r[0]).strip()
            try:
                dt_parts = [int(p) for p in dstr[:10].split('-')]
                dt_obj = datetime(dt_parts[0], dt_parts[1], dt_parts[2])
                iso_yr, iso_wk, _ = dt_obj.isocalendar()
                w_key = f"{iso_yr}-W{iso_wk:02d}"
                if w_key not in week_map_raw:
                    start_d, end_d = get_week_start_end(iso_yr, iso_wk)
                    week_map_raw[w_key] = {'year': iso_yr, 'week_num': iso_wk, 'start_date': start_d, 'end_date': end_d}
            except:
                continue

        sorted_weeks = sorted(week_map_raw.keys(), reverse=True)[:10]
        weeks_list = []
        weekly_map = {}
        for idx, w in enumerate(sorted_weeks):
            w_info = week_map_raw[w]
            start_m_d = w_info['start_date'][5:].replace('-','/')
            end_m_d = w_info['end_date'][5:].replace('-','/')
            w_label = f"{w_info['year']}年第{w_info['week_num']}周 ({start_m_d}-{end_m_d})" + (" (最新)" if idx == 0 else "")
            weeks_list.append({'val': w, 'label': w_label})
            weekly_map[w] = get_category_bundle(f"sold_date BETWEEN '{w_info['start_date']}' AND '{w_info['end_date']}'")

        conn.close()

        return {
            'options': {
                'months': months_list,
                'weeks': weeks_list,
                'years': years_list
            },
            'monthly_map': monthly_map,
            'weekly_map': weekly_map,
            'yearly_map': yearly_map
        }
    except Exception as e:
        print(f"构建动态排行榜失败: {e}")
        return {}

def main():
    print("开始扫描价单目录并构建网页数据库...")
    ensure_dirs()
    hkp_status_map = fetch_hkp_status_map()
    custom_districts = load_custom_districts()
    
    # 提取所有磁盘项目名称集合 (100% 涵盖所有存有一手销控表的新盘项目，严防美联/置业二手盘混入)
    valid_new_project_names = set()
    for d in os.listdir(BASE_DIR):
        dir_path = os.path.join(BASE_DIR, d)
        if not os.path.isdir(dir_path) or d.startswith('.') or d in ['web', 'scratch']:
            continue
        parts = d.split('-')
        if len(parts) >= 3:
            pname = "-".join(parts[2:]).strip()
            valid_new_project_names.add(pname)

    config_path = os.path.join(BASE_DIR, "config_admin.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                _cfg = json.load(f)
                if 'projects_data' in _cfg and isinstance(_cfg['projects_data'], dict):
                    for k in _cfg['projects_data'].keys():
                        valid_new_project_names.add(k)
        except: pass

    real_history = load_real_history_analytics(custom_districts, valid_new_project_names)
    leaderboards = build_leaderboard_data(valid_new_project_names)
    
    projects_list = []
    
    # 统计汇总
    global_stats = {
        'total_projects': 0,
        'total_units': 0,
        'total_sold': 0,
        'total_sale': 0,
        'total_priced': 0,
        'total_stopped': 0,
        'total_pending': 0,
        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # 0. 读取 config_admin.json 配置 与 香港新房精选项目.xlsx
    admin_config = {}
    config_path = os.path.join(BASE_DIR, "config_admin.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                admin_config = json.load(f)
        except Exception as e:
            print(f"读取 config_admin.json 失败: {e}")
            
    focus_projects = admin_config.get('focus_projects', [])
    featured_by_price = admin_config.get('featured_by_price', {})
    projects_data = admin_config.get('projects_data', {})

    custom_dist_file = os.path.join(BASE_DIR, "custom_districts.json")
    custom_districts = {}
    if os.path.exists(custom_dist_file):
        try:
            with open(custom_dist_file, 'r', encoding='utf-8') as f:
                custom_districts = json.load(f)
            print(f"成功载入 {len(custom_districts)} 条用户自定义商圈规则。")
        except Exception as e:
            print(f"读取 custom_districts.json 失败: {e}")

    # 优先从 香港新房精选项目.xlsx 动态更新聚焦精选盘内容
    def load_featured_from_excel():
        excel_paths = [
            os.path.join(BASE_DIR, "聚焦盘精选盘", "香港新房精选项目.xlsx"),
            os.path.join(os.path.dirname(BASE_DIR), "聚焦盘精选盘", "香港新房精选项目.xlsx"),
            "/Users/nb/google/Antigravity/工作/运营/聚焦盘精选盘/香港新房精选项目.xlsx",
            os.path.join(BASE_DIR, "香港新房精选项目.xlsx")
        ]
        target_path = None
        for p in excel_paths:
            if os.path.exists(p):
                target_path = p
                break
                
        if not target_path:
            return None, None

        try:
            wb = openpyxl.load_workbook(target_path, data_only=True)
            sheet = wb.active

            f_by_price = {'1000-2000': [], '2000-5000': [], '5000-10000': [], '10000+': []}
            p_data = {}
            current_tier = '1000-2000'

            header_col3 = str(sheet.cell(2, 3).value or '').strip()
            is_10_col_mode = ('价位' in header_col3 or '1000' in str(sheet.cell(3, 3).value or ''))

            for r in range(3, sheet.max_row + 1):
                name = sheet.cell(r, 1).value
                if not name: continue
                name_str = str(name).strip()
                if '1️⃣' in name_str or '首套刚需类' in name_str or '投资配置类' in name_str:
                    current_tier = '1000-2000'; continue
                if '2️⃣' in name_str or '自用保值类' in name_str:
                    current_tier = '2000-5000'; continue
                if '3️⃣' in name_str or '豪宅购置类' in name_str:
                    current_tier = '5000-10000'; continue
                if '4️⃣' in name_str or '顶豪收藏类' in name_str:
                    current_tier = '10000+'; continue
                if name_str == '项目名': continue

                raw_tier = str(sheet.cell(r, 3).value or '').strip()
                c_layout = sheet.cell(r, 4).value
                c_total = sheet.cell(r, 5).value
                c_sqft = sheet.cell(r, 6).value
                c_rent = ''
                c_roi = ''
                c_reason = sheet.cell(r, 7).value
                c_mainland = sheet.cell(r, 8).value

                tier_key = current_tier
                if '1000-2000' in raw_tier or '1000万-2000' in raw_tier or '500万' in raw_tier:
                    tier_key = '1000-2000'
                elif '2000-5000' in raw_tier or '2000万-5000' in raw_tier:
                    tier_key = '2000-5000'
                elif '5000-1亿' in raw_tier or '5000万-1亿' in raw_tier or '5000-10000' in raw_tier:
                    tier_key = '5000-10000'
                elif '1亿' in raw_tier or '未开售' in raw_tier or '10000+' in raw_tier:
                    tier_key = '10000+'

                if name_str not in f_by_price[tier_key]:
                    f_by_price[tier_key].append(name_str)

                clean_name = re.sub(r'\(第.*?\)', '', name_str).replace('4B', '').replace('4b', '').strip()
                user_cd = (custom_districts.get(name_str) or 
                           custom_districts.get(clean_name) or 
                           custom_districts.get(name_str.lower()) or 
                           custom_districts.get(clean_name.lower()) or {})
                if not user_cd.get('district'):
                    for k, v in custom_districts.items():
                        if k.lower() in name_str.lower() or name_str.lower() in k.lower():
                            user_cd = v
                            break

                roi_val = ''
                if c_roi is not None:
                    try:
                        r_float = float(c_roi)
                        if 0 < r_float < 1:
                            roi_val = f"{round(r_float * 100, 2)}%"
                        else:
                            roi_val = str(c_roi).strip()
                    except:
                        roi_val = str(c_roi).strip()

                meta_item = {
                    'grade': str(sheet.cell(r, 2).value or 'A').strip(),
                    'region': user_cd.get('region') or '九龙',
                    'district': user_cd.get('district') or '',
                    'price_tier': tier_key,
                    'main_layout': str(c_layout or '').strip(),
                    'total_price': str(c_total or '').strip(),
                    'total_price_desc': str(c_total or '').strip(),
                    'sqft_price': str(c_sqft or '').strip(),
                    'sqft_price_desc': str(c_sqft or '').strip(),
                    'rent_range': str(c_rent or '').strip(),
                    'rent_range_desc': str(c_rent or '').strip(),
                    'roi': roi_val,
                    'reason': str(c_reason or '').strip(),
                    'mainland_selling_points': str(c_mainland or '').strip(),
                }
                p_data[f"{tier_key}_{name_str}"] = meta_item
                p_data[name_str] = meta_item
            print(f"成功从 Excel [{target_path}] 载入 {len(p_data)} 个精选聚焦盘源数据。")
            return f_by_price, p_data
        except Exception as e:
            print(f"警告: 从 Excel 读取精选项目失败: {e}")
            return None, None

    excel_f_by_price, excel_p_data = load_featured_from_excel()
    if excel_f_by_price and excel_p_data:
        featured_by_price = excel_f_by_price
        projects_data.update(excel_p_data)

    rental_bm_file = os.path.join(BASE_DIR, "rental_benchmarks.json")
    rental_benchmarks = {}
    if os.path.exists(rental_bm_file):
        try:
            with open(rental_bm_file, 'r', encoding='utf-8') as f:
                rental_benchmarks = json.load(f)
        except Exception as e:
            print(f"读取 rental_benchmarks.json 失败: {e}")

    custom_dist_file = os.path.join(BASE_DIR, "custom_districts.json")
    custom_districts = {}
    if os.path.exists(custom_dist_file):
        try:
            with open(custom_dist_file, 'r', encoding='utf-8') as f:
                custom_districts = json.load(f)
            print(f"成功载入 {len(custom_districts)} 条用户自定义商圈规则。")
            for pname, cd in custom_districts.items():
                if pname not in projects_data:
                    projects_data[pname] = {}
                if 'region' in cd: projects_data[pname]['region'] = cd['region']
                if 'district' in cd: projects_data[pname]['district'] = cd['district']
        except Exception as e:
            print(f"读取 custom_districts.json 失败: {e}")

    drive_mapping_file = os.path.join(BASE_DIR, "google_drive_mapping.json")
    drive_mapping = {}
    if os.path.exists(drive_mapping_file):
        try:
            with open(drive_mapping_file, 'r', encoding='utf-8') as f:
                drive_mapping = json.load(f)
            print(f"成功载入 {len(drive_mapping)} 条 Google Drive 显式映射规则。")
            # 为所有别名及不同写法补全 projects_data 映射
            for k_alias, f_info in drive_mapping.items():
                if k_alias not in projects_data:
                    projects_data[k_alias] = {}
                if isinstance(f_info, dict):
                    projects_data[k_alias]['google_drive_folder'] = f_info.get('folder', '')
                    projects_data[k_alias]['marketing_url'] = f_info.get('url', '#')
                elif isinstance(f_info, str):
                    if f_info.startswith('http'):
                        projects_data[k_alias]['marketing_url'] = f_info
                    else:
                        projects_data[k_alias]['google_drive_folder'] = f_info
                        projects_data[k_alias]['marketing_url'] = f"https://drive.google.com/drive/search?q={urllib.parse.quote('type:folder parent:' + PARENT_DRIVE_ID + ' \"' + f_info + '\"')}"

            def get_drive_info(key_str):
                if not key_str: return None
                k = str(key_str).strip()
                k_clean = re.sub(r'第\s*[0-9A-Za-z\-]+\s*期', '', k).strip()
                k_clean_cn = re.sub(r'第\s*[一二三四五六七八九十]+\s*期', '', k).strip()
                m_info = (drive_mapping.get(k) or 
                          drive_mapping.get(k_clean) or 
                          drive_mapping.get(k_clean_cn) or 
                          drive_mapping.get(k.lower()) or 
                          drive_mapping.get(k_clean.lower()) or 
                          drive_mapping.get(k_clean_cn.lower()))
                return m_info

            for pk, pval in projects_data.items():
                if isinstance(pval, dict):
                    raw_name = pk.split('_', 1)[-1] if '_' in pk else pk
                    m_info = get_drive_info(raw_name) or get_drive_info(pk)
                    if isinstance(m_info, dict):
                        if m_info.get('url'): pval['marketing_url'] = m_info['url']
                        if m_info.get('folder'): pval['google_drive_folder'] = m_info['folder']
                    elif isinstance(m_info, str) and m_info.startswith('http'):
                        pval['marketing_url'] = m_info
        except Exception as e:
            print(f"读取 google_drive_mapping.json 失败: {e}")

    def strip_phase_suffix(name):
        s = str(name).strip()
        s = re.sub(r'\(第.*?\)', '', s)
        s = re.sub(r'第\s*[0-9A-Za-z\-]+期.*$', '', s)
        s = re.sub(r'Phase\s*[0-9A-Za-z\-]+.*$', '', s, flags=re.IGNORECASE)
        s = re.sub(r'第[I|V|X|A-Z0-9]+期', '', s)
        s = re.sub(r'第[0-9A-Za-z]+$', '', s)
        s = re.sub(r'[0-9]+[a-zA-Z]+$', '', s)
        s = re.sub(r'\s+[0-9]+$', '', s)
        s = re.sub(r'\s+I{1,3}$', '', s)
        s = re.sub(r'\s+II$', '', s)
        s = re.sub(r'\s+III$', '', s)
        s = re.sub(r'\(.*?\)', '', s)
        return s.strip()

    # 载入 楼盘介绍映射表.xlsx
    intro_mapping = {}
    intro_excel = "/Users/nb/google/Antigravity/工作/运营/聚焦盘精选盘/楼盘介绍映射表.xlsx"
    if os.path.exists(intro_excel):
        try:
            wb_intro = openpyxl.load_workbook(intro_excel, data_only=True)
            ws_intro = wb_intro.active
            for r in range(2, ws_intro.max_row + 1):
                pname = str(ws_intro.cell(r, 1).value or '').strip()
                url = str(ws_intro.cell(r, 2).value or '').strip()
                if pname and url and url != 'None' and url.startswith('http'):
                    intro_mapping[pname] = url
                    clean_p = strip_phase_suffix(pname)
                    if clean_p and clean_p not in intro_mapping:
                        intro_mapping[clean_p] = url
            print(f"成功从 Excel [楼盘介绍映射表.xlsx] 载入 {len(intro_mapping)} 条楼盘介绍链接。")
        except Exception as e:
            print(f"读取 楼盘介绍映射表.xlsx 失败: {e}")

    # 载入 营销素材网盘地址.xlsx 状态 (包含“是否有资料”标识)
    marketing_materials_map = {}
    marketing_excel = "/Users/nb/google/Antigravity/工作/运营/聚焦盘精选盘/营销素材网盘地址.xlsx"
    if os.path.exists(marketing_excel):
        try:
            wb_mat = openpyxl.load_workbook(marketing_excel, data_only=True)
            ws_mat = wb_mat.active
            headers = [str(ws_mat.cell(1, c).value or '').strip() for c in range(1, ws_mat.max_column + 1)]
            
            folder_col = 2
            url_col = 3
            has_mat_col = None
            for idx, h in enumerate(headers, 1):
                if '文件夹' in h or '项目' in h: folder_col = idx
                elif '地址' in h or '链接' in h or 'url' in h.lower(): url_col = idx
                elif '是否有资料' in h or '有资料' in h or '资料' in h: has_mat_col = idx
            
            for r in range(2, ws_mat.max_row + 1):
                folder_val = str(ws_mat.cell(r, folder_col).value or '').strip()
                url_val = str(ws_mat.cell(r, url_col).value or '').strip()
                has_mat_val = str(ws_mat.cell(r, has_mat_col).value or '').strip() if has_mat_col else ''
                
                if folder_val:
                    pname = folder_val.split('-')[-1].strip()
                    if has_mat_col:
                        has_mat = (has_mat_val == '是')
                    else:
                        has_mat = bool(url_val and url_val.startswith('http'))
                    
                    info = {'url': url_val, 'has_materials': has_mat}
                    marketing_materials_map[folder_val] = info
                    marketing_materials_map[pname] = info
                    clean_p = strip_phase_suffix(pname)
                    if clean_p and clean_p not in marketing_materials_map:
                        marketing_materials_map[clean_p] = info
            print(f"成功从 Excel [营销素材网盘地址.xlsx] 解析 {len(marketing_materials_map)} 条营销资料状态。")
        except Exception as e:
            print(f"读取 营销素材网盘地址.xlsx 失败: {e}")

    # 遍历 BASE_DIR 下的子文件夹
    for d in sorted(os.listdir(BASE_DIR)):
        dir_path = os.path.join(BASE_DIR, d)
        if not os.path.isdir(dir_path):
            continue
            
        if d.startswith('.') or d in ['web', 'scratch']:
            continue
            
        parts = d.split('-')
        if len(parts) < 3:
            continue
            
        region = parts[0].strip()
        district = parts[1].strip()
        project_name = "-".join(parts[2:]).strip()
        
        EXCLUDE_DISTRICTS = {"将军澳", "茶果岭、油塘及鲤鱼门", "长沙湾", "牛头角及九龙湾", "慈云山、钻石山及新蒲岗"}
        if region == "九龙" and district in EXCLUDE_DISTRICTS:
            continue
        
        excel_filename = f"{project_name}_销控明细表.xlsx"
        src_excel_path = os.path.join(dir_path, excel_filename)
        
        if not os.path.exists(src_excel_path):
            found = False
            for file in os.listdir(dir_path):
                if file.endswith("_销控明细表.xlsx"):
                    src_excel_path = os.path.join(dir_path, file)
                    excel_filename = file
                    found = True
                    break
            if not found:
                continue
        
        stats = parse_project_stats(src_excel_path)
        
        if stats['total'] > 0 and stats['sold'] == stats['total']:
            continue
        
        dest_filename = f"{region}-{district}-{project_name}.xlsx"
        dest_excel_path = os.path.join(FILES_DIR, dest_filename)
        
        try:
            for d in [FILES_DIR, ROOT_FILES_DIR]:
                shutil.copy2(src_excel_path, os.path.join(d, dest_filename))
            file_size_kb = round(os.path.getsize(dest_excel_path) / 1024, 1)
        except Exception as e:
            continue
            
        mtime = os.path.getmtime(src_excel_path)
        last_updated_date = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d')
        
        hkp_st_info = hkp_status_map.get(project_name, {})
        sell_status = hkp_st_info.get('sell_status', 'on_sale')
        sell_status_cn = hkp_st_info.get('sell_status_cn', '出售中')

        # 🚨 严格过滤已售罄项目 (sold_out / 已售罄)
        if sell_status == 'sold_out' or sell_status_cn == '已售罄':
            continue
        
        custom_map = {
            # MONACO 全系合并
            'Monaco 第1期': ('九龙', '启德', 'MONACO'),
            'Monaco One': ('九龙', '启德', 'MONACO'),
            'Monaco Marine': ('九龙', '启德', 'MONACO'),
            'MONACO': ('九龙', '启德', 'MONACO'),
            'MONACO ONE': ('九龙', '启德', 'MONACO'),
            'MONACO MARINE': ('九龙', '启德', 'MONACO'),
            
            # 启德海湾全系合并
            '启德海湾': ('九龙', '启德', '启德海湾'),
            '启德海湾 1': ('九龙', '启德', '启德海湾'),
            '启德海湾 2': ('九龙', '启德', '启德海湾'),
            '启德海湾1期': ('九龙', '启德', '启德海湾'),
            '启德海湾2期': ('九龙', '启德', '启德海湾'),
            
            # 汇玺全系合并
            '汇玺III': ('九龙', '西南九龙', '汇玺'),
            '汇玺3期': ('九龙', '西南九龙', '汇玺'),
            '汇玺': ('九龙', '西南九龙', '汇玺'),
            
            # THE HENLEY 全系合并
            'Henley Park': ('九龙', '启德', 'THE HENLEY'),
            'The Henley I': ('九龙', '启德', 'THE HENLEY'),
            'The Henley II': ('九龙', '启德', 'THE HENLEY'),
            'The Henley III': ('九龙', '启德', 'THE HENLEY'),
            'THE HENLEY': ('九龙', '启德', 'THE HENLEY'),
            
            # Double Coast 全系合并
            'Double Coast I': ('九龙', '启德', 'Double Coast'),
            'Double Coast III': ('九龙', '启德', 'Double Coast'),

            'Mount Nicholson I': ('港岛', '山顶区', 'Mount Nicholson'),
            'Mount Nicholson II': ('港岛', '山顶区', 'Mount Nicholson'),
            'Blue Coast': ('港岛', '黄竹坑', 'Blue Coast'),
            'Blue Coast II': ('港岛', '黄竹坑', 'Blue Coast'),
            '滶晨': ('港岛', '黄竹坑', '滶晨'),
            '滶晨 II': ('港岛', '黄竹坑', '滶晨'),
            'Central Peak I': ('港岛', '半山区东部', 'Central Peak'),
            'Central Peak II': ('港岛', '半山区东部', 'Central Peak'),
            '维港汇 I': ('九龙', '西南九龙', '维港汇'),
            '维港汇 III': ('九龙', '西南九龙', '维港汇'),
            '海璇': ('港岛', '北角', '海璇'),
            '海璇 II': ('港岛', '北角', '海璇'),
            '海璇 II (第2B-3期)': ('港岛', '北角', '海璇'),
            '利奥坊．凯岸': ('九龙', '旺角', '利奥坊'),
            '利奥坊．壹隅': ('九龙', '旺角', '利奥坊'),
            '利奥坊．曦岸': ('九龙', '旺角', '利奥坊'),
            '必嘉坊．曦汇': ('九龙', '红磡', '必嘉坊'),
            '必嘉坊．迎汇': ('九龙', '红磡', '必嘉坊'),
            '瑜一': ('九龙', '何文田', '瑜一'),
            '瑜一第IC期': ('九龙', '何文田', '瑜一'),
            '瑜一．天海': ('九龙', '何文田', '瑜一'),
            'The Monet 第1期': ('九龙', '九龙塘', 'The Monet'),
            'The Monet 第2期': ('九龙', '九龙塘', 'The Monet'),
            'The Monet 第3期': ('九龙', '九龙塘', 'The Monet'),
            '朗贤峰第IIA期': ('九龙', '何文田', '朗贤峰'),
            '朗贤峰第IIB期': ('九龙', '何文田', '朗贤峰'),
            '海盈山第4A期': ('港岛', '黄竹坑', '海盈山'),
            '海盈山第4B期': ('港岛', '黄竹坑', '海盈山'),
            'Deep Water South 第6A期': ('港岛', '寿臣山及浅水湾', 'Deep Water South'),
            'Deep Water South 第6B期': ('港岛', '寿臣山及浅水湾', 'Deep Water South'),
            '天御第1期': ('港岛', '半山区西部', '天御'),
            '天御第2期': ('港岛', '半山区西部', '天御'),
            '壹沐第1期': ('九龙', '马头角', '壹沐'),
            '壹沐第2期': ('九龙', '马头角', '壹沐'),
            '首岸第1期': ('九龙', '红磡', '首岸'),
            '首岸第2期': ('九龙', '红磡', '首岸'),
            '首岸第3期': ('九龙', '红磡', '首岸'),
            '首岸第4期': ('九龙', '红磡', '首岸'),
        }

        def strip_phase_suffix(name):
            s = str(name).strip()
            s = re.sub(r'\(第.*?\)', '', s)
            s = re.sub(r'第\s*[0-9A-Za-z\-]+期.*$', '', s)
            s = re.sub(r'Phase\s*[0-9A-Za-z\-]+.*$', '', s, flags=re.IGNORECASE)
            s = re.sub(r'第[I|V|X|A-Z0-9]+期', '', s)
            s = re.sub(r'第[0-9A-Za-z]+$', '', s)
            s = re.sub(r'[0-9]+[a-zA-Z]+$', '', s)
            s = re.sub(r'\s+[0-9]+$', '', s)
            s = re.sub(r'\s+I{1,3}$', '', s)
            s = re.sub(r'\s+II$', '', s)
            s = re.sub(r'\s+III$', '', s)
            s = re.sub(r'\(.*?\)', '', s)
            return s.strip()

        p_meta = projects_data.get(project_name, {})
        clean_pname = strip_phase_suffix(project_name)
        reg_val = p_meta.get('region') or region
        dist_val = p_meta.get('district') or district
        user_custom = custom_districts.get(project_name) or custom_districts.get(clean_pname)
        if user_custom:
            reg_val = user_custom.get('region') or reg_val
            dist_val = user_custom.get('district') or dist_val
            if project_name in projects_data:
                projects_data[project_name]['region'] = reg_val
                projects_data[project_name]['district'] = dist_val

        PARENT_DRIVE_ID = "15tRwSlG1VTOKuEyj-H131zpNK6v6MY04"
        mapped_folder = drive_mapping.get(project_name) or drive_mapping.get(clean_pname)
        if isinstance(mapped_folder, dict):
            g_folder = mapped_folder.get('folder', '')
        elif isinstance(mapped_folder, str):
            g_folder = mapped_folder
        else:
            g_folder = f"{reg_val}-{dist_val}-{clean_pname}"

        g_folder_str = str(g_folder)
        drive_q = f'type:folder parent:{PARENT_DRIVE_ID} "{g_folder_str}"'
        g_url = f"https://drive.google.com/drive/search?q={urllib.parse.quote(drive_q)}"
        
        dist_info = rental_benchmarks.get('districts', {}).get(district, rental_benchmarks.get('default_fallback', {'base_rent': 50, 'min_rent': 42, 'max_rent': 60}))
        base_rent = dist_info.get('base_rent', 50)
        min_rent = dist_info.get('min_rent', 42)
        max_rent = dist_info.get('max_rent', 60)
        estimated_rent_desc = f"${min_rent} - ${max_rent}/呎"
        
        p_hist = real_history.get(project_name, {})
        m_dict = p_hist.get('monthly', {})
        p_meta = projects_data.get(project_name, {})
        tot_vol = sum(m.get('volume', 0) for m in m_dict.values())
        tot_val = sum(m.get('volume', 0) * m.get('avg_uprice', 0) for m in m_dict.values())
        avg_uprice = int(tot_val / tot_vol) if tot_vol > 0 else 0

        if avg_uprice > 0:
            calc_roi = round((base_rent * 12 / avg_uprice) * 100, 2)
            roi_str = f"{calc_roi}%"
        else:
            roi_str = p_meta.get('roi') or "3.5%"

        mat_info = marketing_materials_map.get(g_folder_str) or marketing_materials_map.get(project_name) or marketing_materials_map.get(clean_pname) or {}
        has_mat = mat_info.get('has_materials', False)
        m_url = mat_info.get('url') or p_meta.get('marketing_url') or p_meta.get('drive_url') or g_url

        tot_price_desc = p_meta.get('total_price', '')
        min_p_val = stats.get('min_price')
        if not tot_price_desc and min_p_val:
            wan_val = round(min_p_val / 10000.0, 2)
            wan_str = f"{wan_val:g}" if wan_val == int(wan_val) else f"{wan_val:.2f}"
            tot_price_desc = f"${wan_str}万起"

        proj_info = {
            'name': project_name,
            'region': p_meta.get('region') or region,
            'district': p_meta.get('district') or district,
            'filename': dest_filename,
            'file_size_kb': file_size_kb,
            'stats': stats,
            'last_updated': last_updated_date,
            'sell_status': sell_status,
            'sell_status_cn': sell_status_cn,
            'is_suspended': (sell_status == 'sales_suspended') or (stats.get('stopped', 0) > 0 and stats.get('sale', 0) == 0),
            'is_coming_soon': sell_status == 'coming_soon',
            'is_registration': sell_status == 'registration',
            'grade': p_meta.get('grade', 'C'),
            'basic_info': p_meta.get('basic_info', ''),
            'selling_points': p_meta.get('selling_points', ''),
            'mainland_selling_points': p_meta.get('mainland_selling_points', ''),
            'centaline_url': p_meta.get('centaline_url', ''),
            'intro_url': intro_mapping.get(project_name) or intro_mapping.get(clean_pname) or p_meta.get('intro_url', ''),
            'price_tier': p_meta.get('price_tier', ''),
            'google_drive_folder': g_folder,
            'marketing_url': m_url,
            'has_marketing_materials': has_mat,
            'main_layout': p_meta.get('main_layout', ''),
            'total_price_desc': tot_price_desc,
            'min_price': min_p_val,
            'sqft_price_desc': p_meta.get('sqft_price', ''),
            'rent_range_desc': p_meta.get('rent_range') or estimated_rent_desc,
            'roi': roi_str if avg_uprice > 0 else (p_meta.get('roi') or roi_str),
            'avg_uprice': avg_uprice,
            'base_sqft_rent': base_rent,
            'estimated_sqft_rent_desc': estimated_rent_desc,
            'reason': p_meta.get('reason', ''),
            'is_focus': project_name in focus_projects
        }
        projects_list.append(proj_info)
        
        global_stats['total_projects'] += 1
        global_stats['total_units'] += stats['total']
        global_stats['total_sold'] += stats['sold']
        global_stats['total_sale'] += stats['sale']
        global_stats['total_priced'] += stats['priced']
        global_stats['total_stopped'] += stats['stopped']
        global_stats['total_pending'] += stats['pending']

    # 追加补全 HKP API 中已存在但尚无具体单元销控 Excel 的项目 (如花语海第1期、花语海第2期等，仅限港岛与九龙)
    existing_names = {p['name'] for p in projects_list}
    EXCLUDE_DISTRICTS = {"将军澳", "茶果岭、油塘及鲤鱼门", "长沙湾", "牛头角及九龙湾", "慈云山、钻石山及新蒲岗"}
    for hkp_name, hkp_item in hkp_status_map.items():
        if hkp_name not in existing_names:
            p_meta = projects_data.get(hkp_name, {})
            clean_pname = strip_phase_suffix(hkp_name)
            reg_val = hkp_item.get('region') or p_meta.get('region') or ''
            dist_val = hkp_item.get('district') or p_meta.get('district') or ''
            user_custom = custom_districts.get(hkp_name) or custom_districts.get(clean_pname)
            if user_custom:
                reg_val = user_custom.get('region') or reg_val
                dist_val = user_custom.get('district') or dist_val

            # 🚨 严格执行区域规范：仅限【港岛】和【九龙】，严禁新界或离岛加入！
            if reg_val not in ("港岛", "九龙"):
                continue

            # 🚨 严格执行区域规范：排除九龙 5 个禁用商圈
            if reg_val == "九龙" and dist_val in EXCLUDE_DISTRICTS:
                continue

            st = hkp_item.get('sell_status', 'coming_soon')
            st_cn = hkp_item.get('sell_status_cn', '即將發售')

            # 🚨 严格过滤已售罄项目 (sold_out / 已售罄)
            if st == 'sold_out' or st_cn == '已售罄':
                continue

            mat_info = marketing_materials_map.get(hkp_name) or marketing_materials_map.get(clean_pname) or {}
            has_mat = mat_info.get('has_materials', False)
            m_url = mat_info.get('url') or p_meta.get('marketing_url') or f"https://drive.google.com/drive/search?q={urllib.parse.quote(hkp_name)}"

            no_excel_proj = {
                'name': hkp_name,
                'region': reg_val,
                'district': dist_val,
                'filename': '',
                'file_size_kb': 0,
                'has_excel': False,
                'stats': {
                    'total': hkp_item.get('total_unit', 0),
                    'sold': 0,
                    'sale': 0,
                    'priced': 0,
                    'stopped': 0,
                    'pending': hkp_item.get('total_unit', 0),
                    'sold_rate': 0.0
                },
                'last_updated': datetime.now().strftime('%Y-%m-%d'),
                'sell_status': st,
                'sell_status_cn': st_cn,
                'is_suspended': (st == 'sales_suspended'),
                'is_coming_soon': (st == 'coming_soon'),
                'is_registration': (st == 'registration'),
                'grade': p_meta.get('grade', 'B'),
                'basic_info': p_meta.get('basic_info', ''),
                'selling_points': p_meta.get('selling_points', ''),
                'mainland_selling_points': p_meta.get('mainland_selling_points', ''),
                'intro_url': intro_mapping.get(hkp_name) or intro_mapping.get(clean_pname) or '',
                'price_tier': p_meta.get('price_tier', ''),
                'marketing_url': m_url,
                'has_marketing_materials': has_mat,
                'main_layout': p_meta.get('main_layout', ''),
                'total_price_desc': p_meta.get('total_price', '售價待定'),
                'sqft_price_desc': p_meta.get('sqft_price', '呎價待定'),
                'rent_range_desc': p_meta.get('rent_range') or '租金估算中',
                'roi': p_meta.get('roi') or '3.5%',
                'avg_uprice': 0,
                'reason': p_meta.get('reason', '全新即將發售新盤'),
                'is_focus': hkp_name in focus_projects
            }
            projects_list.append(no_excel_proj)

    # 🚨 真实成交数据与官网实时动态数据双向补全管线（严格遵循：规划总套数 = 官网规划套数）
    for proj in projects_list:
        pname = proj['name']
        st = proj.get('stats', {})
        has_ex = proj.get('has_excel', True)
        old_sold = st.get('sold', 0)
        official_total = st.get('total', 0)
        
        # 1. 获取官网 HKP API 实时已售 (total_sold) 与在售 (total_sale, 包含10B单位)
        hkp_info = hkp_status_map.get(pname, {})
        hkp_sold = hkp_info.get('total_sold', 0)
        hkp_sale = hkp_info.get('total_sale', 0)

        # 2. 获取 3.18万+ SQLite 离线成交大库数据
        p_hist = real_history.get(pname, {})
        m_dict = p_hist.get('monthly', {})
        tx_sold = sum(v.get('volume', 0) for v in m_dict.values())
        
        if not has_ex:
            # 无物理 Excel 表的项目：优先取官网 API 实时已售 (total_sold) 与在售 (total_sale)
            real_sold = hkp_sold if hkp_sold > 0 else tx_sold
            st['sold'] = real_sold
            
            # 精准在售判定：若 API 明确返回 total_sale > 0 则用 total_sale；对于尾盘剩 1 套 (如 Elize Park 10B) 则归为 1 套在售；未推盘则归 0 在售
            if hkp_sale > 0:
                real_sale = hkp_sale
            elif official_total > 0 and real_sold > 0 and (official_total - real_sold == 1):
                real_sale = 1
            else:
                real_sale = 0

            st['sale'] = real_sale
            st['priced'] = real_sale
            st['pending'] = max(0, official_total - real_sold - real_sale)
        else:
            st['sold'] = max(old_sold, tx_sold)

        if official_total > 0:
            new_rate = round((st['sold'] / official_total) * 100, 1)
            if new_rate > 100.0:
                new_rate = 100.0
        else:
            new_rate = 0.0
            
        st['sold_rate'] = new_rate
        proj['stats'] = st

    # 重新汇总全站全局统计看板
    global_stats['total_projects'] = len(projects_list)
    global_stats['total_units'] = sum(p['stats']['total'] for p in projects_list)
    global_stats['total_sold'] = sum(p['stats']['sold'] for p in projects_list)
    global_stats['total_sale'] = sum(p['stats']['sale'] for p in projects_list)
    global_stats['total_priced'] = sum(p['stats']['priced'] for p in projects_list)
    global_stats['total_stopped'] = sum(p['stats']['stopped'] for p in projects_list)
    global_stats['total_pending'] = sum(p['stats']['pending'] for p in projects_list)
    if global_stats['total_units'] > 0:
        global_stats['overall_sold_rate'] = round((global_stats['total_sold'] / global_stats['total_units']) * 100, 1)
    else:
        global_stats['overall_sold_rate'] = 0.0

    version_hash = datetime.now().strftime('%Y%m%d%H%M%S')
    global_stats['version_hash'] = version_hash

    # 1. 核心轻量版 data.json (包含主页渲染、销售检索及热销排行榜必需数据，极速 ~340KB)
    db_data = {
        'version_hash': version_hash,
        'global_stats': global_stats,
        'projects': projects_list,
        'focus_projects': focus_projects,
        'featured_by_price': featured_by_price,
        'projects_data': projects_data,
        'leaderboards': leaderboards
    }
    
    # 2. 专项离线成交走势库 analytics_data.json (懒加载 ~1.2MB)
    analytics_data = {
        'version_hash': version_hash,
        'real_history_analytics': real_history,
        'leaderboards': leaderboards
    }

    # 同时写入 WEB_DIR 和 BASE_DIR 根目录
    for target_dir in [WEB_DIR, BASE_DIR]:
        j_path = os.path.join(target_dir, "data.json")
        with open(j_path, 'w', encoding='utf-8') as f:
            json.dump(db_data, f, ensure_ascii=False, separators=(',', ':'))

        s_path = os.path.join(target_dir, "data.js")
        with open(s_path, 'w', encoding='utf-8') as f:
            f.write("window.APP_DATA=" + json.dumps(db_data, ensure_ascii=False, separators=(',', ':')) + ";")

        aj_path = os.path.join(target_dir, "analytics_data.json")
        with open(aj_path, 'w', encoding='utf-8') as f:
            json.dump(analytics_data, f, ensure_ascii=False, separators=(',', ':'))

        as_path = os.path.join(target_dir, "analytics_data.js")
        with open(as_path, 'w', encoding='utf-8') as f:
            f.write("window.data_real_history=" + json.dumps(real_history, ensure_ascii=False, separators=(',', ':')) + ";\nwindow.data_leaderboards=" + json.dumps(leaderboards, ensure_ascii=False, separators=(',', ':')) + ";")
        
    # 检查是否有新增项目需要建 Google Drive 文件夹提醒
    known_folders_cache_file = os.path.join(BASE_DIR, "known_folders.json")
    known_folders = set()
    if os.path.exists(known_folders_cache_file):
        try:
            with open(known_folders_cache_file, 'r', encoding='utf-8') as f:
                known_folders = set(json.load(f))
        except:
            pass

    current_folders = set(str(p['google_drive_folder']) for p in projects_list if 'google_drive_folder' in p and isinstance(p['google_drive_folder'], str))
    new_folders = current_folders - known_folders

    if new_folders and len(known_folders) > 0:
        print("\n" + "="*70)
        print("🔔【新增盘源 Google Drive 建文件夹提醒】")
        for nf in sorted(new_folders):
            print(f" ➔ 发现新增盘源，请在 Google Drive 中新建对应的文件夹: [{nf}]")
        print("="*70 + "\n")

    # 保存最新已知文件夹缓存
    with open(known_folders_cache_file, 'w', encoding='utf-8') as f:
        json.dump(sorted(list(current_folders)), f, ensure_ascii=False, indent=2)

    print(f"\n构建成功! 共处理 {global_stats['total_projects']} 个项目，包含 {len(real_history)} 个项目的真实成交历史统计。")
    print(f"数据索引已同步写入根目录及 web 目录: data.json 及 data.js")

if __name__ == "__main__":
    main()
